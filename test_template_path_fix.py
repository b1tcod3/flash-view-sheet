"""
Test para verificar la correcci√≥n del problema de validaci√≥n de plantilla
"""
import os
import tempfile
import pandas as pd
import openpyxl
from PySide6.QtWidgets import QApplication
from app.widgets.export_separated_dialog import ExportSeparatedDialog

def test_template_path_stored():
    """Test que verifica que la ruta de plantilla se almacena correctamente"""
    print("üß™ Iniciando test de correcci√≥n de ruta de plantilla...")
    
    # Crear una aplicaci√≥n Qt
    app = QApplication.instance()
    if app is None:
        app = QApplication([])
    
    # Crear datos de prueba
    df = pd.DataFrame({
        'Region': ['Norte', 'Sur', 'Norte', 'Este'],
        'Ventas': [1000, 2000, 1500, 1800]
    })
    
    # Crear plantilla Excel temporal
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        template_path = temp_file.name
    
    try:
        # Crear archivo Excel de prueba
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Regi√≥n'
        ws['B1'] = 'Ventas'
        wb.save(template_path)
        wb.close()
        
        # Crear di√°logo
        dialog = ExportSeparatedDialog(df)
        
        # Simular selecci√≥n de plantilla
        dialog._template_path = template_path  # Esto es lo que agregamos en la correcci√≥n
        dialog.template_path_label.setText(f"üìÑ {os.path.basename(template_path)}")
        dialog.template_path_label.setToolTip(template_path)
        
        # Verificar que la ruta se almacena correctamente
        assert hasattr(dialog, '_template_path'), "El di√°logo debe tener el atributo _template_path"
        assert dialog._template_path == template_path, f"La ruta debe ser {template_path}"
        
        # Verificar que la configuraci√≥n puede obtener la ruta
        config = dialog.get_configuration(validate=False)
        assert config is not None, "La configuraci√≥n no debe ser None"
        assert config.template_path == template_path, "La plantilla en la configuraci√≥n debe coincidir"
        
        print("‚úÖ Test pasado: La ruta de plantilla se almacena y recupera correctamente")
        return True
        
    except Exception as e:
        print(f"‚ùå Test fall√≥: {str(e)}")
        return False
    finally:
        # Limpiar archivo temporal
        if os.path.exists(template_path):
            os.unlink(template_path)

def test_validation_with_stored_template():
    """Test que verifica la validaci√≥n con plantilla almacenada"""
    print("üß™ Iniciando test de validaci√≥n con plantilla almacenada...")
    
    # Crear una aplicaci√≥n Qt
    app = QApplication.instance()
    if app is None:
        app = QApplication([])
    
    # Crear datos de prueba
    df = pd.DataFrame({
        'Region': ['Norte', 'Sur'],
        'Ventas': [1000, 2000]
    })
    
    # Crear plantilla Excel temporal
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        template_path = temp_file.name
    
    try:
        # Crear archivo Excel de prueba
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Regi√≥n'
        ws['B1'] = 'Ventas'
        wb.save(template_path)
        wb.close()
        
        # Crear di√°logo y configurar
        dialog = ExportSeparatedDialog(df)
        dialog._template_path = template_path
        dialog.column_combo.setCurrentText('Region')
        dialog.dest_folder_label.setText("/tmp")
        dialog.filename_template_edit.setText("{valor}.xlsx")
        
        # Intentar validaci√≥n (esto no debe fallar ahora)
        try:
            validation_result = dialog.validate_configuration()
            print("‚úÖ Test pasado: La validaci√≥n funciona correctamente con plantilla almacenada")
            return True
        except Exception as e:
            print(f"‚ùå Test fall√≥ en validaci√≥n: {str(e)}")
            return False
        
    except Exception as e:
        print(f"‚ùå Test fall√≥: {str(e)}")
        return False
    finally:
        # Limpiar archivo temporal
        if os.path.exists(template_path):
            os.unlink(template_path)

def test_filename_extension_validation():
    """Test que verifica validaci√≥n de extensi√≥n de archivo de exportaci√≥n"""
    print("üß™ Iniciando test de validaci√≥n de extensi√≥n de archivo...")
    
    # Crear una aplicaci√≥n Qt
    app = QApplication.instance()
    if app is None:
        app = QApplication([])
    
    # Crear datos de prueba
    df = pd.DataFrame({
        'Region': ['Norte', 'Sur'],
        'Ventas': [1000, 2000]
    })
    
    # Crear plantilla Excel temporal
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        template_path = temp_file.name
    
    try:
        # Crear archivo Excel de prueba
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Regi√≥n'
        ws['B1'] = 'Ventas'
        wb.save(template_path)
        wb.close()
        
        # Crear di√°logo
        dialog = ExportSeparatedDialog(df)
        dialog._template_path = template_path
        dialog.column_combo.setCurrentText('Region')
        dialog.dest_folder_label.setText("/tmp")
        
        # Test 1: Plantilla con extensi√≥n v√°lida .xlsx
        dialog.filename_template_edit.setText("{valor}.xlsx")
        config = dialog.get_configuration(validate=False)
        assert config is not None, "Configuraci√≥n debe ser v√°lida con extensi√≥n .xlsx"
        print("‚úÖ Extensi√≥n .xlsx v√°lida - Correcta")
        
        # Test 2: Plantilla con extensi√≥n v√°lida .xlsm
        dialog.filename_template_edit.setText("{valor}.xlsm")
        config = dialog.get_configuration(validate=False)
        assert config is not None, "Configuraci√≥n debe ser v√°lida con extensi√≥n .xlsm"
        print("‚úÖ Extensi√≥n .xlsm v√°lida - Correcta")
        
        # Test 3: Plantilla con extensi√≥n inv√°lida
        dialog.filename_template_edit.setText("{valor}.doc")
        config = dialog.get_configuration(validate=False)
        assert config is None, "Configuraci√≥n debe ser None con extensi√≥n .doc inv√°lida"
        print("‚úÖ Extensi√≥n .doc inv√°lida - Correctamente rechazada")
        
        # Test 4: Plantilla sin extensi√≥n
        dialog.filename_template_edit.setText("{valor}")
        config = dialog.get_configuration(validate=False)
        assert config is None, "Configuraci√≥n debe ser None sin extensi√≥n"
        print("‚úÖ Sin extensi√≥n - Correctamente rechazada")
        
        # Test 5: Plantilla con extensi√≥n .XLSX (may√∫sculas)
        dialog.filename_template_edit.setText("{valor}.XLSX")
        config = dialog.get_configuration(validate=False)
        assert config is not None, "Configuraci√≥n debe ser v√°lida con extensi√≥n .XLSX (may√∫sculas)"
        print("‚úÖ Extensi√≥n .XLSX (may√∫sculas) v√°lida - Correcta")
        
        print("‚úÖ Todos los tests de validaci√≥n de extensi√≥n pasaron")
        return True
        
    except Exception as e:
        print(f"‚ùå Test de validaci√≥n de extensi√≥n fall√≥: {str(e)}")
        return False
    finally:
        # Limpiar archivo temporal
        if os.path.exists(template_path):
            os.unlink(template_path)

if __name__ == "__main__":
    print("üîç Test de Correcci√≥n: Validaci√≥n de Plantilla y Nombre de Archivo")
    print("=" * 60)
    
    test1_passed = test_template_path_stored()
    test2_passed = test_validation_with_stored_template()
    test3_passed = test_filename_extension_validation()
    
    print("\n" + "=" * 60)
    if test1_passed and test2_passed and test3_passed:
        print("üéâ Todos los tests pasaron. Las correcciones funcionan correctamente.")
    else:
        print("‚ö†Ô∏è  Algunos tests fallaron. Revisar la implementaci√≥n.")