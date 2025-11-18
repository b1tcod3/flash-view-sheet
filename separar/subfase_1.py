# Subfase 1.2: Análisis de Impacto en la Arquitectura
## Evaluación de Integración de la Nueva Funcionalidad

### 1. Análisis de la Arquitectura Actual

#### 1.1 Estructura de Menús y Navegación
**Estado Actual**:
```
Archivo
├── Abrir...
├── Exportar como...
│   ├── PDF...
│   ├── Imagen...
│   └── SQL...
└── Salir
```

**Impacto de la Nueva Funcionalidad**:
```
Archivo
├── Abrir...
├── Exportar como...
│   ├── PDF...
│   ├── Imagen...
│   └── SQL...
└── Salir

Separar          ← NUEVO MENÚ PRINCIPAL
├── Exportar Datos Separados...    ← NUEVA OPCIÓN
└── Configurar Plantillas...       ← OPCIONAL: gestión de plantillas
```

**Consideraciones**:
- **Posicionamiento**: El menú "Separar" debe estar al mismo nivel que "Archivo"
- **Accesibilidad**: Requiere datos cargados para habilitarse
- **Consistencia**: Seguir el mismo patrón de señal-slot del sistema actual
- **Estado**: Deshabilitado cuando no hay DataFrame cargado

#### 1.2 Patrones de Diálogo Existentes
**Análisis del diálogo LoadOptionsDialog**:

```python
class LoadOptionsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Opciones de Carga")
        self.resize(600, 400)
        # ... configuración inicial
        
    def setup_ui(self):
        """Configurar la interfaz del diálogo"""
        main_layout = QVBoxLayout(self)
        # ... creación de UI
        
    def get_options(self):
        """Obtener las opciones configuradas"""
        # ... lógica de obtención de datos
```

**Aplicación al Nuevo Diálogo**:
- **ExportSeparatedDialog** seguirá el mismo patrón
- Uso de QGroupBox para organizar secciones
- QDialogButtonBox para botones Ok/Cancel
- Métodos setup_ui() y get_configuration()

#### 1.3 Sistema de Exportación Actual
**Funciones Export en data_handler.py**:
```python
def exportar_a_pdf(df: pd.DataFrame, filepath: str) -> bool:
def exportar_a_sql(df: pd.DataFrame, filepath: str, nombre_tabla: str) -> bool:
def exportar_a_imagen(table_view, filepath: str) -> bool:
```

**Patrón Identificado**:
- Función toma DataFrame + parámetros específicos
- Retorna bool para indicar éxito/fallo
- Manejo de errores interno con try/except
- Logging de errores para debugging

**Nueva Función a Agregar**:
```python
def exportar_datos_separados(df: pd.DataFrame, config: dict) -> dict:
    """
    Exporta DataFrame separado por columna usando plantilla Excel
    
    Args:
        df: DataFrame a separar
        config: Diccionario con configuración:
            - 'columna_separacion': str
            - 'plantilla_excel': str (ruta)
            - 'hoja_excel': str
            - 'celda_inicial': str
            - 'mapeo_columnas': dict
            - 'plantilla_nombre': str
            - 'carpeta_destino': str
    
    Returns:
        dict: {
            'exito': bool,
            'archivos_generados': list,
            'errores': list,
            'estadisticas': dict
        }
    """
```

### 2. Componentes a Crear

#### 2.1 Nuevo Widget: ExportSeparatedDialog
**Ubicación**: `app/widgets/export_separated_dialog.py`

**Responsabilidades**:
- Interface de configuración principal
- Validación en tiempo real
- Preview de archivos a generar
- Conexión con lógica de exportación

**Integración**:
- Importado en main.py
- Instanciado cuando se selecciona menú "Separar"
- Recibe DataFrame actual como parámetro
- Emite señal con configuración cuando se acepta

#### 2.2 Extensión de data_handler.py
**Nueva Clase**: `ExcelTemplateSplitter`

**Métodos Principales**:
```python
class ExcelTemplateSplitter:
    def __init__(self, dataframe, config):
        self.df = dataframe
        self.config = config
        
    def validar_configuracion(self) -> list:
        """Validar todos los parámetros de configuración"""
        
    def generar_preview_archivos(self) -> list:
        """Generar preview de archivos que se crearán"""
        
    def procesar_separacion(self, callback_progreso=None) -> dict:
        """Procesar la separación completa"""
        
    def crear_archivo_individual(self, grupo_datos, nombre_archivo) -> bool:
        """Crear un archivo Excel individual con plantilla"""
```

#### 2.3 Integración con Menú Principal
**Modificaciones en main.py**:

1. **Agregar menú "Separar"**:
```python
# En create_menu_bar()
separar_menu = menu_bar.addMenu("&Separar")

# Acción Exportar Datos Separados
separar_action = separar_menu.addAction("&Exportar Datos Separados...")
separar_action.setShortcut("Ctrl+E")
separar_action.triggered.connect(self.exportar_datos_separados)
separar_action.setEnabled(False)  # Solo habilitado con datos
```

2. **Nuevo método**:
```python
def exportar_datos_separados(self):
    """Slot para exportar datos separados"""
    if self.df_vista_actual is None:
        QMessageBox.warning(self, "Advertencia", "No hay datos para exportar.")
        return
        
    from app.widgets.export_separated_dialog import ExportSeparatedDialog
    dialog = ExportSeparatedDialog(self.df_vista_actual, self)
    
    if dialog.exec():
        config = dialog.get_configuration()
        self.procesar_exportacion_separada(config)
```

### 3. Compatibilidad con Sistema de Loaders

#### 3.1 Integración con Loader Existente
**Análisis del Sistema Actual**:
```python
# En data_handler.py
from core.loaders import get_file_loader
loader = get_file_loader(filepath)
df = loader.load()
```

**Compatibilidad con Nueva Funcionalidad**:
- La separación trabaja con DataFrames ya cargados
- No interfiere con el sistema de loaders
- Aprovecha el DataFrame ya procesado por loaders
- Puede beneficiarse de optimizaciones de memoria existentes

#### 3.2 Optimizaciones para Datasets Grandes
**Compatibilidad con Chunking**:
- Reutilizar lógica de chunking del sistema de loaders
- Aplicar chunking durante la separación si es necesario
- Mantener compatibilidad con system de virtualización existente

**Configuración de Memoria**:
- Aprovechar `optimization_config` existente
- Usar misma lógica de threshold para datasets grandes
- Mantener consistencia con sistema de memory management

### 4. Investigación de Librerías Excel

#### 4.1 Evaluación: openpyxl vs Alternativas

**Openpyxl (Recomendado)**:
```python
# Ventajas identificadas
- Preserva formato Excel completamente
- Soporte para estilos, fórmulas, gráficos
- Lectura y escritura eficiente
- Mantiene objetos Excel complejos
- Activamente mantenido

# Implementación
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment

# Preservar formato
wb = load_workbook(template_path)
ws = wb.active
# Insertar datos preservando estilos
for row_idx, row_data in enumerate(data_rows, start_row):
    for col_idx, value in enumerate(row_data, start_col):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        # Formato se mantiene automáticamente
wb.save(output_path)
```

**Alternativas Evaluadas**:

1. **XlsxWriter**:
   - ❌ Solo escritura (no lee plantillas)
   - ❌ No preserva formato existente
   - ✅ Bueno para crear desde cero
   - ❌ No apto para funcionalidad de plantillas

2. **Pandas ExcelWriter**:
   - ❌ Modifica formato de plantillas
   - ❌ Opciones limitadas de preservación
   - ✅ Interface simple
   - ❌ No adecuado para preservación de formato

3. **Xlrd/Xlwt** (Legacy):
   - ❌ Solo formatos .xls antiguos
   - ❌ No soporte para .xlsx moderno
   - ❌ Funcionalidad limitada
   - ❌ Proyecto abandonado

**Conclusión**: **Openpyxl es la única opción viable** para preservar completamente el formato de plantillas Excel mientras se insertan datos.

#### 4.2 Implementación con Openpyxl

**Estrategia de Preservación de Formato**:
```python
def crear_archivo_con_plantilla(template_path, data, output_path, 
                               start_cell, column_mapping):
    """
    Crear archivo Excel copiando plantilla y insertando datos
    """
    # 1. Cargar plantilla preservando formato
    wb = load_workbook(template_path, data_only=False)
    ws = wb.active
    
    # 2. Determinar posición inicial
    start_row, start_col = parse_excel_cell(start_cell)
    
    # 3. Insertar datos preservando formato
    for row_offset, (_, row_data) in enumerate(data.iterrows()):
        for col_offset, (df_col, excel_col) in enumerate(column_mapping.items()):
            # Convertir letra de columna a número
            excel_col_num = excel_column_to_number(excel_col)
            cell = ws.cell(row=start_row + row_offset, 
                          column=excel_col_num)
            cell.value = row_data[df_col]
            # Formato de plantilla se mantiene
    
    # 4. Guardar preservando todo el formato
    wb.save(output_path)
```

### 5. Consideraciones de Rendimiento

#### 5.1 Optimizaciones de Memoria
**Compatibilidad con Sistema Actual**:
- Reutilizar `optimization_config` existente
- Aplicar mismos thresholds para chunking
- Mantener consistencia con virtualización de DataView

**Nuevas Optimizaciones**:
```python
# En ExcelTemplateSplitter
def procesar_en_chunks(self, chunk_size=1000):
    """Procesar separación por chunks para datasets grandes"""
    grupos_unicos = self.df[self.config['columna_separacion']].unique()
    
    for i in range(0, len(grupos_unicos), chunk_size):
        chunk_grupos = grupos_unicos[i:i+chunk_size]
        chunk_df = self.df[self.df[self.config['columna_separacion']].isin(chunk_grupos)]
        
        # Procesar chunk
        self._procesar_grupo_chunk(chunk_df)
        
        # Limpiar memoria
        del chunk_df
        gc.collect()
```

#### 5.2 Optimizaciones de I/O
**Estrategias Identificadas**:
1. **Cache de Plantillas**: Cargar plantilla una vez por cada grupo único
2. **Buffer de Escritura**: Agrupar escrituras para reducir I/O
3. **Compresión Temporal**: Usar archivos temporales para operaciones grandes

### 6. Impacto en la UI/UX

#### 6.1 Integración con Sistema de Vistas Actual
**Compatibilidad con StackedWidget**:
- La funcionalidad de separación es modal (diálogo)
- No interfiere con las vistas existentes
- Puede acceder al DataFrame actual desde cualquier vista

**Estado del Menú**:
- Menú "Separar" se habilita solo con datos cargados
- Iconografía consistente con resto de la aplicación
- Tooltips explicativos para nuevos usuarios

#### 6.2 Consistencia de Diseño
**Seguir Patrones Existentes**:
- Mismos estilos de botones y layouts
- Consistent color scheme (#4a90e2 primary)
- Mismos patrones de validación y feedback
- Misma estructura de GroupBox y FormLayout

### 7. Gestión de Errores y Fallbacks

#### 7.1 Integración con Sistema de Logging
**Compatibilidad con Logging Existente**:
```python
# En data_handler.py - patrón existente
except Exception as e:
    print(f"Error al exportar a PDF: {str(e)}")
    return False

# Aplicar mismo patrón para nueva funcionalidad
except Exception as e:
    print(f"Error en separación con plantillas: {str(e)}")
    logging.error(f"Error separación: {str(e)}")
    return {'exito': False, 'error': str(e)}
```

#### 7.2 Manejo de Casos Extremos
**Casos Identificados y Estrategias**:
1. **Plantilla Corrupta**: Validación previa, mensaje claro
2. **Sin Permisos**: Verificación anticipada, alternativa temporal
3. **Disco Lleno**: Monitoreo de espacio, cleanup automático
4. **Memoria Insuficiente**: Chunking automático, fallback a procesamiento reducido

### 8. Plan de Migración e Integración

#### 8.1 Fases de Implementación
1. **Fase A**: Crear ExcelTemplateSplitter en data_handler.py
2. **Fase B**: Crear ExportSeparatedDialog widget
3. **Fase C**: Integrar menú en main.py
4. **Fase D**: Testing e integración completa
5. **Fase E**: Optimizaciones de rendimiento

#### 8.2 Backward Compatibility
**Garantías**:
- No modificar funciones de exportación existentes
- No cambiar interfaces de widgets existentes
- Mantener compatibilidad con DataFrames existentes
- Preservar configuración de optimización actual

#### 8.3 Testing Strategy
**Tests de Integración**:
- Verificar que menús existentes siguen funcionando
- Validar que carga de archivos no se ve afectada
- Confirmar que rendimiento general se mantiene
- Testing con datasets de diferentes tamaños

### 9. Dependencias Nuevas

#### 9.1 Librerías Python
**Agregar a requirements.txt**:
```txt
openpyxl>=3.1.0  # Para manejo de Excel con preservación de formato
# Dependencias existentes se mantienen
pandas>=1.5.0
PySide6>=6.0.0
```

#### 9.2 Recursos del Sistema
**Consideraciones**:
- Espacio temporal para archivos intermedios
- Memoria para cache de plantillas
- Permisos de escritura en carpetas de destino
- Compatibilidad con diferentes sistemas de archivos

### 10. Métricas de Éxito de Integración

#### 10.1 Métricas Técnicas
- **Tiempo de Startup**: < 5 segundos (actual: ~3 segundos)
- **Memoria Base**: < 200MB RAM sin datos (actual: ~150MB)
- **Tiempo de Respuesta UI**: < 100ms para habilitar/deshabilitar menú
- **Compatibilidad**: 100% con funcionalidades existentes

#### 10.2 Métricas de Calidad
- **Cobertura de Tests**: > 90% para nueva funcionalidad
- **Regresiones**: 0 regresiones en funcionalidades existentes
- **Estabilidad**: < 1% crashes durante uso normal
- **Usabilidad**: Tiempo de configuración < 2 minutos

---

**Estado**: ✅ COMPLETADO - Subfase 1.2
**Próximo**: Subfase 1.3 - Diseño de Interfaz de Usuario
**Fecha**: 2025-11-04