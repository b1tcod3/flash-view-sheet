#!/usr/bin/env python3
"""
Script de prueba para validar preservaci√≥n de formato Excel
Ejecuta test completo para verificar que la soluci√≥n funciona
"""

import sys
import os
import tempfile
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment

# A√±adir directorios al path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from core.data_handler import ExcelTemplateSplitter, ExportSeparatedConfig
from core.excel_format_preserver import ExcelFormatPreserver


def create_test_template_with_format():
    """Crear plantilla de prueba con formato complejo"""
    template_path = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    
    # 1. T√≠tulo con formato
    title = ws['A1']
    title.value = "REPORTE DE VENTAS"
    title.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')
    
    # 2. Headers con formato
    headers = ['Regi√≥n', 'Ventas', 'Meta']
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i)
        cell.value = header
        cell.font = Font(name='Calibri', size=12, bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.border = Border(
            left=openpyxl.styles.Side(border_style='thin', color='000000'),
            right=openpyxl.styles.Side(border_style='thin', color='000000'),
            top=openpyxl.styles.Side(border_style='thin', color='000000'),
            bottom=openpyxl.styles.Side(border_style='thin', color='000000')
        )
    
    # 3. Formato de n√∫mero
    num_cell = ws['B4']
    num_cell.value = 999999
    num_cell.number_format = '#,##0.00'
    
    # 4. Anchos personalizados
    ws.column_dimensions['A'].width = 20.0
    ws.column_dimensions['B'].width = 15.0
    ws.column_dimensions['C'].width = 12.0
    
    # 5. Datos de ejemplo en fila 5
    ws['A5'] = "Norte"
    ws['B5'] = 5000
    ws['C5'] = 6000
    
    # 6. Celda con formato especial
    special = ws['D2']
    special.value = "FORMATO ESPECIAL"
    special.font = Font(italic=True, color='FF0000', size=10)
    
    wb.save(template_path.name)
    print(f"‚úì Plantilla de prueba creada: {template_path.name}")
    
    return template_path.name


def test_format_preservation():
    """Test principal de preservaci√≥n de formato"""
    print("\n=== INICIANDO TEST DE PRESERVACI√ìN DE FORMATO ===")
    
    # 1. Crear plantilla con formato
    template_path = create_test_template_with_format()
    
    # 2. Crear DataFrame de prueba
    df = pd.DataFrame({
        'Region': ['Norte', 'Sur', 'Este', 'Oeste'],
        'Ventas': [1000, 800, 900, 750],
        'Meta': [1200, 750, 950, 800]
    })
    
    # 3. Crear configuraci√≥n
    output_folder = tempfile.mkdtemp()
    config = ExportSeparatedConfig(
        separator_column='Region',
        template_path=template_path,
        output_folder=output_folder,
        file_template='Reporte_{valor}.xlsx',
        column_mapping={
            'Region': 'A',
            'Ventas': 'B',
            'Meta': 'C'
        },
        start_cell='A5'
    )
    
    # 4. Crear splitter y procesar
    try:
        print("‚úì Creando ExcelTemplateSplitter...")
        splitter = ExcelTemplateSplitter(df, config)
        
        print("‚úì Ejecutando separaci√≥n y exportaci√≥n...")
        result = splitter.separate_and_export()
        
        if result['success']:
            print("‚úì Separaci√≥n exitosa!")
            print(f"‚úì Archivos creados: {len(result['files_created'])}")
            
            # 5. Verificar que se preserv√≥ el formato
            for file_path in result['files_created']:
                print(f"\n--- Verificando formato en: {os.path.basename(file_path)}")
                verify_format_preservation(file_path, template_path)
            
        else:
            print("‚ùå Error en la separaci√≥n:")
            for error in result.get('errors', []):
                print(f"  - {error}")
            
    except Exception as e:
        print(f"‚ùå Error durante el test: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        # Limpiar archivos
        try:
            os.unlink(template_path)
            import shutil
            shutil.rmtree(output_folder)
        except:
            pass


def verify_format_preservation(output_path: str, original_template: str):
    """Verificar que el formato se preserv√≥ correctamente"""
    try:
        # Cargar archivo original para comparar
        original_wb = openpyxl.load_workbook(original_template, data_only=False)
        original_sheet = original_wb.active
        
        # Cargar archivo de salida
        output_wb = openpyxl.load_workbook(output_path, data_only=False)
        output_sheet = output_wb.active
        
        # Verificar t√≠tulo
        print("  üìã Verificando t√≠tulo...")
        original_title = original_sheet['A1']
        output_title = output_sheet['A1']
        
        title_checks = [
            ("valor", original_title.value, output_title.value),
            ("font.name", original_title.font.name, output_title.font.name),
            ("font.size", original_title.font.size, output_title.font.size),
            ("font.bold", original_title.font.bold, output_title.font.bold),
            ("fill.color", original_title.fill.start_color.rgb, output_title.fill.start_color.rgb)
        ]
        
        for check_name, original_val, output_val in title_checks:
            if original_val == output_val:
                print(f"    ‚úì {check_name}: {output_val}")
            else:
                print(f"    ‚ùå {check_name}: esperado {original_val}, obtenido {output_val}")
        
        # Verificar headers
        print("  üìã Verificando headers...")
        for col in ['A', 'B', 'C']:
            original_cell = original_sheet[f'{col}3']
            output_cell = output_sheet[f'{col}3']
            
            format_checks = [
                ("valor", original_cell.value, output_cell.value),
                ("font.name", original_cell.font.name, output_cell.font.name),
                ("font.size", original_cell.font.size, output_cell.font.size),
                ("font.bold", original_cell.font.bold, output_cell.font.bold),
                ("fill.color", original_cell.fill.start_color.rgb, output_cell.fill.start_color.rgb),
                ("border", original_cell.border.left.style is not None, output_cell.border.left.style is not None)
            ]
            
            for check_name, original_val, output_val in format_checks:
                if original_val == output_val:
                    print(f"    ‚úì {col}3 {check_name}: {output_val}")
                else:
                    print(f"    ‚ùå {col}3 {check_name}: esperado {original_val}, obtenido {output_val}")
        
        # Verificar anchos de columna
        print("  üìã Verificando anchos de columna...")
        for col in ['A', 'B', 'C']:
            original_width = original_sheet.column_dimensions[col].width
            output_width = output_sheet.column_dimensions[col].width
            
            if original_width == output_width:
                print(f"    ‚úì {col} width: {output_width}")
            else:
                print(f"    ‚ùå {col} width: esperado {original_width}, obtenido {output_width}")
        
        # Verificar celda con formato especial
        print("  üìã Verificando celda especial...")
        original_special = original_sheet['D2']
        output_special = output_sheet['D2']
        
        special_checks = [
            ("valor", original_special.value, output_special.value),
            ("font.italic", original_special.font.italic, output_special.font.italic),
            ("font.color", original_special.font.color.rgb, output_special.font.color.rgb)
        ]
        
        for check_name, original_val, output_val in special_checks:
            if original_val == output_val:
                print(f"    ‚úì {check_name}: {output_val}")
            else:
                print(f"    ‚ùå {check_name}: esperado {original_val}, obtenido {output_val}")
        
        original_wb.close()
        output_wb.close()
        
        print("  ‚úÖ Verificaci√≥n de formato completada")
        
    except Exception as e:
        print(f"  ‚ùå Error verificando formato: {e}")
        import traceback
        traceback.print_exc()


def test_excel_format_preserver_directly():
    """Test directo del ExcelFormatPreserver"""
    print("\n=== TEST DIRECTO DE EXCELFORMATPRESERVER ===")
    
    try:
        # Crear preserver
        preserver = ExcelFormatPreserver()
        print("‚úì ExcelFormatPreserver creado")
        
        # Crear plantilla de prueba
        template_path = create_test_template_with_format()
        
        # Cargar y cachear formatos
        wb = openpyxl.load_workbook(template_path, data_only=False)
        sheet = wb.active
        
        print("‚úì Caching formatos...")
        formats_cache = preserver.cache_workbook_formats(wb)
        print(f"  - Cached {len(formats_cache.get('worksheet_formats', {}))} worksheets")
        print(f"  - Cached {len(formats_cache.get('cell_formats', {}))} cell formats")
        
        # Simular inserci√≥n de datos
        test_data = {0: {'Region': 'Test', 'Ventas': 1234, 'Meta': 5678}}
        
        print("‚úì Insertando datos...")
        preserver.insert_data_preserving_format(
            sheet, 
            test_data, 
            {'Region': 'A', 'Ventas': 'B', 'Meta': 'C'},
            'A5'
        )
        
        # Restaurar formatos
        print("‚úì Restaurando formatos...")
        preserver.restore_workbook_formats(wb, formats_cache)
        
        # Guardar archivo de prueba
        test_output = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(test_output.name)
        wb.close()
        
        print("‚úì Verificando resultado...")
        verify_format_preservation(test_output.name, template_path)
        
        # Limpiar
        os.unlink(template_path)
        os.unlink(test_output.name)
        
    except Exception as e:
        print(f"‚ùå Error en test directo: {e}")
        import traceback
        traceback.print_exc()


def main():
    """Funci√≥n principal"""
    print("üîç DIAGN√ìSTICO Y PRUEBA DE PRESERVACI√ìN DE FORMATO EXCEL")
    print("=" * 60)
    
    # Test 1: Test directo del preserver
    test_excel_format_preserver_directly()
    
    # Test 2: Test completo con separaci√≥n
    test_format_preservation()
    
    print("\n" + "=" * 60)
    print("‚úÖ TESTS COMPLETADOS")
    print("\nRESUMEN:")
    print("- Se ha implementado ExcelFormatPreserver para preservar formato")
    print("- Se ha modificado _create_excel_file_with_template para usar preservaci√≥n")
    print("- Los tests verifican que el formato se mantiene durante la inserci√≥n de datos")


if __name__ == '__main__':
    main()