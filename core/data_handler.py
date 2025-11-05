"""
Módulo para manejo de datos - carga, análisis, transformación y exportación
Incluye integración con el sistema avanzado de transformaciones (Fase 7)
"""

import pandas as pd
import numpy as np
import os
from typing import Optional, Tuple, Dict, Any
import sys

# Añadir directorio raíz para importar config
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import optimization_config


def cargar_datos(filepath: str, chunk_size: int = None) -> pd.DataFrame:
    """
    Cargar datos desde un archivo usando el nuevo sistema de loaders

    Args:
        filepath: Ruta del archivo a cargar
        chunk_size: Tamaño de chunk para lectura (si el formato lo soporta)

    Returns:
        DataFrame de Pandas con los datos cargados

    Raises:
        FileNotFoundError: Si el archivo no existe
        ValueError: Si el archivo no es soportado
        Exception: Para otros errores de carga
    """
    from core.loaders import get_file_loader

    # Usar el factory pattern para cargar el archivo
    loader = get_file_loader(filepath)
    
    # Aplicar optimización para archivos grandes
    if chunk_size or (loader.can_load_chunks() and loader.get_memory_usage_info().get('file_size_mb', 0) > 100):
        if chunk_size is None:
            # Usar configuración de optimización
            file_info = loader.get_memory_usage_info()
            estimated_rows = file_info.get('estimated_data_rows', 1000)
            if estimated_rows > optimization_config.VIRTUALIZATION_THRESHOLD:
                chunk_size = 1000
            else:
                chunk_size = 10000
        
        try:
            return loader.load_in_chunks(chunk_size)
        except Exception as e:
            # Si falla el chunk loading, usar carga normal
            print(f"Chunk loading falló, usando carga normal: {str(e)}")
            return loader.load()

    return loader.load()


def cargar_datos_con_opciones(filepath: str, skip_rows: int = 0, column_names: dict = None, chunk_size: int = None) -> pd.DataFrame:
    """
    Cargar datos desde un archivo con opciones adicionales usando el sistema de loaders

    Args:
        filepath: Ruta del archivo a cargar
        skip_rows: Número de filas a saltar al inicio (la siguiente fila se usa como header)
        column_names: Diccionario con nombres de columnas a renombrar {original: nuevo}
        chunk_size: Tamaño de chunk para lectura (si el formato lo soporta)

    Returns:
        DataFrame de Pandas con los datos cargados y opciones aplicadas
    """
    from core.loaders import get_file_loader

    # Usar el factory pattern para cargar el archivo
    loader = get_file_loader(filepath)
    
    # Aplicar optimización para archivos grandes
    if chunk_size or (loader.can_load_chunks() and loader.get_memory_usage_info().get('file_size_mb', 0) > 100):
        if chunk_size is None:
            # Usar configuración de optimización
            file_info = loader.get_memory_usage_info()
            estimated_rows = file_info.get('estimated_data_rows', 1000)
            if estimated_rows > optimization_config.VIRTUALIZATION_THRESHOLD:
                chunk_size = 1000
            else:
                chunk_size = 10000
        
        try:
            df = loader.load_in_chunks(chunk_size)
        except Exception as e:
            # Si falla el chunk loading, usar carga normal
            print(f"Chunk loading falló, usando carga normal: {str(e)}")
            df = loader.load(skip_rows, column_names)
    else:
        # Carga normal con opciones
        df = loader.load(skip_rows, column_names)

    return df


def get_supported_file_formats() -> list:
    """
    Get list of all supported file formats
    
    Returns:
        List of supported file extensions
    """
    from core.loaders import get_supported_formats
    return get_supported_formats()


def is_file_format_supported(filepath: str) -> bool:
    """
    Check if a file format is supported
    
    Args:
        filepath: Path to the file
        
    Returns:
        True if the file format is supported
    """
    from core.loaders import is_file_supported
    return is_file_supported(filepath)


def obtener_metadata(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Obtener metadata del DataFrame
    
    Args:
        df: DataFrame de Pandas
        
    Returns:
        Diccionario con información sobre los datos
    """
    metadata = {
        'filas': df.shape[0],
        'columnas': df.shape[1],
        'nombres_columnas': df.columns.tolist(),
        'tipos_datos': df.dtypes.astype(str).tolist(),
        'columnas_numericas': df.select_dtypes(include=['number']).columns.tolist(),
        'columnas_texto': df.select_dtypes(include=['object']).columns.tolist(),
        'valores_nulos': df.isnull().sum().to_dict()
    }
    
    return metadata


def obtener_estadisticas(df: pd.DataFrame, columnas: list = None, percentiles: list = None) -> pd.DataFrame:
    """
    Obtener estadísticas descriptivas del DataFrame con optimización para datasets grandes

    Args:
        df: DataFrame de Pandas
        columnas: Lista de columnas específicas (si None, usar todas las numéricas)
        percentiles: Lista de percentiles a calcular (si None, usar [25, 50, 75])

    Returns:
        DataFrame con estadísticas descriptivas
    """
    try:
        if len(df) == 0:
            return pd.DataFrame()

        # Si no se especifican columnas, usar solo las numéricas para optimizar
        if columnas is None:
            columnas = df.select_dtypes(include=['number']).columns.tolist()
            if not columnas:
                # Si no hay columnas numéricas, usar todas
                columnas = df.columns.tolist()

        # Configurar percentiles por defecto (convertir a decimales)
        if percentiles is None:
            percentiles = [0.25, 0.5, 0.75]  # Usar decimales en lugar de porcentajes
        else:
            # Convertir percentiles de porcentaje a decimal si es necesario
            percentiles = [p/100 if p > 1 else p for p in percentiles]

        # Para datasets muy grandes, usar sample para estadísticas aproximadas
        if optimization_config.should_sample_stats(len(df)):
            print(f"Dataset grande detectado ({len(df)} filas), calculando estadísticas con sample...")
            sample_size = min(optimization_config.STATS_SAMPLE_SIZE, len(df))
            df_sample = df.sample(n=sample_size, random_state=42)
        else:
            df_sample = df

        # Calcular estadísticas solo para las columnas especificadas
        estadisticas = df_sample[columnas].describe(percentiles=percentiles, include='all')

        return estadisticas
    except Exception as e:
        print(f"Error al calcular estadísticas: {str(e)}")
        # Si hay error, devolver DataFrame vacío
        return pd.DataFrame()


def obtener_estadisticas_basicas(df: pd.DataFrame) -> dict:
    """
    Obtener estadísticas básicas optimizadas para datasets grandes

    Args:
        df: DataFrame de Pandas

    Returns:
        Diccionario con estadísticas básicas
    """
    try:
        basic_stats = {
            'total_filas': len(df),
            'total_columnas': len(df.columns),
            'columnas_numericas': len(df.select_dtypes(include=['number']).columns),
            'columnas_texto': len(df.select_dtypes(include=['object']).columns),
            'columnas_fecha': len(df.select_dtypes(include=['datetime']).columns),
            'memoria_uso_mb': df.memory_usage(deep=True).sum() / 1024 / 1024,
            'filas_duplicadas': df.duplicated().sum(),
            'valores_nulos_total': df.isnull().sum().sum()
        }

        return basic_stats
    except Exception as e:
        print(f"Error al calcular estadísticas básicas: {str(e)}")
        return {}


def aplicar_filtro(df: pd.DataFrame, columna: str, termino: str, use_index: bool = True) -> pd.DataFrame:
    """
    Aplicar filtro a los datos con optimización para datasets grandes

    Args:
        df: DataFrame original
        columna: Nombre de la columna a filtrar
        termino: Término de búsqueda
        use_index: Usar búsqueda indexada para mejor rendimiento

    Returns:
        DataFrame filtrado
    """
    if columna not in df.columns:
        raise ValueError(f"Columna no encontrada: {columna}")

    if len(df) == 0:
        return df

    # Para datasets muy grandes, usar optimizaciones
    if optimization_config.should_optimize_filtering(len(df)) and use_index:
        return _aplicar_filtro_indexado(df, columna, termino)
    else:
        return _aplicar_filtro_simple(df, columna, termino)


def _aplicar_filtro_simple(df: pd.DataFrame, columna: str, termino: str) -> pd.DataFrame:
    """
    Aplicar filtro simple (método original)

    Args:
        df: DataFrame original
        columna: Nombre de la columna a filtrar
        termino: Término de búsqueda

    Returns:
        DataFrame filtrado
    """
    # Filtrar por contenido de texto (case-insensitive)
    df_filtrado = df[df[columna].astype(str).str.contains(termino, case=False, na=False)]
    return df_filtrado


def _aplicar_filtro_indexado(df: pd.DataFrame, columna: str, termino: str) -> pd.DataFrame:
    """
    Aplicar filtro optimizado usando indexación para datasets grandes

    Args:
        df: DataFrame original
        columna: Nombre de la columna a filtrar
        termino: Término de búsqueda

    Returns:
        DataFrame filtrado
    """
    try:
        # Convertir columna a string para búsqueda de texto
        columna_str = df[columna].astype(str)

        # Crear una serie booleana para el filtro
        if termino.startswith('^') and termino.endswith('$'):
            # Búsqueda exacta (regex)
            pattern = termino[1:-1]
            mask = columna_str.str.match(pattern, case=False, na=False)
        elif termino.startswith('*') or termino.endswith('*'):
            # Búsqueda con wildcards
            pattern = termino.replace('*', '.*')
            mask = columna_str.str.contains(pattern, case=False, na=False, regex=True)
        else:
            # Búsqueda normal
            mask = columna_str.str.contains(termino, case=False, na=False)

        # Aplicar filtro
        df_filtrado = df[mask]

        print(f"Filtro aplicado: {len(df_filtrado)} de {len(df)} filas encontradas")
        return df_filtrado

    except Exception as e:
        print(f"Error en filtrado indexado, usando método simple: {str(e)}")
        return _aplicar_filtro_simple(df, columna, termino)


def exportar_a_pdf(df: pd.DataFrame, filepath: str) -> bool:
    """
    Exportar DataFrame a archivo PDF
    
    Args:
        df: DataFrame a exportar
        filepath: Ruta de destino
        
    Returns:
        True si la exportación fue exitosa
    """
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.lib import colors
        
        # Crear documento
        doc = SimpleDocTemplate(filepath, pagesize=letter)
        
        # Convertir DataFrame a lista de listas
        data = [df.columns.tolist()] + df.values.tolist()
        
        # Crear tabla
        tabla = Table(data)
        estilo = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla.setStyle(estilo)
        
        # Construir documento
        doc.build([tabla])
        return True
        
    except Exception as e:
        print(f"Error al exportar a PDF: {str(e)}")
        return False


def exportar_a_sql(df: pd.DataFrame, filepath: str, nombre_tabla: str) -> bool:
    """
    Exportar DataFrame a base de datos SQL

    Args:
        df: DataFrame a exportar
        filepath: Ruta de la base de datos
        nombre_tabla: Nombre de la tabla en la base de datos

    Returns:
        True si la exportación fue exitosa
    """
    try:
        from sqlalchemy import create_engine

        # Crear engine para SQLite
        engine = create_engine(f'sqlite:///{filepath}')

        # Exportar a SQL
        df.to_sql(nombre_tabla, engine, if_exists='replace', index=False)
        return True

    except Exception as e:
        print(f"Error al exportar a SQL: {str(e)}")
        return False


def exportar_a_imagen(table_view, filepath: str) -> bool:
    """
    Exportar vista de tabla a imagen

    Args:
        table_view: QTableView a capturar
        filepath: Ruta de destino para la imagen

    Returns:
        True si la exportación fue exitosa
    """
    try:
        from PySide6.QtWidgets import QApplication
        from PySide6.QtGui import QPixmap

        # Capturar la vista de la tabla como pixmap
        pixmap = table_view.grab()

        # Guardar como imagen (PNG por defecto)
        success = pixmap.save(filepath)

        if success:
            return True
        else:
            print(f"Error al guardar la imagen en {filepath}")
            return False

    except Exception as e:
        print(f"Error al exportar a imagen: {str(e)}")
        return False


def limpiar_datos(df: pd.DataFrame, opciones: dict = None) -> pd.DataFrame:
    """
    Limpiar datos del DataFrame aplicando varias operaciones de limpieza

    Args:
        df: DataFrame original
        opciones: Diccionario con opciones de limpieza
            - eliminar_duplicados: bool (default True)
            - eliminar_nulos: bool (default False)
            - rellenar_nulos: dict con columna: valor para rellenar
            - eliminar_columnas: list de columnas a eliminar
            - convertir_tipos: dict con columna: tipo

    Returns:
        DataFrame limpio
    """
    df_clean = df.copy()

    if opciones is None:
        opciones = {}

    # Eliminar duplicados
    if opciones.get('eliminar_duplicados', True):
        df_clean = df_clean.drop_duplicates()

    # Eliminar filas con nulos si se especifica
    if opciones.get('eliminar_nulos', False):
        df_clean = df_clean.dropna()

    # Rellenar nulos con valores especificados
    rellenar_nulos = opciones.get('rellenar_nulos', {})
    for columna, valor in rellenar_nulos.items():
        if columna in df_clean.columns:
            df_clean[columna] = df_clean[columna].fillna(valor)

    # Eliminar columnas especificadas
    eliminar_columnas = opciones.get('eliminar_columnas', [])
    for columna in eliminar_columnas:
        if columna in df_clean.columns:
            df_clean = df_clean.drop(columns=[columna])

    # Convertir tipos de datos
    convertir_tipos = opciones.get('convertir_tipos', {})
    for columna, tipo in convertir_tipos.items():
        if columna in df_clean.columns:
            try:
                if tipo == 'numeric':
                    df_clean[columna] = pd.to_numeric(df_clean[columna], errors='coerce')
                elif tipo == 'datetime':
                    df_clean[columna] = pd.to_datetime(df_clean[columna], errors='coerce')
                elif tipo == 'string':
                    df_clean[columna] = df_clean[columna].astype(str)
            except Exception as e:
                print(f"Error al convertir tipo de {columna}: {str(e)}")

    return df_clean


def agregar_datos(df: pd.DataFrame, operaciones: list) -> pd.DataFrame:
    """
    Realizar operaciones de agregación en el DataFrame

    Args:
        df: DataFrame original
        operaciones: Lista de diccionarios con operaciones
            Cada operación: {
                'grupo': ['col1', 'col2'],  # Columnas para agrupar
                'funciones': {'col3': 'sum', 'col4': ['mean', 'count']},
                'nombre': 'resultado'  # Nombre para el resultado
            }

    Returns:
        DataFrame con resultados de agregación
    """
    resultados = []

    for op in operaciones:
        grupo = op.get('grupo', [])
        funciones = op.get('funciones', {})
        nombre = op.get('nombre', 'agregado')

        try:
            if grupo:
                # Agregación por grupos
                df_agregado = df.groupby(grupo).agg(funciones).reset_index()
            else:
                # Agregación global
                df_agregado = df.agg(funciones).to_frame().T.reset_index(drop=True)

            # Aplanar columnas si hay múltiples funciones
            df_agregado.columns = ['_'.join(col).strip() if isinstance(col, tuple) else col
                                   for col in df_agregado.columns]

            df_agregado['operacion'] = nombre
            resultados.append(df_agregado)

        except Exception as e:
            print(f"Error en agregación '{nombre}': {str(e)}")

    if resultados:
        return pd.concat(resultados, ignore_index=True)
    else:
        return pd.DataFrame()


def pivotar_datos(df: pd.DataFrame, index: str, columns: str, values: str,
                  aggfunc: str = 'mean') -> pd.DataFrame:
    """
    Crear tabla pivote del DataFrame

    Args:
        df: DataFrame original
        index: Columna para usar como índice
        columns: Columna para usar como columnas
        values: Columna para usar como valores
        aggfunc: Función de agregación (mean, sum, count, etc.)

    Returns:
        DataFrame pivoteado
    """
    try:
        if aggfunc == 'mean':
            func = np.mean
        elif aggfunc == 'sum':
            func = np.sum
        elif aggfunc == 'count':
            func = 'count'
        else:
            func = aggfunc

        df_pivot = df.pivot_table(index=index, columns=columns, values=values, aggfunc=func)
        return df_pivot.reset_index()
    except Exception as e:
        print(f"Error al crear tabla pivote: {str(e)}")
        return pd.DataFrame()


# ===============================
# SISTEMA AVANZADO DE TRANSFORMACIONES (FASE 7)
# ===============================

def aplicar_transformacion(df: pd.DataFrame, tipo_transformacion: str, parametros: Dict[str, Any] = None) -> pd.DataFrame:
    """
    Aplicar una transformación avanzada al DataFrame usando el nuevo sistema
    
    Args:
        df: DataFrame de entrada
        tipo_transformacion: Tipo de transformación a aplicar
        parametros: Parámetros específicos para la transformación
        
    Returns:
        DataFrame transformado
        
    Raises:
        ValueError: Si el tipo de transformación no es válido
    """
    try:
        from core.transformations import get_transformation_manager
        
        if parametros is None:
            parametros = {}
            
        # Usar el gestor de transformaciones
        manager = get_transformation_manager()
        resultado = manager.execute_transformation(df, tipo_transformacion, parametros)
        
        print(f"Transformación '{tipo_transformacion}' aplicada exitosamente")
        return resultado
        
    except ImportError:
        print("Sistema de transformaciones avanzadas no disponible, usando método básico")
        return _aplicar_transformacion_basica(df, tipo_transformacion, parametros)
    except Exception as e:
        print(f"Error al aplicar transformación avanzada: {str(e)}")
        # Fallback al sistema básico si está disponible
        return _aplicar_transformacion_basica(df, tipo_transformacion, parametros)


def aplicar_pipeline_transformaciones(df: pd.DataFrame, pasos_transformacion: list) -> pd.DataFrame:
    """
    Aplicar un pipeline de transformaciones al DataFrame
    
    Args:
        df: DataFrame de entrada
        pasos_transformacion: Lista de pasos de transformación
            Cada paso: {'tipo': 'transformacion', 'parametros': {...}}
            
    Returns:
        DataFrame transformado
    """
    try:
        from core.transformations import get_transformation_manager
        
        if not pasos_transformacion:
            return df.copy()
            
        # Usar el gestor de transformaciones
        manager = get_transformation_manager()
        resultado = manager.execute_pipeline(df, pasos_transformacion)
        
        print(f"Pipeline de {len(pasos_transformacion)} pasos aplicado exitosamente")
        return resultado
        
    except ImportError:
        print("Sistema de transformaciones avanzadas no disponible")
        return df.copy()
    except Exception as e:
        print(f"Error al aplicar pipeline de transformaciones: {str(e)}")
        return df.copy()


def obtener_transformaciones_disponibles() -> Dict[str, Dict[str, Any]]:
    """
    Obtener lista de transformaciones disponibles en el sistema avanzado
    
    Returns:
        Diccionario con transformaciones disponibles
    """
    try:
        from core.transformations import get_transformation_manager
        
        manager = get_transformation_manager()
        return manager.get_available_transformations()
        
    except ImportError:
        print("Sistema de transformaciones avanzadas no disponible")
        return {}
    except Exception as e:
        print(f"Error al obtener transformaciones: {str(e)}")
        return {}


def ejecutar_transformacion_con_compatibilidad(df: pd.DataFrame, operacion: str, parametros: Dict[str, Any] = None) -> pd.DataFrame:
    """
    Ejecutar transformación con compatibilidad entre sistema básico y avanzado
    
    Args:
        df: DataFrame de entrada
        operacion: Tipo de operación a realizar
        parametros: Parámetros para la operación
        
    Returns:
        DataFrame procesado
    """
    if parametros is None:
        parametros = {}
    
    # Mapeo de operaciones básicas a transformaciones avanzadas
    mapeo_transformaciones = {
        'limpiar_datos': 'text_cleaning',  # Mapear a limpieza de texto básica
        'renombrar_columnas': 'rename_columns',
        'crear_columna_calculada': 'create_calculated_column',
        'aplicar_funcion': 'apply_function',
        'eliminar_columnas': 'drop_columns',
        'normalizar_datos': 'normalization',
        'escalar_datos': 'scaling',
        'aplicar_logaritmo': 'logarithmic',
        'convertir_texto': 'case_conversion',
        'parsear_fechas': 'date_parsing'
    }
    
    # Si la operación tiene mapeo directo, usar el sistema avanzado
    if operacion in mapeo_transformaciones:
        try:
            tipo_avanzado = mapeo_transformaciones[operacion]
            return aplicar_transformacion(df, tipo_avanzado, parametros)
        except Exception as e:
            print(f"Error en transformación avanzada, usando sistema básico: {str(e)}")
    
    # Si no hay mapeo, usar el sistema básico
    if operacion == 'limpiar_datos':
        return limpiar_datos(df, parametros)
    elif operacion == 'agregar_datos':
        return agregar_datos(df, parametros)
    elif operacion == 'pivotar_datos':
        index = parametros.get('index', '')
        columns = parametros.get('columns', '')
        values = parametros.get('values', '')
        aggfunc = parametros.get('aggfunc', 'mean')
        return pivotar_datos(df, index, columns, values, aggfunc)
    else:
        print(f"Operación '{operacion}' no reconocida, devolviendo DataFrame original")
        return df.copy()


def _aplicar_transformacion_basica(df: pd.DataFrame, tipo: str, parametros: Dict[str, Any] = None) -> pd.DataFrame:
    """
    Aplicar transformación básica usando métodos existentes (fallback)
    
    Args:
        df: DataFrame de entrada
        tipo: Tipo de transformación
        parametros: Parámetros
        
    Returns:
        DataFrame transformado
    """
    if parametros is None:
        parametros = {}
    
    # Implementar transformaciones básicas como fallback
    if tipo == 'text_cleaning':
        # Limpieza básica de texto
        df_clean = df.copy()
        for col in df_clean.select_dtypes(include=['object']).columns:
            df_clean[col] = df_clean[col].astype(str).str.strip()
        return df_clean
    
    elif tipo == 'rename_columns':
        mapping = parametros.get('column_mapping', {})
        if mapping:
            return df.rename(columns=mapping)
        return df
    
    elif tipo == 'drop_columns':
        columns = parametros.get('columns', [])
        if columns:
            return df.drop(columns=columns, errors='ignore')
        return df
    
    else:
        print(f"Transformación básica '{tipo}' no implementada, devolviendo DataFrame original")
        return df.copy()


def obtener_estadisticas_transformaciones() -> Dict[str, Any]:
    """
    Obtener estadísticas del sistema de transformaciones
    
    Returns:
        Diccionario con estadísticas
    """
    try:
        from core.transformations import get_transformation_manager
        
        manager = get_transformation_manager()
        return {
            'sistema_disponible': True,
            'transformaciones_registradas': len(manager.get_available_transformations()),
            'pipelines_guardados': len(manager.get_saved_pipelines()),
            'operaciones_en_historial': len(manager.get_history()),
            'metricas_rendimiento': manager.get_performance_report()
        }
        
    except ImportError:
        return {
            'sistema_disponible': False,
            'mensaje': 'Sistema de transformaciones avanzadas no disponible'
        }
    except Exception as e:
        return {
            'sistema_disponible': False,
            'error': str(e)
        }


# ===============================
# SISTEMA DE EXPORTACIÓN SEPARADA CON PLANTILLAS EXCEL (FASE 3)
# ===============================

import pandas as pd
import numpy as np
import os
import tempfile
import time
import json
import shutil
import hashlib
import re
import gc
from datetime import datetime
from typing import Optional, List, Dict, Any, Union, Tuple, Iterator, Callable
from dataclasses import dataclass, field
from pathlib import Path
from collections import namedtuple, defaultdict
from enum import Enum

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
    from openpyxl.styles import Font, PatternFill, Border, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Importar optimizaciones de rendimiento
try:
    from core.performance_optimizer import (
        PerformanceOptimizer, ExcelFormatOptimizer, ProgressMonitor,
        ChunkingStrategy, SystemResources, PerformanceConfig,
        PerformanceResult, ProgressInfo
    )
    PERFORMANCE_OPTIMIZER_AVAILABLE = True
except ImportError:
    PERFORMANCE_OPTIMIZER_AVAILABLE = False
    # Clases fallback si no está disponible el optimizador
    class ChunkingStrategy(Enum):
        NONE = "none"
        MODERATE = "moderate"
        SIZE_BASED = "size"
        GROUP_BASED = "group"
        AGGRESSIVE = "aggressive"
    
    class SystemResources:
        def __init__(self):
            self.memory_usage_mb = 0.0
            self.cpu_percent = 0.0
            self.disk_free_gb = 0.0
    
    class PerformanceConfig:
        def __init__(self):
            self.memory_threshold_mb = 2048
            self.chunk_size = 1000
            self.max_concurrent_operations = 2
            self.progress_interval = 10


@dataclass
class ExportSeparatedConfig:
    """Configuración completa para separación de datos"""
    
    # Datos de origen - Argumentos requeridos primero
    separator_column: str  # Columna para separar
    template_path: str  # Ruta a plantilla .xlsx
    output_folder: str  # Carpeta destino
    
    # Argumentos opcionales con valores por defecto
    start_cell: str = "A1"  # Celda inicial para datos
    file_template: str = "{valor}.xlsx"  # Plantilla nombre archivo
    
    # Mapeo de columnas
    column_mapping: Dict[str, str] = field(default_factory=dict)
    # Ej: {'columna_df': 'A', 'otra_columna': 'C'}
    
    # Opciones avanzadas
    handle_duplicates: str = "overwrite"  # 'overwrite', 'append', 'skip'
    create_summary: bool = True  # Crear archivo resumen
    preserve_format: bool = True  # Preservar formato Excel
    
    # Opciones de rendimiento
    enable_chunking: bool = True  # Habilitar chunking automático
    max_memory_mb: int = 2048  # Límite de memoria
    progress_callback: Optional[callable] = None  # Callback de progreso
    
    def validate(self) -> Dict[str, Any]:
        """Validar configuración completa"""
        errors = []
        warnings = []
        
        # Validar columna de separación
        if not self.separator_column:
            errors.append("Columna de separación es requerida")
        
        # Validar plantilla Excel
        if not os.path.exists(self.template_path):
            errors.append(f"Plantilla no encontrada: {self.template_path}")
        elif not self.template_path.lower().endswith(('.xlsx', '.xlsm')):
            errors.append("Plantilla debe ser archivo .xlsx o .xlsm")
        
        # Validar carpeta destino
        try:
            os.makedirs(self.output_folder, exist_ok=True)
            if not os.access(self.output_folder, os.W_OK):
                errors.append(f"Sin permisos de escritura en: {self.output_folder}")
        except Exception as e:
            errors.append(f"Error en carpeta destino: {str(e)}")
        
        # Validar celda inicial
        try:
            coordinate_to_tuple(self.start_cell)
        except Exception:
            errors.append(f"Celda inicial inválida: {self.start_cell}")
        
        # Validar plantilla de nombre
        if not self.file_template:
            errors.append("Plantilla de nombre de archivo es requerida")
        elif not self.file_template.endswith('.xlsx'):
            warnings.append("Plantilla de nombre no tiene extensión .xlsx")
        
        return {
            'valid': len(errors) == 0,
            'errors': errors,
            'warnings': warnings
        }
    
    def get_default_mapping(self, df_columns: List[str]) -> Dict[str, str]:
        """Obtener mapeo por defecto (posicional)"""
        mapping = {}
        for i, col in enumerate(df_columns):
            if i < 26:  # A-Z
                mapping[col] = get_column_letter(i + 1)
            else:  # AA, AB, etc.
                mapping[col] = get_column_letter(i + 1)
        return mapping


@dataclass
class ValidationResult:
    """Resultado de validación"""
    is_valid: bool = True
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    info: List[str] = field(default_factory=list)
    
    def add_error(self, error: str):
        self.errors.append(error)
        self.is_valid = False
    
    def add_warning(self, warning: str):
        self.warnings.append(warning)
    
    def add_info(self, info: str):
        self.info.append(info)


@dataclass
class ExportResult:
    """Resultado de exportación individual"""
    success: bool = False
    file_path: str = ""
    group_name: str = ""
    rows_processed: int = 0
    processing_time: float = 0.0
    error: str = ""
    timestamp: Optional[datetime] = None


class SeparationError(Exception):
    """Error base para separación de datos"""
    def __init__(self, message: str, error_code: str = None, details: dict = None):
        super().__init__(message)
        self.error_code = error_code
        self.details = details or {}


class TemplateError(SeparationError):
    """Error específico de plantilla Excel"""
    pass


class ConfigurationError(SeparationError):
    """Error de configuración inválida"""
    pass


class MemoryError(SeparationError):
    """Error de memoria insuficiente"""
    pass


class ExcelTemplateSplitter:
    """Clase principal para separación de datos con plantillas Excel"""
    
    def __init__(self, df: pd.DataFrame, config: ExportSeparatedConfig):
        """
        Inicializar separador con DataFrame y configuración
        
        Args:
            df: DataFrame a separar
            config: Configuración de separación
        """
        self.df = df
        self.config = config
        self.progress_callback = config.progress_callback
        self.logger = self._setup_logger()
        self.created_files = []
        self.failed_groups = {}
        self._cancelled = False
        
        # Verificar dependencias
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl es requerido para separación con plantillas Excel")
        
        # Configurar optimizaciones de rendimiento
        self._setup_performance_optimization()
    
    def _setup_logger(self):
        """Configurar logging consistente con sistema existente"""
        import logging
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)
        return logger
    
    def _setup_performance_optimization(self):
        """Configurar optimizaciones de rendimiento avanzadas"""
        # Inicializar optimizador de rendimiento si está disponible
        if PERFORMANCE_OPTIMIZER_AVAILABLE and self.config.enable_chunking:
            self.performance_optimizer = PerformanceOptimizer(
                memory_threshold_mb=self.config.max_memory_mb
            )
            
            # Determinar estrategia de chunking óptima
            self.chunking_strategy = self.performance_optimizer.determine_optimal_chunking_strategy(
                self.df, self.config.separator_column
            )
            
            # Configurar optimizador de formato Excel
            self.excel_optimizer = ExcelFormatOptimizer()
            
            # Configurar monitor de progreso
            self.progress_monitor = ProgressMonitor()
            
            self.using_advanced_optimization = True
        else:
            # Fallback a optimización básica
            self.performance_optimizer = None
            self.excel_optimizer = None
            self.progress_monitor = None
            self.using_advanced_optimization = False
            
            # Usar optimización_config para determinar chunking básico
            if (len(self.df) > optimization_config.VIRTUALIZATION_THRESHOLD or
                self.df.memory_usage(deep=True).sum() > self.config.max_memory_mb * 1024 * 1024):
                self.enable_chunking = True
                self.chunk_size = min(
                    optimization_config.DEFAULT_CHUNK_SIZE,
                    max(1000, len(self.df) // 20)
                )
            else:
                self.enable_chunking = False
                self.chunk_size = len(self.df)
    
    def validate_configuration(self) -> ValidationResult:
        """Validar configuración completa antes de proceder"""
        result = ValidationResult()
        
        # Validación básica de configuración
        config_validation = self.config.validate()
        if not config_validation['valid']:
            for error in config_validation['errors']:
                result.add_error(error)
            for warning in config_validation['warnings']:
                result.add_warning(warning)
        
        # Validación de datos
        if self.df.empty:
            result.add_error("DataFrame está vacío")
            return result
        
        if self.config.separator_column not in self.df.columns:
            result.add_error(f"Columna '{self.config.separator_column}' no existe en DataFrame")
            return result
        
        # Validar valores únicos en columna de separación
        unique_values = self.df[self.config.separator_column].nunique()
        null_count = self.df[self.config.separator_column].isnull().sum()
        
        if unique_values == 0:
            result.add_error("No se encontraron valores únicos en columna de separación")
        elif unique_values + (1 if null_count > 0 else 0) == 0:
            result.add_error("No hay datos para separar")
        
        # Validación de plantilla Excel
        if os.path.exists(self.config.template_path):
            try:
                workbook = load_workbook(self.config.template_path, read_only=True)
                workbook.close()
                result.add_info(f"Plantilla validada: {os.path.basename(self.config.template_path)}")
            except Exception as e:
                result.add_error(f"Error validando plantilla Excel: {str(e)}")
        
        return result
    
    def analyze_data(self) -> Dict[str, Any]:
        """Analizar DataFrame para generar preview y validar separación"""
        analysis = {}
        
        # Análisis básico
        analysis['total_rows'] = len(self.df)
        analysis['total_columns'] = len(self.df.columns)
        analysis['memory_usage_mb'] = self.df.memory_usage(deep=True).sum() / 1024 / 1024
        
        # Análisis de columna de separación
        if self.config.separator_column in self.df.columns:
            separator_series = self.df[self.config.separator_column]
            analysis['separator_column'] = self.config.separator_column
            analysis['unique_values'] = separator_series.nunique()
            analysis['null_count'] = separator_series.isnull().sum()
            analysis['null_percentage'] = (analysis['null_count'] / len(separator_series)) * 100
            
            # Top valores más frecuentes
            value_counts = separator_series.value_counts()
            analysis['top_values'] = value_counts.head(10).to_dict()
            analysis['estimated_groups'] = analysis['unique_values'] + (1 if analysis['null_count'] > 0 else 0)
        
        # Estimación de rendimiento
        total_mb = analysis['memory_usage_mb']
        estimated_processing_time = total_mb * 0.5  # 0.5 segundos por MB (estimado)
        analysis['estimated_processing_time'] = estimated_processing_time
        analysis['recommended_chunking'] = (
            total_mb > 500 or
            analysis['estimated_groups'] > 100 or
            self.enable_chunking
        )
        
        # Recomendaciones
        recommendations = []
        if analysis['null_percentage'] > 20:
            recommendations.append("Alto porcentaje de nulos detectado - considerar limpieza de datos")
        if analysis['estimated_groups'] > 500:
            recommendations.append("Muchos grupos detectados - usar chunking para mejor rendimiento")
        if total_mb > 1024:
            recommendations.append("Dataset grande - habilitar chunking agresivo")
        
        analysis['recommendations'] = recommendations
        
        return analysis
    
    def generate_file_preview(self) -> List[Dict[str, Any]]:
        """Generar preview de archivos que se crearán"""
        preview = []
        
        try:
            # Obtener grupos únicos
            groups = list(self.df.groupby(self.config.separator_column))
            
            for group_name, group_df in groups:
                # Generar nombre de archivo
                filename = self._generate_filename_for_group(group_name, len(group_df))
                file_path = os.path.join(self.config.output_folder, filename)
                
                # Estimar tamaño de archivo (aproximado)
                estimated_size_kb = len(group_df) * 0.5  # ~0.5KB por fila (estimado)
                
                preview.append({
                    'filename': filename,
                    'group_name': str(group_name),
                    'rows': len(group_df),
                    'estimated_size_kb': estimated_size_kb,
                    'file_path': file_path,
                    'status': 'ready'
                })
        
        except Exception as e:
            self.logger.error(f"Error generando preview: {str(e)}")
            return []
        
        return preview
    
    def _generate_filename_for_group(self, group_name: str, row_count: int) -> str:
        """Generar nombre de archivo para un grupo específico"""
        group_info = {
            'valor': group_name,
            'columna': self.config.separator_column,
            'fecha': datetime.now().strftime('%Y-%m-%d'),
            'fecha_hora': datetime.now().strftime('%Y-%m-%d_%H%M'),
            'filas': row_count,
            'timestamp': str(int(time.time()))
        }
        
        return self._process_file_template(self.config.file_template, group_info)
    
    def _process_file_template(self, template: str, group_info: Dict[str, Any]) -> str:
        """Procesar plantilla de nombre de archivo con placeholders"""
        processed = template
        
        # Placeholders soportados
        placeholders = {
            '{valor}': str(group_info.get('valor', '')),
            '{columna}': str(group_info.get('columna', '')),
            '{fecha}': str(group_info.get('fecha', '')),
            '{fecha_hora}': str(group_info.get('fecha_hora', '')),
            '{filas}': str(group_info.get('filas', '')),
            '{timestamp}': str(group_info.get('timestamp', ''))
        }
        
        # Reemplazar placeholders
        for placeholder, value in placeholders.items():
            processed = processed.replace(placeholder, value)
        
        # Sanitizar nombre de archivo
        return self._sanitize_filename(processed)
    
    def _sanitize_filename(self, filename: str) -> str:
        """Sanitizar nombre para compatibilidad del SO"""
        # Caracteres prohibidos
        forbidden_chars = r'[<>:"/\\|?*]'
        sanitized = re.sub(forbidden_chars, '_', filename)
        
        # Remover puntos múltiples al final
        sanitized = re.sub(r'\.+$', '', sanitized)
        
        # Normalizar espacios
        sanitized = ' '.join(sanitized.split())
        
        # Asegurar que no esté vacío
        if not sanitized.strip():
            sanitized = 'archivo_sin_nombre.xlsx'
        
        # Límite de longitud
        if len(sanitized) > 255:
            name_part, ext = os.path.splitext(sanitized)
            available_length = 255 - len(ext)
            sanitized = name_part[:available_length-3] + '...' + ext
        
        return sanitized
    
    def separate_and_export(self) -> Dict[str, Any]:
        """Ejecutar separación completa y exportación"""
        start_time = time.time()
        
        try:
            # 1. Validar configuración
            validation = self.validate_configuration()
            if not validation.is_valid:
                return {
                    'success': False,
                    'error': 'Configuración inválida: ' + '; '.join(validation.errors),
                    'validation_errors': validation.errors,
                    'validation_warnings': validation.warnings
                }
            
            # 2. Analizar datos
            analysis = self.analyze_data()
            if analysis.get('estimated_groups', 0) == 0:
                return {
                    'success': False,
                    'error': 'No hay grupos para separar'
                }
            
            # 3. Procesar separación
            results = []
            groups_processed = 0
            
            for group_name, group_df in self.df.groupby(self.config.separator_column):
                try:
                    # Verificar cancelación
                    if hasattr(self, '_cancelled') and self._cancelled:
                        break
                    
                    # Procesar grupo individual
                    result = self._export_group(str(group_name), group_df)
                    results.append(result)
                    
                    groups_processed += 1
                    
                    # Callback de progreso
                    if self.progress_callback:
                        self.progress_callback(groups_processed, analysis['estimated_groups'])
                    
                except Exception as e:
                    self.logger.error(f"Error procesando grupo {group_name}: {str(e)}")
                    error_result = ExportResult(
                        success=False,
                        group_name=str(group_name),
                        error=str(e)
                    )
                    results.append(error_result)
                    self.failed_groups[str(group_name)] = str(e)
            
            # 4. Generar resumen
            processing_time = time.time() - start_time
            successful_exports = [r for r in results if r.success]
            failed_exports = [r for r in results if not r.success]
            
            return {
                'success': len(successful_exports) > 0,
                'files_created': [r.file_path for r in successful_exports],
                'groups_processed': groups_processed,
                'total_rows': self.df.shape[0],
                'successful_exports': len(successful_exports),
                'failed_exports': len(failed_exports),
                'processing_time': processing_time,
                'errors': validation.errors,
                'warnings': validation.warnings,
                'failed_groups': self.failed_groups,
                'analysis': analysis
            }
            
        except Exception as e:
            self.logger.error(f"Error crítico en separación: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'processing_time': time.time() - start_time
            }
    
    def _export_group(self, group_name: str, group_df: pd.DataFrame) -> ExportResult:
        """Exportar un grupo individual"""
        start_time = time.time()
        result = ExportResult()
        result.group_name = group_name
        result.timestamp = datetime.now()
        
        try:
            # Generar nombre de archivo
            filename = self._generate_filename_for_group(group_name, len(group_df))
            file_path = os.path.join(self.config.output_folder, filename)
            
            # Verificar conflictos de nombres
            file_path = self._resolve_filename_conflicts(file_path)
            
            # Crear archivo Excel con plantilla
            success = self._create_excel_file_with_template(file_path, group_df)
            
            if success:
                result.success = True
                result.file_path = file_path
                result.rows_processed = len(group_df)
                result.processing_time = time.time() - start_time
                self.created_files.append(file_path)
            else:
                result.success = False
                result.error = "Error creando archivo Excel"
            
            return result
            
        except Exception as e:
            result.success = False
            result.error = str(e)
            result.processing_time = time.time() - start_time
            return result
    
    def _create_excel_file_with_template(self, output_path: str, data: pd.DataFrame) -> bool:
        """Crear archivo Excel usando plantilla preservando formato"""
        try:
            # Cargar plantilla
            workbook = load_workbook(self.config.template_path, data_only=False)
            sheet = workbook.active
            
            # Determinar posición inicial
            start_row, start_col = coordinate_to_tuple(self.config.start_cell)
            
            # Aplicar mapeo de columnas
            if not self.config.column_mapping:
                # Usar mapeo por defecto
                self.config.column_mapping = self.config.get_default_mapping(data.columns.tolist())
            
            # Insertar datos preservando formato
            for row_offset, (_, row_data) in enumerate(data.iterrows()):
                excel_row = start_row + row_offset
                
                for df_col, excel_col_letter in self.config.column_mapping.items():
                    if df_col in data.columns:
                        excel_col_idx = column_index_from_string(excel_col_letter)
                        cell = sheet.cell(row=excel_row, column=excel_col_idx)
                        
                        # Insertar valor
                        value = row_data[df_col]
                        if pd.isna(value):
                            cell.value = None
                        else:
                            cell.value = value
            
            # Guardar archivo
            workbook.save(output_path)
            workbook.close()
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error creando archivo Excel: {str(e)}")
            return False
    
    def _resolve_filename_conflicts(self, file_path: str) -> str:
        """Resolver conflictos de nombres de archivo"""
        if not os.path.exists(file_path):
            return file_path
        
        # Auto-numeración
        name_part, ext = os.path.splitext(file_path)
        counter = 1
        
        while True:
            new_path = f"{name_part}_{counter:02d}{ext}"
            if not os.path.exists(new_path):
                return new_path
            counter += 1
            
            # Límite de seguridad
            if counter > 999:
                timestamp = str(int(time.time()))
                return f"{name_part}_{timestamp}{ext}"
    
    def cancel_operation(self):
        """Cancelar operación en curso"""
        self._cancelled = True
    
    def cleanup_temp_files(self):
        """Limpiar archivos temporales en caso de cancelación"""
        for file_path in self.created_files:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception as e:
                self.logger.warning(f"No se pudo limpiar archivo temporal {file_path}: {str(e)}")


def exportar_datos_separados(df: pd.DataFrame, config_dict: dict) -> dict:
    """
    Exportar DataFrame a archivos Excel separados usando plantillas
    
    Args:
        df: DataFrame a separar
        config_dict: Configuración de separación
            - separator_column: str
            - template_path: str
            - start_cell: str (ej: 'A5')
            - output_folder: str
            - file_template: str
            - column_mapping: Dict[str, str]
            - handle_duplicates: str ('overwrite', 'append', 'skip')
    
    Returns:
        dict con resultado:
            - success: bool
            - files_created: List[str]
            - groups_processed: int
            - total_rows: int
            - processing_time: float
            - errors: List[str]
            - warnings: List[str]
    """
    try:
        # Crear configuración
        config = ExportSeparatedConfig(
            separator_column=config_dict.get('separator_column', ''),
            template_path=config_dict.get('template_path', ''),
            start_cell=config_dict.get('start_cell', 'A1'),
            output_folder=config_dict.get('output_folder', ''),
            file_template=config_dict.get('file_template', '{valor}.xlsx'),
            column_mapping=config_dict.get('column_mapping', {}),
            handle_duplicates=config_dict.get('handle_duplicates', 'overwrite'),
            create_summary=config_dict.get('create_summary', True),
            preserve_format=config_dict.get('preserve_format', True),
            enable_chunking=config_dict.get('enable_chunking', True),
            max_memory_mb=config_dict.get('max_memory_mb', 2048)
        )
        
        # Crear splitter y ejecutar
        splitter = ExcelTemplateSplitter(df, config)
        result = splitter.separate_and_export()
        
        return {
            'success': result.get('success', False),
            'files_created': result.get('files_created', []),
            'groups_processed': result.get('groups_processed', 0),
            'total_rows': result.get('total_rows', 0),
            'processing_time': result.get('processing_time', 0.0),
            'errors': result.get('errors', []),
            'warnings': result.get('warnings', []),
            'failed_groups': result.get('failed_groups', {}),
            'analysis': result.get('analysis', {})
        }
        
    except Exception as e:
        print(f"Error en separación de datos: {str(e)}")
        return {
            'success': False,
            'files_created': [],
            'groups_processed': 0,
            'total_rows': 0,
            'processing_time': 0.0,
            'errors': [str(e)],
            'warnings': [],
            'failed_groups': {},
            'analysis': {}
        }