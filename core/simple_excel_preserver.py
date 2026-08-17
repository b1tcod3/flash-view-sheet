"""
Solución simplificada para preservar formato Excel sin usar deepcopy.
Evita problemas de recursión infinita.
"""

import openpyxl
from openpyxl.utils import coordinate_to_tuple, column_index_from_string
from typing import Any
import copy

class SimpleExcelFormatPreserver:
    """Versión simplificada para preservar formato sin problemas de recursión"""
    
    def __init__(self) -> None:
        self.saved_cell_data: dict[str, Any] = {}
    
    def save_cell_format(self, cell: Any) -> dict[str, Any]:
        """
        Guardar formato de una celda sin usar deepcopy
        
        Args:
            cell: Celda de openpyxl
            
        Returns:
            Dict con formato de la celda
        """
        format_info = {
            'value': cell.value,
            'font': {
                'name': cell.font.name,
                'size': cell.font.size,
                'bold': cell.font.bold,
                'italic': cell.font.italic,
                'color': cell.font.color.rgb if (cell.font.color and cell.font.color.type == 'rgb') else None
            },
            'fill': {
                'start_color': cell.fill.start_color.rgb if cell.fill.start_color else None,
                'fill_type': cell.fill.fill_type
            },
            'border': {
                'left': self._serialize_border(cell.border.left),
                'right': self._serialize_border(cell.border.right),
                'top': self._serialize_border(cell.border.top),
                'bottom': self._serialize_border(cell.border.bottom)
            },
            'alignment': {
                'horizontal': cell.alignment.horizontal,
                'vertical': cell.alignment.vertical,
                'wrap_text': cell.alignment.wrap_text
            },
            'number_format': cell.number_format
        }
        
        return format_info
    
    @staticmethod
    def _serialize_border(side: Any) -> dict[str, Any]:
        """Serializar border side"""
        return {
            'style': side.style,
            'color': side.color.rgb if (side.color and side.color.type == 'rgb') else None
        }
    
    @staticmethod
    def restore_cell_format(cell: Any, format_info: dict[str, Any]) -> None:
        """
        Restaurar formato de una celda
        
        Args:
            cell: Celda de openpyxl
            format_info: Dict con formato a restaurar
        """
        try:
            # Restaurar font construyendo un nuevo objeto (evita mutar estilos inmutables)
            if format_info.get('font'):
                from openpyxl.styles import Font, Color
                font_info = format_info['font']
                kwargs = {}
                if font_info.get('name'):
                    kwargs['name'] = font_info['name']
                if font_info.get('size'):
                    kwargs['size'] = font_info['size']
                if font_info.get('bold') is not None:
                    kwargs['bold'] = font_info['bold']
                if font_info.get('italic') is not None:
                    kwargs['italic'] = font_info['italic']
                if font_info.get('color'):
                    kwargs['color'] = Color(rgb=font_info['color'])
                if kwargs:
                    cell.font = Font(**kwargs)
            
            # Restaurar fill
            if format_info.get('fill'):
                fill_info = format_info['fill']
                if fill_info.get('start_color') and fill_info.get('fill_type'):
                    from openpyxl.styles import PatternFill
                    cell.fill = PatternFill(
                        start_color=fill_info['start_color'],
                        end_color=fill_info['start_color'],
                        fill_type=fill_info['fill_type']
                    )
            
            # Restaurar border
            if format_info.get('border'):
                border_info = format_info['border']
                from openpyxl.styles import Border, Side
                
                sides = {}
                for side_name, side_info in border_info.items():
                    if side_info and side_info.get('style'):
                        side = Side(
                            border_style=side_info['style'],
                            color=Color(rgb=side_info['color']) if side_info.get('color') else None
                        )
                        sides[side_name] = side
                
                if sides:
                    cell.border = Border(**sides)
            
            # Restaurar alignment
            if format_info.get('alignment'):
                align_info = format_info['alignment']
                from openpyxl.styles import Alignment
                alignment_kwargs = {}
                if align_info.get('horizontal'):
                    alignment_kwargs['horizontal'] = align_info['horizontal']
                if align_info.get('vertical'):
                    alignment_kwargs['vertical'] = align_info['vertical']
                if align_info.get('wrap_text') is not None:
                    alignment_kwargs['wrap_text'] = align_info['wrap_text']
                
                if alignment_kwargs:
                    cell.alignment = Alignment(**alignment_kwargs)
            
            # Restaurar number format
            if format_info.get('number_format') and format_info['number_format'] != 'General':
                cell.number_format = format_info['number_format']
        
        except Exception as e:
            print(f"Warning: No se pudo restaurar formato completo: {e}")
    
    def backup_area_formatting(self, worksheet: Any, start_cell: str, area_size: tuple[int, int]) -> dict[str, Any]:
        """
        Backup del formato en un área específica
        
        Args:
            worksheet: Worksheet de openpyxl
            start_cell: Celda inicial (ej: 'A5')
            area_size: Tamaño del área (filas, columnas)
            
        Returns:
            Dict con formatos respaldados
        """
        start_row, start_col = coordinate_to_tuple(start_cell)
        rows, cols = area_size
        
        backup = {}
        
        # Backup de celdas
        for row_offset in range(rows):
            for col_offset in range(cols):
                cell_row = start_row + row_offset
                cell_col = start_col + col_offset
                cell = worksheet.cell(row=cell_row, column=cell_col)
                
                if cell.value is not None:
                    coord = f"{openpyxl.utils.get_column_letter(cell_col)}{cell_row}"
                    backup[coord] = self.save_cell_format(cell)
        
        return backup
    
    def insert_data_simple_preservation(self, worksheet: Any, data: dict[str, Any], 
                                      column_mapping: dict[str, str], start_cell: str) -> None:
        """
        Insertar datos preservando formato de manera simple
        
        Args:
            worksheet: Worksheet de openpyxl
            data: DataFrame a insertar como dict
            column_mapping: Mapeo de columnas
            start_cell: Celda inicial
        """
        start_row, _ = coordinate_to_tuple(start_cell)
        
        # Backup formato de área donde se insertarán datos
        max_rows = len(data)
        max_cols = len(column_mapping)
        
        area_backup = self.backup_area_formatting(worksheet, start_cell, (max_rows + 5, max_cols + 2))
        
        # Insertar datos
        for row_offset, (_, row_data) in enumerate(data.items()):
            excel_row = start_row + row_offset
            
            for df_col, excel_col_letter in column_mapping.items():
                if df_col in row_data:
                    excel_col_idx = column_index_from_string(excel_col_letter)
                    cell = worksheet.cell(row=excel_row, column=excel_col_idx)
                    
                    # Insertar valor
                    value = row_data[df_col]
                    if value is None:
                        cell.value = None
                    else:
                        cell.value = value
        
        # Restaurar formato de área (solo para celdas que tenían formato)
        for cell_coord, format_info in area_backup.items():
            try:
                cell = worksheet[cell_coord]
                self.restore_cell_format(cell, format_info)
            except Exception as e:
                print(f"Warning: No se pudo restaurar formato de {cell_coord}: {e}")
        
        # Heredar formato de la fila de encabezados hacia las filas de datos
        self._inherit_header_format(worksheet, column_mapping, start_cell, max_rows, area_backup)
    
    @staticmethod
    def _has_style(cell: Any) -> bool:
        """Determinar si una celda tiene formato aplicado (distinto del predeterminado)"""
        try:
            font = cell.font
            fill = cell.fill
            border = cell.border
            
            if font.bold or font.italic:
                return True
            if font.color and font.color.type == 'rgb' and font.color.rgb not in ('FF000000', '00000000'):
                return True
            if fill.fill_type not in (None, 'none'):
                return True
            if any(getattr(border, side).style for side in ('left', 'right', 'top', 'bottom')):
                return True
            if cell.number_format != 'General':
                return True
        except Exception:
            pass
        return False
    
    def _inherit_header_format(self, worksheet: Any, column_mapping: dict[str, str],
                               start_cell: str, max_rows: int,
                               restored_coords: set[str] | None = None) -> None:
        """
        Copiar el formato de la fila de encabezados (fila 1) hacia las filas de datos
        que no tengan formato propio.
        """
        start_row, _ = coordinate_to_tuple(start_cell)
        if start_row <= 1:
            return
        
        # Pre-calcular columnas cuyo encabezado tiene formato
        styled_cols = []
        for excel_col_letter in column_mapping.values():
            excel_col_idx = column_index_from_string(excel_col_letter)
            if self._has_style(worksheet.cell(row=1, column=excel_col_idx)):
                styled_cols.append(excel_col_idx)
        
        if not styled_cols:
            return
        
        # Construir los estilos base una sola vez por columna (objetos reutilizables)
        base_styles = {}
        for excel_col_idx in styled_cols:
            header_cell = worksheet.cell(row=1, column=excel_col_idx)
            base_styles[excel_col_idx] = (
                copy.copy(header_cell.font),
                copy.copy(header_cell.fill),
                copy.copy(header_cell.border),
                header_cell.number_format
            )
        
        restored = restored_coords or set()
        
        for row_offset in range(max_rows):
            excel_row = start_row + row_offset
            for excel_col_idx in styled_cols:
                coord = f"{openpyxl.utils.get_column_letter(excel_col_idx)}{excel_row}"
                # Celdas con formato de plantilla restaurado: no sobreescribir
                if coord in restored:
                    continue
                base_font, base_fill, base_border, base_numfmt = base_styles[excel_col_idx]
                data_cell = worksheet.cell(row=excel_row, column=excel_col_idx)
                data_cell.font = base_font
                data_cell.fill = base_fill
                data_cell.border = base_border
                data_cell.number_format = base_numfmt

def create_excel_with_simple_format_preservation(template_path: str, output_path: str, 
                                               data: dict[str, Any], column_mapping: dict[str, str],
                                               start_cell: str) -> bool:
    """
    Función utilitaria simple para preservar formato Excel
    
    Args:
        template_path: Ruta de plantilla original
        output_path: Ruta de salida
        data: Data a insertar
        column_mapping: Mapeo de columnas
        start_cell: Celda inicial
        
    Returns:
        bool: True si es exitoso
    """
    try:
        # Cargar plantilla
        workbook = openpyxl.load_workbook(template_path, data_only=False)
        sheet = workbook.active
        
        # Crear preserver simple
        preserver = SimpleExcelFormatPreserver()
        
        # Insertar datos con preservación
        preserver.insert_data_simple_preservation(
            sheet, data, column_mapping, start_cell
        )
        
        # Guardar archivo
        workbook.save(output_path)
        workbook.close()
        
        return True
        
    except Exception as e:
        print(f"Error creando archivo con formato preservado: {e}")
        return False