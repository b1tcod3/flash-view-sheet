"""
Módulo para preservar formato Excel durante la inserción de datos.
Soluciona el problema de que las plantillas cambian de formato al insertar valores.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment, NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Dict, Any, List, Tuple
import copy


class ExcelFormatPreserver:
    """Clase para preservar formato Excel durante modificaciones"""
    
    def __init__(self):
        self.cached_formats = {}
        self.original_styles = {}
    
    def cache_workbook_formats(self, workbook) -> Dict[str, Any]:
        """
        Cache todos los formatos de un workbook para preservar durante modificaciones
        
        Args:
            workbook: Workbook de openpyxl a cachear
            
        Returns:
            Dict con todos los formatos cacheados
        """
        formats_cache = {
            'worksheet_formats': {},
            'cell_formats': {},
            'column_formats': {},
            'row_formats': {},
            'named_styles': {},
            'merged_cells': []
        }
        
        # Cachear estilos nombrados
        if hasattr(workbook, 'named_styles'):
            if isinstance(workbook.named_styles, dict):
                for style_name, style in workbook.named_styles.items():
                    formats_cache['named_styles'][style_name] = copy.deepcopy(style)
            else:
                # named_styles es una lista
                for style in workbook.named_styles:
                    if hasattr(style, 'name'):
                        formats_cache['named_styles'][style.name] = copy.deepcopy(style)
        
        # Cachear formatos por worksheet
        if hasattr(workbook, 'worksheets'):
            if isinstance(workbook.worksheets, dict):
                worksheets_items = workbook.worksheets.items()
            else:
                # worksheets es una lista
                worksheets_items = [(ws.title, ws) for ws in workbook.worksheets]
            
            for sheet_name, worksheet in worksheets_items:
                sheet_cache = {
                    'cell_formats': {},
                    'column_widths': {},
                    'row_heights': {},
                    'merged_cells': list(worksheet.merged_cells.ranges),
                    'page_setup': getattr(worksheet, 'page_setup', None),
                    'print_options': getattr(worksheet, 'print_options', None)
                }
                
                # Cachear formato de celdas individuales
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is not None or self._has_formatting(cell):
                            cell_format = self._extract_cell_format(cell)
                            sheet_cache['cell_formats'][cell.coordinate] = cell_format
                
                # Cachear anchos de columna
                for col_letter, col_dim in worksheet.column_dimensions.items():
                    sheet_cache['column_widths'][col_letter] = col_dim.width
                
                # Cachear alturas de fila
                for row_num, row_dim in worksheet.row_dimensions.items():
                    sheet_cache['row_heights'][row_num] = row_dim.height
                
                formats_cache['worksheet_formats'][sheet_name] = sheet_cache
        
        return formats_cache
    
    def restore_workbook_formats(self, workbook, formats_cache: Dict[str, Any]) -> None:
        """
        Restaurar todos los formatos de un workbook desde el cache
        
        Args:
            workbook: Workbook de openpyxl
            formats_cache: Cache de formatos a restaurar
        """
        # Restaurar estilos nombrados
        for style_name, style in formats_cache.get('named_styles', {}).items():
            workbook.named_styles[style_name] = style
        
        # Restaurar formatos por worksheet
        for sheet_name, sheet_cache in formats_cache.get('worksheet_formats', {}).items():
            # Verificar si la hoja existe (soporte para diferentes versiones de openpyxl)
            worksheet = None
            if hasattr(workbook, 'sheetnames') and sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            elif hasattr(workbook, 'worksheets'):
                for ws in workbook.worksheets:
                    if ws.title == sheet_name:
                        worksheet = ws
                        break
            
            if worksheet:
                # Restaurar merged cells
                if hasattr(worksheet.merged_cells, 'ranges'):
                    worksheet.merged_cells.ranges.clear()
                    for merged_range in sheet_cache.get('merged_cells', []):
                        worksheet.merge_cells(str(merged_range))
                
                # Restaurar anchos de columna
                for col_letter, width in sheet_cache.get('column_widths', {}).items():
                    if hasattr(worksheet.column_dimensions, 'items'):
                        worksheet.column_dimensions[col_letter].width = width
                
                # Restaurar alturas de fila
                for row_num, height in sheet_cache.get('row_heights', {}).items():
                    if hasattr(worksheet.row_dimensions, 'items'):
                        worksheet.row_dimensions[row_num].height = height
                
                # Restaurar formato de celdas individuales
                for cell_coord, cell_format in sheet_cache.get('cell_formats', {}).items():
                    try:
                        cell = worksheet[cell_coord]
                        self._apply_cell_format(cell, cell_format)
                    except Exception as e:
                        print(f"Warning: No se pudo restaurar formato de celda {cell_coord}: {e}")
    
    def _extract_cell_format(self, cell) -> Dict[str, Any]:
        """
        Extraer formato completo de una celda
        
        Args:
            cell: Celda de openpyxl
            
        Returns:
            Dict con formato de la celda
        """
        format_dict = {
            'font': None,
            'fill': None,
            'border': None,
            'alignment': None,
            'number_format': None,
            'protection': None
        }
        
        # Extraer formato si existe
        if cell.font != Font():
            format_dict['font'] = copy.deepcopy(cell.font)
        
        if cell.fill != PatternFill():
            format_dict['fill'] = copy.deepcopy(cell.fill)
        
        if cell.border != Border():
            format_dict['border'] = copy.deepcopy(cell.border)
        
        if cell.alignment != Alignment():
            format_dict['alignment'] = copy.deepcopy(cell.alignment)
        
        if cell.number_format != 'General':
            format_dict['number_format'] = cell.number_format
        
        if cell.protection:
            format_dict['protection'] = copy.deepcopy(cell.protection)
        
        return format_dict
    
    def _apply_cell_format(self, cell, format_dict: Dict[str, Any]) -> None:
        """
        Aplicar formato a una celda preservando valor
        
        Args:
            cell: Celda de openpyxl
            format_dict: Dict con formato a aplicar
        """
        if format_dict.get('font'):
            cell.font = format_dict['font']
        
        if format_dict.get('fill'):
            cell.fill = format_dict['fill']
        
        if format_dict.get('border'):
            cell.border = format_dict['border']
        
        if format_dict.get('alignment'):
            cell.alignment = format_dict['alignment']
        
        if format_dict.get('number_format'):
            cell.number_format = format_dict['number_format']
        
        if format_dict.get('protection'):
            cell.protection = format_dict['protection']
    
    def _has_formatting(self, cell) -> bool:
        """
        Verificar si una celda tiene formato personalizado
        
        Args:
            cell: Celda de openpyxl
            
        Returns:
            bool: True si la celda tiene formato
        """
        return (
            cell.font != Font() or
            cell.fill != PatternFill() or
            cell.border != Border() or
            cell.alignment != Alignment() or
            cell.number_format != 'General'
        )
    
    def insert_data_preserving_format(self, worksheet, data: Dict[Tuple[int, int], Any], 
                                    column_mapping: Dict[str, str], start_cell: str) -> None:
        """
        Insertar datos en worksheet preservando formato existente
        
        Args:
            worksheet: Worksheet de openpyxl
            data: DataFrame a insertar (dict de {row_offset: {col_name: value}})
            column_mapping: Mapeo de columnas DataFrame -> Excel
            start_cell: Celda inicial (ej: 'A1')
        """
        from openpyxl.utils import coordinate_to_tuple
        
        # Obtener posición inicial
        start_row, start_col = coordinate_to_tuple(start_cell)
        
        # Cachear formatos antes de insertar
        self._cache_worksheet_formats(worksheet)
        
        try:
            # Insertar datos SIN tocar el formato
            for row_offset, (_, row_data) in enumerate(data.items()):
                excel_row = start_row + row_offset
                
                for df_col, excel_col_letter in column_mapping.items():
                    if df_col in row_data:
                        excel_col_idx = column_index_from_string(excel_col_letter)
                        cell = worksheet.cell(row=excel_row, column=excel_col_idx)
                        
                        # Guardar valor actual para preservarlo
                        current_value = cell.value
                        
                        # Insertar nuevo valor
                        value = row_data[df_col]
                        if value is None:
                            cell.value = None
                        else:
                            cell.value = value
                        
                        # Restaurar formato si es necesario
                        if self._has_formatting(cell):
                            self._restore_cell_format(cell, excel_row, excel_col_idx)
        
        except Exception as e:
            print(f"Error insertando datos con formato preservado: {e}")
            # Intentar restaurar formatos en caso de error
            self._restore_all_formats(worksheet)
    
    def _cache_worksheet_formats(self, worksheet) -> None:
        """Cachear formatos de worksheet para restaurar después"""
        self.worksheet_cache = {}
        
        for row in worksheet.iter_rows():
            for cell in row:
                if self._has_formatting(cell):
                    self.worksheet_cache[cell.coordinate] = self._extract_cell_format(cell)
    
    def _restore_cell_format(self, cell, row: int, col: int) -> None:
        """Restaurar formato de una celda específica"""
        coord = f"{get_column_letter(col)}{row}"
        if coord in self.worksheet_cache:
            self._apply_cell_format(cell, self.worksheet_cache[coord])
    
    def _restore_all_formats(self, worksheet) -> None:
        """Restaurar todos los formatos del worksheet"""
        for cell_coord, format_dict in self.worksheet_cache.items():
            try:
                cell = worksheet[cell_coord]
                self._apply_cell_format(cell, format_dict)
            except Exception as e:
                print(f"Warning: No se pudo restaurar formato de {cell_coord}: {e}")


def create_template_with_preserved_format(template_path: str, output_path: str, 
                                        data: Dict[str, Any], column_mapping: Dict[str, str],
                                        start_cell: str) -> bool:
    """
    Función utilitaria para crear archivo Excel con formato preservado
    
    Args:
        template_path: Ruta de plantilla original
        output_path: Ruta de salida
        data: Data a insertar
        column_mapping: Mapeo de columnas
        start_cell: Celda inicial
        
    Returns:
        bool: True si es exitoso
    """
    try:
        # Cargar plantilla preservando formato
        workbook = openpyxl.load_workbook(
            template_path, 
            data_only=False,  # No interpretar datos, preservar formato
            keep_vba=True,    # Preservar macros si existen
            keep_links=True,  # Preservar links
            keep_images=True  # Preservar imágenes
        )
        
        # Crear preserver
        preserver = ExcelFormatPreserver()
        
        # Procesar cada worksheet
        for sheet_name, sheet in workbook.worksheets:
            # Cachear formatos originales
            formats_cache = preserver.cache_workbook_formats(workbook)
            
            # Insertar datos preservando formato
            preserver.insert_data_preserving_format(
                sheet, data, column_mapping, start_cell
            )
            
            # Restaurar formatos (ya preservados por insert_data_preserving_format)
        
        # Guardar workbook
        workbook.save(output_path)
        workbook.close()
        
        return True
        
    except Exception as e:
        print(f"Error creando archivo con formato preservado: {e}")
        return False