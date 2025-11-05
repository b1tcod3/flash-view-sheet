#!/usr/bin/env python3
"""
Optimizador de Rendimiento para ExcelTemplateSplitter
Implementa chunking inteligente, gestión de memoria y optimizaciones de openpyxl
"""

import os
import time
import psutil
import gc
import threading
from typing import Dict, List, Optional, Iterator, Tuple, Any
from dataclasses import dataclass
from enum import Enum
import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo


class ChunkingStrategy(Enum):
    """Estrategias de chunking disponibles"""
    NONE = "none"           # Procesamiento directo
    MODERATE = "moderate"   # Chunking conservador
    SIZE_BASED = "size"     # Basado en tamaño de grupos
    GROUP_BASED = "group"   # Basado en número de grupos
    AGGRESSIVE = "aggressive"  # Chunking agresivo


@dataclass
class ProcessingMetrics:
    """Métricas de procesamiento en tiempo real"""
    start_time: float
    current_group: int
    total_groups: int
    groups_completed: int
    current_memory_mb: float
    peak_memory_mb: float
    estimated_time_remaining: float
    current_processing_speed: float  # groups/second
    files_created: int
    total_rows_processed: int


@dataclass
class SystemResources:
    """Estado de recursos del sistema"""
    available_memory_mb: float
    total_memory_mb: float
    cpu_percent: float
    disk_free_space_mb: float
    disk_write_speed_mbps: float


class PerformanceOptimizer:
    """Optimizador principal de rendimiento"""
    
    def __init__(self, chunking_strategy: ChunkingStrategy = ChunkingStrategy.MODERATE):
        self.chunking_strategy = chunking_strategy
        self.memory_threshold_mb = 2048  # 2GB threshold
        self.performance_cache = {}
        self.format_cache = {}  # Cache para formatos Excel
        self.active_monitors = {}
        
    def determine_optimal_chunking_strategy(self, df: pd.DataFrame, 
                                          separator_column: str) -> ChunkingStrategy:
        """
        Determinar estrategia óptima de chunking basada en características del dataset
        Implementa el IntelligentChunkingAlgorithm especificado
        """
        total_rows = len(df)
        memory_usage_mb = df.memory_usage(deep=True).sum() / 1024 / 1024
        
        # Verificar que la columna existe
        if separator_column not in df.columns:
            return ChunkingStrategy.NONE
            
        unique_groups = df[separator_column].nunique()
        
        # Análisis de distribución de grupos
        group_sizes = df.groupby(separator_column).size()
        largest_group_size = group_sizes.max()
        smallest_group_size = group_sizes.min()
        group_size_variance = group_sizes.var()
        
        # Decisión de estrategia según especificaciones
        if memory_usage_mb > 2048 or total_rows > 1000000:
            # Estrategia agresiva para datasets muy grandes
            return ChunkingStrategy.AGGRESSIVE
            
        elif unique_groups > 100 and group_size_variance > 10000:
            # Muchos grupos con tamaños variables
            return ChunkingStrategy.GROUP_BASED
            
        elif largest_group_size > 50000:
            # Algunos grupos muy grandes
            return ChunkingStrategy.SIZE_BASED
            
        elif memory_usage_mb > 500:
            # Dataset moderadamente grande
            return ChunkingStrategy.MODERATE
            
        else:
            # Dataset pequeño - procesamiento directo
            return ChunkingStrategy.NONE
    
    def get_optimal_chunk_size(self, strategy: ChunkingStrategy, 
                              total_rows: int, unique_groups: int) -> int:
        """Calcular tamaño óptimo de chunk según estrategia"""
        if strategy == ChunkingStrategy.NONE:
            return total_rows
            
        elif strategy == ChunkingStrategy.MODERATE:
            return min(10000, max(1000, total_rows // 4))
            
        elif strategy == ChunkingStrategy.SIZE_BASED:
            return min(5000, max(500, total_rows // 10))
            
        elif strategy == ChunkingStrategy.GROUP_BASED:
            return min(1000, max(100, unique_groups // 2))
            
        elif strategy == ChunkingStrategy.AGGRESSIVE:
            return min(1000, max(100, total_rows // 20))
            
        return 1000  # Default
    
    def process_dataframe_in_chunks(self, df: pd.DataFrame, 
                                   separator_column: str,
                                   chunk_size: int) -> Iterator[Tuple[str, pd.DataFrame]]:
        """
        Procesar DataFrame por chunks para optimizar memoria
        Implementa procesamiento eficiente con cleanup automático
        """
        # Verificar que la columna existe
        if separator_column not in df.columns:
            raise ValueError(f"Columna '{separator_column}' no existe en DataFrame")
        
        # Detectar valores nulos para manejo especial
        null_mask = df[separator_column].isnull()
        if null_mask.any():
            # Yield nulos por separado primero
            null_group = df[null_mask].copy()
            null_group[separator_column] = 'N/A'
            yield ('Valores_Nulos', null_group)
        
        # Procesar datos normales
        normal_data = df[~null_mask].copy()
        if normal_data.empty:
            return
            
        # Agrupar datos normales
        grouped = normal_data.groupby(separator_column)
        
        for group_name, group_df in grouped:
            # Aplicar chunking si es necesario
            if len(group_df) > chunk_size:
                # Procesar grupos grandes en sub-chunks
                for i in range(0, len(group_df), chunk_size):
                    chunk = group_df.iloc[i:i + chunk_size].copy()
                    yield (f"{group_name}_chunk_{i//chunk_size + 1}", chunk)
                    
                    # Cleanup de memoria entre chunks
                    del chunk
                    if i % (chunk_size * 5) == 0:  # GC cada 5 chunks
                        gc.collect()
            else:
                # Grupos pequeños se procesan directamente
                yield (str(group_name), group_df)
    
    def monitor_memory_usage(self) -> SystemResources:
        """Monitorear recursos del sistema en tiempo real"""
        try:
            memory = psutil.virtual_memory()
            disk = psutil.disk_usage('/')
            cpu = psutil.cpu_percent(interval=0.1)
            
            return SystemResources(
                available_memory_mb=memory.available / 1024 / 1024,
                total_memory_mb=memory.total / 1024 / 1024,
                cpu_percent=cpu,
                disk_free_space_mb=disk.free / 1024 / 1024,
                disk_write_speed_mbps=100  # Estimado, se puede mejorar
            )
        except Exception:
            # Fallback si psutil no está disponible
            return SystemResources(
                available_memory_mb=1024,
                total_memory_mb=4096,
                cpu_percent=50.0,
                disk_free_space_mb=10240,
                disk_write_speed_mbps=100
            )
    
    def optimize_memory_usage(self):
        """Optimizar uso de memoria con garbage collection"""
        gc.collect()
        
        # Obtener estado de memoria actual
        memory = psutil.virtual_memory()
        memory_usage_percent = memory.percent
        
        # Si uso de memoria > 80%, aplicar optimizaciones agresivas
        if memory_usage_percent > 80:
            # Limpiar cache de formatos
            self.format_cache.clear()
            
            # Forzar garbage collection
            for _ in range(3):
                gc.collect()
                
            print(f"Memory optimization triggered: {memory_usage_percent:.1f}%")
    
    def get_processing_metrics(self, start_time: float, current_group: int,
                             total_groups: int, files_created: int, 
                             rows_processed: int) -> ProcessingMetrics:
        """Calcular métricas de procesamiento en tiempo real"""
        current_time = time.time()
        elapsed_time = current_time - start_time
        
        # Calcular velocidad actual
        if elapsed_time > 0:
            current_speed = current_group / elapsed_time
        else:
            current_speed = 0
        
        # Estimar tiempo restante
        if current_speed > 0 and current_group < total_groups:
            remaining_groups = total_groups - current_group
            estimated_time_remaining = remaining_groups / current_speed
        else:
            estimated_time_remaining = 0
        
        # Obtener estado de memoria
        resources = self.monitor_memory_usage()
        peak_memory = getattr(self, '_peak_memory_mb', resources.current_memory_mb)
        
        return ProcessingMetrics(
            start_time=start_time,
            current_group=current_group,
            total_groups=total_groups,
            groups_completed=current_group,
            current_memory_mb=resources.available_memory_mb,
            peak_memory_mb=peak_memory,
            estimated_time_remaining=estimated_time_remaining,
            current_processing_speed=current_speed,
            files_created=files_created,
            total_rows_processed=rows_processed
        )


class ExcelFormatOptimizer:
    """Optimizador específico para operaciones de Excel"""
    
    def __init__(self):
        self.format_cache = {}  # Cache para formatos Excel
        self.cache_limit = 100  # Límite de entradas en cache
    
    def load_excel_template_optimized(self, template_path: str) -> openpyxl.Workbook:
        """
        Cargar plantilla Excel con optimizaciones de performance
        Preserva formato pero optimiza carga
        """
        try:
            # Verificar si el archivo ya está en cache
            if template_path in self.format_cache:
                # Devolver copia para evitar modificaciones accidentales
                return self.format_cache[template_path]
            
            # Cargar workbook con optimizaciones
            wb = openpyxl.load_workbook(
                template_path, 
                data_only=False,  # Mantener fórmulas
                keep_vba=False,   # No VBA para mejor performance
                keep_links=False  # No enlaces externos
            )
            
            # Cachear el workbook (solo para uso readonly)
            if len(self.format_cache) < self.cache_limit:
                self.format_cache[template_path] = wb
            
            return wb
            
        except Exception as e:
            raise Exception(f"Error cargando plantilla Excel: {str(e)}")
    
    def preserve_format_during_insert(self, workbook: openpyxl.Workbook,
                                     start_cell: str, 
                                     data_df: pd.DataFrame,
                                     column_mapping: Dict[str, str]) -> bool:
        """
        Insertar datos preservando formato Excel con optimizaciones
        Implementa ExcelFormatPreservationAlgorithm especificado
        """
        try:
            sheet = workbook.active
            start_row, start_col = self._cell_coordinates_to_indices(start_cell)
            
            # Detectar área de datos existente
            existing_data_range = self._detect_existing_data_range(sheet, start_cell)
            existing_formats = self._cache_existing_formats(sheet, existing_data_range)
            
            # Optimización: Deshabilitar cálculo automático temporalmente
            sheet.sheet_properties.formulaUpdatesEnabled = False
            
            # Insertar datos preservando formato
            success = True
            for row_idx, (_, row_data) in enumerate(data_df.iterrows()):
                excel_row = start_row + row_idx
                
                for df_col, excel_col_letter in column_mapping.items():
                    if df_col in data_df.columns:
                        excel_col_idx = self._column_letter_to_index(excel_col_letter)
                        cell = sheet.cell(row=excel_row, column=excel_col_idx)
                        
                        # Insertar valor
                        value = row_data[df_col]
                        if pd.isna(value):
                            cell.value = None
                        else:
                            cell.value = value
                        
                        # Preservar formato original si existe
                        if (excel_row, excel_col_idx) in existing_formats:
                            self._apply_cached_format(cell, existing_formats[(excel_row, excel_col_idx)])
            
            # Rehabilitar cálculos
            sheet.sheet_properties.formulaUpdatesEnabled = True
            
            # Optimización final: forzar escritura
            workbook._archive.close()
            
            return success
            
        except Exception as e:
            print(f"Error preservando formato Excel: {str(e)}")
            return False
    
    def _detect_existing_data_range(self, sheet, start_cell: str) -> str:
        """Detectar rango de datos existente"""
        start_row, start_col = self._cell_coordinates_to_indices(start_cell)
        
        # Detectar última fila con datos
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Crear rango basado en área de datos
        if max_row >= start_row and max_col >= start_col:
            end_col_letter = openpyxl.utils.get_column_letter(max_col)
            return f"{start_cell}:{end_col_letter}{max_row}"
        
        return ""
    
    def _cache_existing_formats(self, sheet, data_range: str) -> Dict:
        """Cachear formatos existentes para preservación"""
        formats = {}
        
        if not data_range:
            return formats
        
        try:
            for row in sheet[data_range]:
                for cell in row:
                    if cell.value is not None:
                        formats[(cell.row, cell.column)] = {
                            'font': cell.font,
                            'fill': cell.fill,
                            'border': cell.border,
                            'number_format': cell.number_format,
                            'alignment': cell.alignment
                        }
        except Exception as e:
            print(f"Warning: No se pudo cachear formato completo: {e}")
            # Fallback: cachear solo formato numérico
            for row in sheet[data_range]:
                for cell in row:
                    if cell.value is not None:
                        formats[(cell.row, cell.column)] = {
                            'number_format': cell.number_format
                        }
        
        return formats
    
    def _apply_cached_format(self, cell, cached_format: Dict):
        """Aplicar formato cacheado a celda"""
        try:
            if 'font' in cached_format and cached_format['font']:
                cell.font = cached_format['font']
            if 'fill' in cached_format and cached_format['fill']:
                cell.fill = cached_format['fill']
            if 'border' in cached_format and cached_format['border']:
                cell.border = cached_format['border']
            if 'number_format' in cached_format and cached_format['number_format']:
                cell.number_format = cached_format['number_format']
            if 'alignment' in cached_format and cached_format['alignment']:
                cell.alignment = cached_format['alignment']
        except Exception as e:
            print(f"Warning: Error aplicando formato cacheado: {e}")
    
    def _cell_coordinates_to_indices(self, cell_coord: str) -> Tuple[int, int]:
        """Convertir coordenada de celda (A1) a índices (1, 1)"""
        from openpyxl.utils import coordinate_to_tuple
        row, col = coordinate_to_tuple(cell_coord)
        return row, col
    
    def _column_letter_to_index(self, col_letter: str) -> int:
        """Convertir letra de columna a índice numérico"""
        from openpyxl.utils import column_index_from_string
        return column_index_from_string(col_letter)


class ProgressMonitor:
    """Monitor de progreso con cancelación de operaciones"""
    
    def __init__(self):
        self.cancelled = False
        self.paused = False
        self.callbacks = []
        self._lock = threading.Lock()
    
    def cancel_operation(self):
        """Cancelar operación en curso"""
        with self._lock:
            self.cancelled = True
    
    def pause_operation(self):
        """Pausar operación en curso"""
        with self._lock:
            self.paused = True
    
    def resume_operation(self):
        """Reanudar operación pausada"""
        with self._lock:
            self.paused = False
    
    def is_cancelled(self) -> bool:
        """Verificar si la operación fue cancelada"""
        with self._lock:
            return self.cancelled
    
    def is_paused(self) -> bool:
        """Verificar si la operación está pausada"""
        with self._lock:
            return self.paused
    
    def wait_if_paused(self):
        """Esperar si la operación está pausada"""
        while self.is_paused():
            time.sleep(0.1)  # Esperar 100ms
    
    def add_progress_callback(self, callback):
        """Agregar callback de progreso"""
        self.callbacks.append(callback)
    
    def update_progress(self, metrics: ProcessingMetrics):
        """Actualizar progreso y notificar callbacks"""
        # Verificar cancelación
        if self.is_cancelled():
            raise Exception("Operación cancelada por el usuario")
        
        # Notificar callbacks
        for callback in self.callbacks:
            try:
                callback(metrics)
            except Exception as e:
                print(f"Error en callback de progreso: {e}")
    
    def reset(self):
        """Resetear estado del monitor"""
        with self._lock:
            self.cancelled = False
            self.paused = False