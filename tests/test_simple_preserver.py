#!/usr/bin/env python3
"""
Test específico del SimpleExcelPreserver que está funcionando
"""

import sys
import tempfile
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment

sys.path.insert(0, str(Path(__file__).resolve().parent))

from core.simple_excel_preserver import SimpleExcelFormatPreserver


def test_simple_preserver() -> None:
    """Test del SimpleExcelPreserver que está funcionando"""
    print("🔍 TEST DEL SIMPLE EXCEL PRESERVER")
    print("=" * 50)
    
    # Crear plantilla con formato
    template_path = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    
    # Formato complejo
    ws['A1'] = "TITULO CON FORMATO"
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = "Header 1"
    ws['A3'].font = Font(bold=True, color='0000FF')
    ws['A3'].border = Border(
        left=openpyxl.styles.Side(border_style='thin', color='000000')
    )
    
    ws['B3'] = "Header 2"
    ws['B3'].number_format = '#,##0.00'
    
    # Anchos personalizados
    ws.column_dimensions['A'].width = 25.0
    ws.column_dimensions['B'].width = 15.0
    
    # Datos de ejemplo que se van a sobrescribir
    ws['A5'] = "Dato 1"
    ws['B5'] = 1000
    
    wb.save(template_path.name)
    print(f"✓ Plantilla creada: {template_path.name}")
    
    # Test del preserver
    preserver = SimpleExcelFormatPreserver()
    
    # Cargar plantilla
    test_wb = openpyxl.load_workbook(template_path.name, data_only=False)
    test_ws = test_wb.active
    
    print("✓ Carga de plantilla exitosa")
    
    # Datos para insertar
    test_data = {
        0: {'Región': 'Norte', 'Ventas': 1234},
        1: {'Región': 'Sur', 'Ventas': 5678}
    }
    
    # Preservar formato
    print("✓ Insertando datos con preservación de formato...")
    preserver.insert_data_simple_preservation(
        test_ws, 
        test_data, 
        {'Región': 'A', 'Ventas': 'B'}, 
        'A5'
    )
    
    # Guardar resultado
    output_path = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    test_wb.save(output_path.name)
    test_wb.close()
    
    print("✓ Verificando resultado...")
    
    # Verificar formato
    result_wb = openpyxl.load_workbook(output_path.name, data_only=False)
    result_ws = result_wb.active
    
    # Test de título (debe estar intacto)
    title_cell = result_ws['A1']
    if (title_cell.value == "TITULO CON FORMATO" and 
        title_cell.font.name == 'Arial' and
        title_cell.font.size == 14 and
        title_cell.font.bold and
        title_cell.fill.start_color.rgb == 'FFFF0000'):  # Rojo
        print("✅ Título preservado correctamente")
    else:
        print("❌ Error en título")
    
    # Test de header (debe estar intacto)
    header_cell = result_ws['A3']
    if (header_cell.value == "Header 1" and
        header_cell.font.bold and
        header_cell.font.color.rgb == 'FF0000FF'):  # Azul
        print("✅ Header preservado correctamente")
    else:
        print("❌ Error en header")
    
    # Test de anchos
    if (result_ws.column_dimensions['A'].width == 25.0 and
        result_ws.column_dimensions['B'].width == 15.0):
        print("✅ Anchos de columna preservados")
    else:
        print("❌ Error en anchos de columna")
    
    # Test de datos insertados
    if (result_ws['A5'].value == 'Norte' and
        result_ws['B5'].value == 1234):
        print("✅ Datos insertados correctamente")
    else:
        print("❌ Error insertando datos")
    
    result_wb.close()
    
    # Limpiar
    Path(template_path.name).unlink()
    Path(output_path.name).unlink()
    
    print("\n✅ Test del SimpleExcelPreserver completado exitosamente")
    return True


if __name__ == '__main__':
    test_simple_preserver()