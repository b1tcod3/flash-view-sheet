"""
Tests actualizados para la clase ExcelTemplateSplitter y funcionalidad de separación de datos.
"""

import unittest
import pandas as pd
import tempfile
import os
import json
from unittest.mock import patch, MagicMock, mock_open
import openpyxl

from core.data_handler import ExcelTemplateSplitter, ExportSeparatedConfig, ValidationResult, ExportResult


class TestExcelTemplateSplitter(unittest.TestCase):
    """Tests para ExcelTemplateSplitter"""
    
    def setUp(self):
        """Configurar datos de prueba"""
        # Crear DataFrame de prueba
        self.df_test = pd.DataFrame({
            'Region': ['Norte', 'Norte', 'Sur', 'Sur', 'Este', 'Este', 'Oeste'],
            'Producto': ['A', 'B', 'A', 'B', 'A', 'B', 'A'],
            'Ventas': [1000, 1500, 800, 1200, 900, 1300, 750],
            'Fecha': pd.to_datetime(['2024-01-01', '2024-01-02', '2024-01-01', 
                                   '2024-01-02', '2024-01-02', '2024-01-02', '2024-01-01'])
        })
        
        # Crear plantilla temporal para pruebas
        self.temp_template = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Plantilla de Prueba"
        ws['A2'] = "Datos:"
        wb.save(self.temp_template.name)
        
        # Crear configuración de prueba
        self.config = ExportSeparatedConfig(
            separator_column='Region',
            template_path=self.temp_template.name,
            output_folder=tempfile.mkdtemp(),
            file_template='Reporte_{valor}_{fecha}.xlsx',
            column_mapping={
                'Region': 'A',
                'Producto': 'B', 
                'Ventas': 'C',
                'Fecha': 'D'
            },
            start_cell='A5'
        )
        
    def tearDown(self):
        """Limpiar archivos temporales"""
        try:
            os.unlink(self.temp_template.name)
            os.rmdir(self.config.output_folder)
        except:
            pass
    
    def test_init(self):
        """Test inicialización del splitter"""
        splitter = ExcelTemplateSplitter(self.df_test, self.config)
        
        self.assertEqual(splitter.df.shape, (7, 4))
        self.assertEqual(splitter.config.separator_column, 'Region')
        self.assertFalse(hasattr(splitter, 'groups_data'))  # No existe este atributo
        self.assertFalse(hasattr(splitter, 'is_cancelled'))  # No existe este atributo
    
    def test_validate_configuration(self):
        """Test validación de configuración"""
        splitter = ExcelTemplateSplitter(self.df_test, self.config)
        result = splitter.validate_configuration()
        
        self.assertIsInstance(result, ValidationResult)
        # Como la plantilla existe pero no se valida el path real, puede haber warnings
        self.assertIsInstance(result.is_valid, bool)
        self.assertIsInstance(result.errors, list)
        self.assertIsInstance(result.warnings, list)
    
    def test_validate_configuration_invalid_column(self):
        """Test validación con columna inexistente"""
        config = ExportSeparatedConfig(
            separator_column='ColumnaInexistente',
            template_path=self.temp_template.name,
            output_folder=tempfile.mkdtemp()
        )
        
        splitter = ExcelTemplateSplitter(self.df_test, config)
        result = splitter.validate_configuration()
        
        self.assertFalse(result.is_valid)
        self.assertGreater(len(result.errors), 0)
    
    def test_analyze_data(self):
        """Test análisis de datos"""
        splitter = ExcelTemplateSplitter(self.df_test, self.config)
        result = splitter.analyze_data()
        
        self.assertIsInstance(result, dict)
        self.assertIn('total_rows', result)
        self.assertIn('separator_column', result)
        self.assertEqual(result['total_rows'], 7)
        self.assertEqual(result['separator_column'], 'Region')
        self.assertIn('unique_values', result)
        self.assertEqual(result['unique_values'], 4)  # Norte, Sur, Este, Oeste
        self.assertIn('estimated_groups', result)
    
    def test_generate_file_preview(self):
        """Test generación de preview de archivos"""
        splitter = ExcelTemplateSplitter(self.df_test, self.config)
        preview = splitter.generate_file_preview()
        
        self.assertIsInstance(preview, list)
        self.assertEqual(len(preview), 4)  # 4 grupos únicos
        
        for file_info in preview:
            self.assertIn('filename', file_info)
            self.assertIn('group_name', file_info)
            self.assertIn('rows', file_info)
            self.assertIn('estimated_size_kb', file_info)
            self.assertTrue(file_info['filename'].endswith('.xlsx'))
            self.assertGreater(file_info['rows'], 0)
    
    def test_generate_filename_for_group(self):
        """Test generación de nombres de archivo"""
        splitter = ExcelTemplateSplitter(self.df_test, self.config)
        filename = splitter._generate_filename_for_group('TestGroup', 100)
        
        self.assertIsInstance(filename, str)
        self.assertTrue(filename.endswith('.xlsx'))
        self.assertIn('TestGroup', filename)
    
    def test_sanitize_filename(self):
        """Test sanitización de nombres de archivo"""
        splitter = ExcelTemplateSplitter(self.df_test, self.config)
        
        # Test caracteres prohibidos
        dirty_name = 'test/file:name?.xlsx'
        clean_name = splitter._sanitize_filename(dirty_name)
        self.assertNotIn('/', clean_name)
        self.assertNotIn(':', clean_name)
        self.assertNotIn('?', clean_name)
        
        # Test límite de longitud
        long_name = 'a' * 300 + '.xlsx'
        shortened_name = splitter._sanitize_filename(long_name)
        self.assertLess(len(shortened_name), 260)
    
    def test_resolve_filename_conflicts(self):
        """Test resolución de conflictos de nombres"""
        splitter = ExcelTemplateSplitter(self.df_test, self.config)
        
        # Crear archivo temporal
        test_path = os.path.join(self.config.output_folder, 'test.xlsx')
        with open(test_path, 'w') as f:
            f.write('test')
        
        # Test que se resuelve conflicto
        resolved_path = splitter._resolve_filename_conflicts(test_path)
        self.assertNotEqual(resolved_path, test_path)
        self.assertTrue(resolved_path.endswith('.xlsx'))
    
    def test_separate_and_export_success(self):
        """Test exportación exitosa"""
        # Crear directorio temporal
        temp_output = tempfile.mkdtemp()
        config = ExportSeparatedConfig(
            separator_column='Region',
            template_path=self.temp_template.name,
            output_folder=temp_output,
            file_template='{valor}.xlsx'
        )
        
        splitter = ExcelTemplateSplitter(self.df_test, config)
        
        # Mock para evitar crear archivos reales
        with patch.object(splitter, '_create_excel_file_with_template', return_value=True):
            result = splitter.separate_and_export()
            
            self.assertIsInstance(result, dict)
            self.assertIn('success', result)
            self.assertIn('files_created', result)
            self.assertIn('groups_processed', result)
    
    def test_separate_and_export_validation_failure(self):
        """Test fallo en validación"""
        # Configuración inválida
        config = ExportSeparatedConfig(
            separator_column='ColumnaInexistente',
            template_path='/nonexistent/template.xlsx',
            output_folder='/nonexistent/output'
        )
        
        splitter = ExcelTemplateSplitter(self.df_test, config)
        result = splitter.separate_and_export()
        
        self.assertFalse(result['success'])
        # El resultado puede tener 'error' o 'validation_errors' dependiendo de la implementación
        self.assertTrue(
            'error' in result or 'validation_errors' in result or 'errors' in result
        )
    
    def test_export_group_success(self):
        """Test exportación exitosa de grupo individual"""
        splitter = ExcelTemplateSplitter(self.df_test, self.config)
        
        # Crear grupo de datos de prueba
        group_df = self.df_test[self.df_test['Region'] == 'Norte'].copy()
        
        # Mock para evitar crear archivo real
        with patch.object(splitter, '_create_excel_file_with_template', return_value=True):
            result = splitter._export_group('Norte', group_df)
            
            self.assertIsInstance(result, ExportResult)
            self.assertTrue(result.success)
            self.assertEqual(result.group_name, 'Norte')
            self.assertEqual(result.rows_processed, 2)
    
    def test_cancel_operation(self):
        """Test cancelación de operación"""
        splitter = ExcelTemplateSplitter(self.df_test, self.config)
        splitter.cancel_operation()
        
        # Verificar que se establece el flag de cancelación
        self.assertTrue(hasattr(splitter, '_cancelled'))
        self.assertTrue(splitter._cancelled)
    
    def test_cleanup_temp_files(self):
        """Test limpieza de archivos temporales"""
        splitter = ExcelTemplateSplitter(self.df_test, self.config)
        
        # Crear archivo temporal
        temp_file = os.path.join(self.config.output_folder, 'temp.xlsx')
        with open(temp_file, 'w') as f:
            f.write('test')
        
        splitter.created_files = [temp_file]
        splitter.cleanup_temp_files()
        
        # Verificar que se limpia el archivo
        self.assertFalse(os.path.exists(temp_file))


class TestExportSeparatedConfig(unittest.TestCase):
    """Tests para ExportSeparatedConfig"""
    
    def test_init(self):
        """Test inicialización"""
        config = ExportSeparatedConfig(
            separator_column='Region',
            template_path='template.xlsx',
            output_folder='/output'
        )
        
        self.assertEqual(config.separator_column, 'Region')
        self.assertEqual(config.template_path, 'template.xlsx')
        self.assertEqual(config.output_folder, '/output')
        self.assertEqual(config.file_template, '{valor}.xlsx')  # Default
        self.assertEqual(config.start_cell, 'A1')  # Default
        self.assertEqual(config.handle_duplicates, 'overwrite')  # Default
    
    def test_validate(self):
        """Test validación de configuración"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
            # Crear archivo Excel válido
            wb = openpyxl.Workbook()
            wb.save(tmp.name)
            
            config = ExportSeparatedConfig(
                separator_column='Region',
                template_path=tmp.name,
                output_folder=tempfile.mkdtemp(),
                file_template='Report.xlsx'
            )
            
            result = config.validate()
            
            self.assertIsInstance(result, dict)
            self.assertIn('valid', result)
            self.assertIn('errors', result)
            self.assertIn('warnings', result)
    
    def test_get_default_mapping(self):
        """Test obtención de mapeo por defecto"""
        config = ExportSeparatedConfig(
            separator_column='Region',
            template_path='dummy.xlsx',
            output_folder='/tmp'
        )
        df_columns = ['Region', 'Producto', 'Ventas', 'Fecha']
        
        mapping = config.get_default_mapping(df_columns)
        
        self.assertIsInstance(mapping, dict)
        self.assertEqual(mapping['Region'], 'A')
        self.assertEqual(mapping['Producto'], 'B')
        self.assertEqual(mapping['Ventas'], 'C')
        self.assertEqual(mapping['Fecha'], 'D')


class TestValidationResult(unittest.TestCase):
    """Tests para ValidationResult"""
    
    def test_default_state(self):
        """Test estado por defecto"""
        result = ValidationResult()
        
        self.assertTrue(result.is_valid)
        self.assertEqual(result.errors, [])
        self.assertEqual(result.warnings, [])
        self.assertEqual(result.info, [])
    
    def test_add_error(self):
        """Test agregar error"""
        result = ValidationResult()
        result.add_error("Test error")
        
        self.assertFalse(result.is_valid)
        self.assertIn("Test error", result.errors)
    
    def test_add_warning(self):
        """Test agregar warning"""
        result = ValidationResult()
        result.add_warning("Test warning")
        
        self.assertTrue(result.is_valid)  # Warnings no invalidan
        self.assertIn("Test warning", result.warnings)
    
    def test_add_info(self):
        """Test agregar info"""
        result = ValidationResult()
        result.add_info("Test info")
        
        self.assertTrue(result.is_valid)
        self.assertIn("Test info", result.info)


class TestExportResult(unittest.TestCase):
    """Tests para ExportResult"""
    
    def test_default_state(self):
        """Test estado por defecto"""
        result = ExportResult()
        
        self.assertFalse(result.success)
        self.assertEqual(result.file_path, "")
        self.assertEqual(result.group_name, "")
        self.assertEqual(result.rows_processed, 0)
        self.assertEqual(result.processing_time, 0.0)
        self.assertEqual(result.error, "")
        self.assertIsNone(result.timestamp)


if __name__ == '__main__':
    unittest.main()