#!/usr/bin/env python3
"""
Tests de integración para el sistema de exportación separada completo
Prueba la integración entre main.py, diálogos y core logic
"""

import unittest
import tempfile
from pathlib import Path
import shutil
from unittest.mock import patch, MagicMock, Mock
import pandas as pd

# Configurar QApplication para tests
from PySide6.QtWidgets import QApplication
from PySide6.QtCore import Qt

from main import MainWindow
from core.data_handler import ExcelTemplateSplitter, ExportSeparatedConfig, exportar_datos_separados
from app.menus import MenuActions


class TestExportSeparatedIntegration(unittest.TestCase):
    """Tests de integración para sistema de exportación separada"""
    
    @classmethod
    def setUpClass(cls) -> None:
        """Configurar QApplication para tests de UI"""
        if not QApplication.instance():
            cls.app = QApplication([])
        else:
            cls.app = QApplication.instance()
    
    def setUp(self) -> None:
        """Configurar datos de prueba"""
        # Crear DataFrame de prueba
        self.df_test = pd.DataFrame({
            'Region': ['Norte', 'Norte', 'Sur', 'Sur', 'Este', 'Oeste'],
            'Producto': ['A', 'B', 'A', 'B', 'A', 'B'],
            'Ventas': [1000, 1500, 800, 1200, 900, 1100],
            'Fecha': pd.to_datetime(['2024-01-01', '2024-01-02', '2024-01-01', '2024-01-02', '2024-01-01', '2024-01-02'])
        })
        
        # Crear directorio temporal para tests
        self.temp_dir = tempfile.mkdtemp()
        self.excel_template_path = self.create_test_excel_template()
        
    def tearDown(self) -> None:
        """Limpiar después de tests"""
        if Path(self.temp_dir).exists():
            shutil.rmtree(self.temp_dir)
    
    def create_test_excel_template(self) -> pd.DataFrame:
        """Crear una plantilla Excel de prueba"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla"
        
        # Crear encabezados básicos
        ws['A1'] = "Región"
        ws['B1'] = "Producto"
        ws['C1'] = "Ventas"
        ws['D1'] = "Fecha"
        
        # Formato básico
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Guardar archivo temporal
        template_path = str(Path(self.temp_dir) / "plantilla_test.xlsx")
        wb.save(template_path)
        return template_path
    
    def test_main_window_integration(self) -> None:
        """Test integración básica con MainWindow"""
        window = MainWindow()
        
        # Simular carga de datos
        window.data_service.set_original_data(self.df_test.copy())
        
        # Habilitar acciones de datos via single source of truth
        MenuActions.enable_data_actions(True)
        
        self.assertTrue(MenuActions.exportar_separado_action.isEnabled())
        
        # Verificar que tiene el status tip correcto
        expected_tip = "Exportar datos separados por columna usando plantillas Excel"
        self.assertEqual(MenuActions.exportar_separado_action.statusTip(), expected_tip)
    
    def test_export_processing_integration(self) -> None:
        """Test procesamiento de exportación desde core"""
        window = MainWindow()
        
        # Crear configuración válida
        config = ExportSeparatedConfig(
            separator_column='Region',
            template_path=self.excel_template_path,
            output_folder=self.temp_dir,
            file_template='{valor}_{fecha}.xlsx',
            column_mapping={'Region': 'A', 'Producto': 'B', 'Ventas': 'C', 'Fecha': 'D'},
            start_cell='A2'
        )
        
        # Procesar exportación directamente via core
        try:
            result = exportar_datos_separados(self.df_test, config.to_dict())
            self.assertIsInstance(result, dict)
            self.assertIn('success', result)
        except Exception as e:
            self.fail(f"Error en procesamiento de exportación: {e}")
    
    def test_menu_integration_with_data(self) -> None:
        """Test integración del menú con datos cargados"""
        window = MainWindow()
        
        # Sin datos
        MenuActions.enable_data_actions(False)
        self.assertFalse(MenuActions.exportar_separado_action.isEnabled())
        
        # Con datos válidos
        window.data_service.set_original_data(self.df_test.copy())
        MenuActions.enable_data_actions(True)
        self.assertTrue(MenuActions.exportar_separado_action.isEnabled())
    
    def test_error_handling_integration(self) -> None:
        """Test manejo de errores en la integración"""
        window = MainWindow()
        
        # Configuración inválida
        config = ExportSeparatedConfig(
            separator_column='ColumnaInexistente',
            template_path='/ruta/inexistente.xlsx',
            output_folder='/carpeta/inexistente',
            file_template='{valor}.xlsx'
        )
        
        # Procesar exportación — no debe lanzar excepción no manejada
        try:
            result = exportar_datos_separados(self.df_test, config.to_dict())
            self.assertIsInstance(result, dict)
        except Exception as e:
            self.fail(f"Excepción no manejada en error handling: {e}")
    
    def test_complete_workflow_simulation(self) -> None:
        """Test simulación de workflow completo"""
        window = MainWindow()
        
        # Configuración válida
        config = ExportSeparatedConfig(
            separator_column='Region',
            template_path=self.excel_template_path,
            output_folder=self.temp_dir,
            file_template='Reporte_{valor}.xlsx',
            column_mapping={'Region': 'A', 'Producto': 'B', 'Ventas': 'C'},
            start_cell='A2'
        )
        
        # Procesar exportación directamente via core
        result = exportar_datos_separados(self.df_test, config.to_dict())
        
        # Verificar resultado
        self.assertIsInstance(result, dict)
        self.assertIn('success', result)
    
    def test_menu_action_connection(self) -> None:
        """Test que el menú está conectado correctamente"""
        window = MainWindow()
        
        # Verificar que la acción del menú existe
        action = MenuActions.exportar_separado_action
        self.assertIsNotNone(action)
        
        # Verificar que el método exportar_datos_separados existe en MainWindow
        self.assertTrue(hasattr(window, 'exportar_datos_separados'))
        self.assertTrue(callable(window.exportar_datos_separados))
    
    def test_progress_dialog_integration(self) -> None:
        """Test integración del diálogo de progreso"""
        window = MainWindow()
        
        # Configuración válida
        config = ExportSeparatedConfig(
            separator_column='Region',
            template_path=self.excel_template_path,
            output_folder=self.temp_dir,
            file_template='test.xlsx'
        )
        
        # Procesar exportación — verificar que no lanza excepciones
        try:
            result = exportar_datos_separados(self.df_test, config.to_dict())
            self.assertIsInstance(result, dict)
        except Exception as e:
            self.fail(f"Error en procesamiento de exportación: {e}")


class TestExcelTemplateSplitterIntegration(unittest.TestCase):
    """Tests de integración específicos para ExcelTemplateSplitter"""
    
    def setUp(self) -> None:
        """Configurar datos de prueba"""
        self.df_test = pd.DataFrame({
            'Category': ['A', 'A', 'B', 'B', 'C'],
            'Value': [10, 20, 30, 40, 50],
            'Name': ['Item1', 'Item2', 'Item3', 'Item4', 'Item5']
        })
        
        self.temp_dir = tempfile.mkdtemp()
        
    def tearDown(self) -> None:
        """Limpiar después de tests"""
        if Path(self.temp_dir).exists():
            shutil.rmtree(self.temp_dir)
    
    def test_splitter_with_config_integration(self) -> None:
        """Test integración ExcelTemplateSplitter con ExportSeparatedConfig"""
        # Crear plantilla Excel de prueba
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla"
        
        # Crear encabezados básicos que coincidan con el DataFrame
        ws['A1'] = "Category"
        ws['B1'] = "Value"
        ws['C1'] = "Name"
        
        template_path = str(Path(self.temp_dir) / "plantilla_category.xlsx")
        wb.save(template_path)
        
        # Configuración completa y válida
        config = ExportSeparatedConfig(
            separator_column='Category',
            template_path=template_path,
            output_folder=self.temp_dir,
            file_template='{valor}.xlsx',
            column_mapping={'Category': 'A', 'Value': 'B', 'Name': 'C'},
            start_cell='A2'
        )
        
        # Crear splitter
        splitter = ExcelTemplateSplitter(self.df_test, config)
        
        # Verificar que se inicializa correctamente
        self.assertEqual(len(splitter.df), 5)
        self.assertEqual(splitter.config.separator_column, 'Category')
        
        # Verificar validación con configuración completa
        validation = splitter.validate_configuration()
        # Con una configuración válida, debería pasar la validación básica
        self.assertIsInstance(validation.is_valid, bool)
        # Si hay errores, deberían estar relacionados con aspectos específicos del sistema, no con parámetros faltantes
        self.assertIsInstance(validation.errors, list)


if __name__ == '__main__':
    unittest.main()