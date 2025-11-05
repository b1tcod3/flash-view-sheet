import unittest
from unittest.mock import Mock, patch, MagicMock
import pandas as pd
import numpy as np
import tempfile
import os
import shutil
import sys
import time
import gc
import tracemalloc
import resource
# import psutil  # Opcional - solo si está disponible (import condicional más abajo)
from datetime import datetime
import threading
import concurrent.futures
from contextlib import contextmanager

# Verificar si psutil está disponible
PSUTIL_AVAILABLE = False
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    print("Warning: psutil no está disponible. Usando medición alternativa de memoria.")

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.data_handler import (
    ExcelTemplateSplitter,
    ExportSeparatedConfig,
    exportar_datos_separados,
    OptimizationConfig
)
from core.performance_optimizer import PerformanceOptimizer, ChunkingStrategy


class PerformanceMetrics:
    """Clase para almacenar y analizar métricas de rendimiento"""
    def __init__(self):
        self.measurements = []
        
    def add_measurement(self, name: str, duration: float, memory_peak_mb: float, 
                       rows_processed: int, groups_processed: int, 
                       file_size_mb: float = 0.0, success: bool = True):
        """Agregar una medición de rendimiento"""
        self.measurements.append({
            'name': name,
            'duration_seconds': duration,
            'memory_peak_mb': memory_peak_mb,
            'rows_processed': rows_processed,
            'groups_processed': groups_processed,
            'file_size_mb': file_size_mb,
            'success': success,
            'throughput_rows_per_second': rows_processed / duration if duration > 0 else 0,
            'memory_per_row_kb': (memory_peak_mb * 1024) / rows_processed if rows_processed > 0 else 0,
            'timestamp': datetime.now()
        })
    
    def get_summary(self) -> dict:
        """Obtener resumen de todas las mediciones"""
        if not self.measurements:
            return {}
        
        successful = [m for m in self.measurements if m['success']]
        failed = [m for m in self.measurements if not m['success']]
        
        return {
            'total_tests': len(self.measurements),
            'successful_tests': len(successful),
            'failed_tests': len(failed),
            'avg_duration_seconds': np.mean([m['duration_seconds'] for m in successful]) if successful else 0,
            'avg_memory_peak_mb': np.mean([m['memory_peak_mb'] for m in successful]) if successful else 0,
            'avg_throughput_rows_per_second': np.mean([m['throughput_rows_per_second'] for m in successful]) if successful else 0,
            'max_memory_peak_mb': max([m['memory_peak_mb'] for m in successful]) if successful else 0,
            'min_throughput_rows_per_second': min([m['throughput_rows_per_second'] for m in successful]) if successful else 0,
            'max_throughput_rows_per_second': max([m['throughput_rows_per_second'] for m in successful]) if successful else 0
        }
    
    def generate_report(self) -> str:
        """Generar reporte de rendimiento en texto"""
        summary = self.get_summary()
        
        report = f"""
        === REPORTE DE RENDIMIENTO ===
        Total de pruebas: {summary['total_tests']}
        Pruebas exitosas: {summary['successful_tests']}
        Pruebas fallidas: {summary['failed_tests']}
        
        MÉTRICAS PROMEDIO:
        - Duración: {summary['avg_duration_seconds']:.2f} segundos
        - Memoria pico: {summary['avg_memory_peak_mb']:.2f} MB
        - Rendimiento: {summary['avg_throughput_rows_per_second']:.0f} filas/segundo
        
        MÉTRICAS EXTREMAS:
        - Memoria máxima: {summary['max_memory_peak_mb']:.2f} MB
        - Rendimiento mínimo: {summary['min_throughput_rows_per_second']:.0f} filas/segundo
        - Rendimiento máximo: {summary['max_throughput_rows_per_second']:.0f} filas/segundo
        
        DETALLES POR PRUEBA:
        """
        
        for i, measurement in enumerate(self.measurements, 1):
            status = "✓" if measurement['success'] else "✗"
            report += f"""
        {i}. {measurement['name']} {status}
           - Duración: {measurement['duration_seconds']:.2f}s
           - Memoria: {measurement['memory_peak_mb']:.2f}MB
           - Filas: {measurement['rows_processed']:,}
           - Grupos: {measurement['groups_processed']}
           - Rendimiento: {measurement['throughput_rows_per_second']:.0f} filas/s
           - Memoria/fila: {measurement['memory_per_row_kb']:.2f}KB
        """
        
        return report


class MemoryMonitor:
    """Monitor de memoria que funciona con o sin psutil"""
    
    @staticmethod
    def get_memory_mb():
        """Obtener uso actual de memoria en MB"""
        if PSUTIL_AVAILABLE:
            return psutil.Process().memory_info().rss / 1024 / 1024
        else:
            # Usar resource como fallback
            try:
                usage = resource.getrusage(resource.RUSAGE_SELF)
                # En Linux, ru_maxrss está en KB
                return usage.ru_maxrss / 1024
            except (AttributeError, OSError):
                # Fallback a tracemalloc si está disponible
                return 0.0
    
    @staticmethod
    def get_peak_memory_mb():
        """Obtener memoria pico usando tracemalloc"""
        if tracemalloc.is_tracing():
            current, peak = tracemalloc.get_traced_memory()
            return peak / 1024 / 1024
        return 0.0


@contextmanager
def measure_performance(test_name: str, metrics: PerformanceMetrics):
    """Context manager para medir rendimiento de una operación"""
    # Iniciar medición de memoria
    tracemalloc.start()
    start_memory = MemoryMonitor.get_memory_mb()
    start_time = time.time()
    
    try:
        yield
    finally:
        # Finalizar medición
        end_time = time.time()
        end_memory = MemoryMonitor.get_memory_mb()
        
        # Obtener peak de memoria de tracemalloc
        current, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()
        
        duration = end_time - start_time
        memory_peak = max(end_memory, peak / 1024 / 1024)
        
        # Nota: rows_processed y otros valores se agregarán después
        metrics.measurements.append({
            'name': test_name,
            'duration_seconds': duration,
            'memory_peak_mb': memory_peak,
            'rows_processed': 0,  # Se actualizará después
            'groups_processed': 0,  # Se actualizará después
            'file_size_mb': 0.0,  # Se actualizará después
            'success': True,  # Se actualizará después
            'throughput_rows_per_second': 0,  # Se calculará después
            'memory_per_row_kb': 0,  # Se calculará después
            'timestamp': datetime.now()
        })


class TestExportSeparatedPerformance(unittest.TestCase):
    """
    Suite completa de pruebas de rendimiento para exportación separada
    Incluye benchmarks, medición de memoria, tests de optimización y stress testing
    """
    
    def setUp(self):
        """Configurar entorno de pruebas de rendimiento"""
        # Crear directorio temporal para pruebas
        self.temp_dir = tempfile.mkdtemp(prefix='performance_test_')
        self.output_dir = os.path.join(self.temp_dir, 'output')
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Crear plantilla base para pruebas
        self.template_path = self.create_performance_template()
        
        # Inicializar métricas
        self.metrics = PerformanceMetrics()
        
        # Configuraciones de prueba
        self.small_config = {
            'separator_column': 'Grupo',
            'template_path': self.template_path,
            'start_cell': 'A2',
            'output_folder': self.output_dir,
            'file_template': 'perf_test_{valor}.xlsx',
            'enable_chunking': False,
            'max_memory_mb': 512
        }
        
        self.large_config = {
            'separator_column': 'Grupo',
            'template_path': self.template_path,
            'start_cell': 'A2',
            'output_folder': self.output_dir,
            'file_template': 'perf_test_{valor}.xlsx',
            'enable_chunking': True,
            'max_memory_mb': 256  # Límite bajo para forzar optimización
        }
    
    def create_performance_template(self) -> str:
        """Crear plantilla optimizada para pruebas de rendimiento"""
        template_path = os.path.join(self.temp_dir, 'performance_template.xlsx')
        
        # Importar openpyxl aquí para evitar problemas de dependencias
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl no está disponible para crear plantilla de rendimiento")
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Datos'
        
        # Encabezados simples para máximo rendimiento
        headers = ['Grupo', 'Valor1', 'Valor2', 'Valor3', 'Valor4', 'Valor5']
        for i, header in enumerate(headers, 1):
            sheet.cell(row=1, column=i, value=header)
        
        workbook.save(template_path)
        return template_path
    
    def create_test_dataframe(self, num_rows: int, num_groups: int = None, 
                            include_nulls: bool = False) -> pd.DataFrame:
        """Crear DataFrame de prueba con tamaño controlado"""
        if num_groups is None:
            num_groups = min(num_rows // 100, 50)  # Grupos proporcionales pero limitados
        
        # Generar datos de grupo
        groups = [f'Grupo_{i}' for i in range(num_groups)]
        group_column = [groups[i % num_groups] for i in range(num_rows)]
        
        # Agregar valores nulos si se requiere
        if include_nulls:
            null_indices = np.random.choice(num_rows, size=num_rows // 20, replace=False)
            for idx in null_indices:
                group_column[idx] = None
        
        # Crear DataFrame
        df = pd.DataFrame({
            'Grupo': group_column,
            'Valor1': np.random.randint(1, 1000, num_rows),
            'Valor2': np.random.uniform(0, 100, num_rows),
            'Valor3': [f'Texto_{i}' for i in range(num_rows)],
            'Valor4': pd.date_range('2020-01-01', periods=num_rows, freq='H'),
            'Valor5': np.random.choice(['A', 'B', 'C', 'D'], num_rows)
        })
        
        return df
    
    def measure_export_performance(self, df: pd.DataFrame, config: dict, 
                                 test_name: str) -> dict:
        """Medir rendimiento de una exportación específica"""
        with measure_performance(test_name, self.metrics) as measure_context:
            try:
                # Ejecutar exportación
                result = exportar_datos_separados(df, config)
                
                # Calcular tamaño total de archivos generados
                total_file_size = 0
                if result['success'] and result['files_created']:
                    for file_path in result['files_created']:
                        if os.path.exists(file_path):
                            total_file_size += os.path.getsize(file_path) / 1024 / 1024  # MB
                
                # Actualizar la última medición con datos completos
                if self.metrics.measurements:
                    measurement = self.metrics.measurements[-1]
                    measurement.update({
                        'rows_processed': result.get('total_rows', 0),
                        'groups_processed': result.get('groups_processed', 0),
                        'file_size_mb': total_file_size,
                        'success': result.get('success', False)
                    })
                    
                    # Recalcular métricas derivadas
                    if measurement['rows_processed'] > 0 and measurement['duration_seconds'] > 0:
                        measurement['throughput_rows_per_second'] = (
                            measurement['rows_processed'] / measurement['duration_seconds']
                        )
                        measurement['memory_per_row_kb'] = (
                            measurement['memory_peak_mb'] * 1024 / measurement['rows_processed']
                        )
                
                return result
                
            except Exception as e:
                # Marcar como fallido si hay excepción
                if self.metrics.measurements:
                    self.metrics.measurements[-1]['success'] = False
                raise
    
    def test_small_dataset_performance(self):
        """Test de rendimiento con dataset pequeño (< 1K filas)"""
        print("\n=== Test Dataset Pequeño ===")
        
        df_small = self.create_test_dataframe(500, num_groups=5)
        result = self.measure_export_performance(df_small, self.small_config, 
                                               "Dataset Pequeño (500 filas)")
        
        # Verificar que el resultado es exitoso
        self.assertTrue(result['success'])
        self.assertEqual(result['groups_processed'], 5)
        self.assertEqual(result['total_rows'], 500)
        
        # Verificar métricas de rendimiento
        latest = self.metrics.measurements[-1]
        self.assertLess(latest['duration_seconds'], 10)  # Menos de 10 segundos
        self.assertGreater(latest['throughput_rows_per_second'], 50)  # Al menos 50 filas/s
        
        print(f"✓ Pequeño dataset: {latest['duration_seconds']:.2f}s, "
              f"{latest['throughput_rows_per_second']:.0f} filas/s")
    
    def test_medium_dataset_performance(self):
        """Test de rendimiento con dataset mediano (1K-50K filas)"""
        print("\n=== Test Dataset Mediano ===")
        
        df_medium = self.create_test_dataframe(5000, num_groups=10)  # Reducido para speed
        result = self.measure_export_performance(df_medium, self.large_config, 
                                               "Dataset Mediano (5K filas)")
        
        # Verificar resultado
        self.assertTrue(result['success'])
        self.assertEqual(result['groups_processed'], 10)
        self.assertEqual(result['total_rows'], 5000)
        
        # Verificar métricas
        latest = self.metrics.measurements[-1]
        self.assertLess(latest['duration_seconds'], 60)  # Menos de 1 minuto
        self.assertGreater(latest['throughput_rows_per_second'], 50)  # Al menos 50 filas/s
        
        print(f"✓ Dataset mediano: {latest['duration_seconds']:.2f}s, "
              f"{latest['throughput_rows_per_second']:.0f} filas/s")
    
    def test_memory_usage_optimization(self):
        """Test de optimización de uso de memoria"""
        print("\n=== Test Optimización de Memoria ===")
        
        # Dataset diseñado para probar gestión de memoria
        df_memory_test = self.create_test_dataframe(10000, num_groups=10)
        
        # Configuración con límite de memoria muy bajo
        memory_config = self.large_config.copy()
        memory_config['max_memory_mb'] = 50  # Límite muy bajo
        
        # Ejecutar exportación
        start_memory = MemoryMonitor.get_memory_mb()
        result = self.measure_export_performance(df_memory_test, memory_config, 
                                               "Test Memoria (10K filas)")
        end_memory = MemoryMonitor.get_memory_mb()
        
        # Verificar resultado
        self.assertTrue(result['success'])
        
        # Verificar que la memoria se mantuvo dentro de límites razonables
        memory_increase = end_memory - start_memory
        self.assertLess(memory_increase, 200)  # Aumento de menos de 200MB
        
        latest = self.metrics.measurements[-1]
        print(f"✓ Uso de memoria: {latest['memory_peak_mb']:.1f}MB pico, "
              f"{memory_increase:.1f}MB aumento")
    
    def test_chunking_performance_impact(self):
        """Test del impacto del chunking en el rendimiento"""
        print("\n=== Test Impacto del Chunking ===")
        
        df_chunk_test = self.create_test_dataframe(8000, num_groups=8)  # Dataset moderado
        
        # Test sin chunking
        config_no_chunk = self.small_config.copy()
        config_no_chunk['enable_chunking'] = False
        
        result_no_chunk = self.measure_export_performance(df_chunk_test, config_no_chunk, 
                                                        "Sin Chunking (8K filas)")
        
        # Test con chunking
        config_with_chunk = self.large_config.copy()
        config_with_chunk['enable_chunking'] = True
        config_with_chunk['max_memory_mb'] = 100
        
        result_with_chunk = self.measure_export_performance(df_chunk_test, config_with_chunk, 
                                                          "Con Chunking (8K filas)")
        
        # Verificar ambos resultados
        self.assertTrue(result_no_chunk['success'])
        self.assertTrue(result_with_chunk['success'])
        
        # Obtener métricas
        no_chunk_metrics = self.metrics.measurements[-2]  # Penúltima medición
        with_chunk_metrics = self.metrics.measurements[-1]  # Última medición
        
        print(f"✓ Sin chunking: {no_chunk_metrics['duration_seconds']:.2f}s, "
              f"{no_chunk_metrics['memory_peak_mb']:.1f}MB")
        print(f"✓ Con chunking: {with_chunk_metrics['duration_seconds']:.2f}s, "
              f"{with_chunk_metrics['memory_peak_mb']:.1f}MB")
        
        # El chunking debería usar menos memoria (puede ser más lento)
        self.assertLess(with_chunk_metrics['memory_peak_mb'], 
                       no_chunk_metrics['memory_peak_mb'] * 1.2)  # Max 20% más memoria
    
    def test_stress_test_extreme_conditions(self):
        """Test de stress con condiciones extremas"""
        print("\n=== Test de Stress - Condiciones Extremas ===")
        
        # Dataset extremo con muchos grupos pequeños
        df_stress = self.create_test_dataframe(5000, num_groups=100)  # Muchos grupos pequeños
        
        # Configuración agresiva
        stress_config = self.large_config.copy()
        stress_config['max_memory_mb'] = 100
        stress_config['enable_chunking'] = True
        
        start_time = time.time()
        result = self.measure_export_performance(df_stress, stress_config, 
                                               "Stress Test (5K filas, 100 grupos)")
        total_time = time.time() - start_time
        
        # Verificar que survived el stress test
        if result['success']:
            latest = self.metrics.measurements[-1]
            print(f"✓ Stress test sobrevivido: {latest['duration_seconds']:.2f}s, "
                  f"{latest['memory_peak_mb']:.1f}MB, {latest['groups_processed']} grupos")
            
            # Verificar que procesó la mayoría de grupos
            self.assertGreater(latest['groups_processed'], 80)  # Al menos 80% de grupos
        else:
            print("⚠ Stress test falló (esperado en condiciones extremas)")
            
            # Aún así, debería haber procesado algunos grupos
            self.assertGreater(result.get('groups_processed', 0), 0)
    
    def tearDown(self):
        """Limpiar después de las pruebas"""
        # Limpiar archivos generados
        try:
            if os.path.exists(self.output_dir):
                for file in os.listdir(self.output_dir):
                    file_path = os.path.join(self.output_dir, file)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
            
            # Limpiar directorio temporal
            if os.path.exists(self.temp_dir):
                shutil.rmtree(self.temp_dir)
        except Exception as e:
            print(f"Error limpiando archivos temporales: {str(e)}")
        
        # Forzar garbage collection
        gc.collect()


class TestPerformanceBenchmarks(unittest.TestCase):
    """
    Tests de benchmark específicos para validar rendimiento contra estándares
    """
    
    def setUp(self):
        """Configurar datos de benchmark"""
        self.temp_dir = tempfile.mkdtemp(prefix='benchmark_test_')
        self.output_dir = os.path.join(self.temp_dir, 'output')
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Crear plantilla simple para benchmarks
        self.template_path = os.path.join(self.temp_dir, 'benchmark_template.xlsx')
        try:
            import openpyxl
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Datos'
            
            headers = ['Grupo', 'Valor']
            for i, header in enumerate(headers, 1):
                sheet.cell(row=1, column=i, value=header)
            
            workbook.save(self.template_path)
        except ImportError:
            self.skipTest("openpyxl no disponible para benchmarks")
    
    def test_benchmark_small_dataset(self):
        """Benchmark para dataset pequeño (referencia estándar)"""
        # Crear dataset pequeño estándar
        df = pd.DataFrame({
            'Grupo': ['A', 'B', 'C'] * 100,
            'Valor': list(range(300))
        })
        
        config = {
            'separator_column': 'Grupo',
            'template_path': self.template_path,
            'start_cell': 'A2',
            'output_folder': self.output_dir,
            'file_template': 'benchmark_{valor}.xlsx',
            'enable_chunking': False
        }
        
        start_time = time.time()
        result = exportar_datos_separados(df, config)
        elapsed_time = time.time() - start_time
        
        # Benchmark: debería completarse en menos de 5 segundos
        self.assertTrue(result['success'])
        self.assertLess(elapsed_time, 5.0, f"Benchmark pequeño falló: {elapsed_time:.2f}s > 5s")
        
        print(f"✓ Benchmark pequeño: {elapsed_time:.2f}s (objetivo: < 5s)")
    
    def test_benchmark_memory_efficiency(self):
        """Benchmark de eficiencia de memoria"""
        # Dataset mediano
        df = pd.DataFrame({
            'Grupo': [f'Grupo_{i%10}' for i in range(5000)],
            'Valor': list(range(5000))
        })
        
        # Medir memoria antes
        start_memory = MemoryMonitor.get_memory_mb()
        
        config = {
            'separator_column': 'Grupo',
            'template_path': self.template_path,
            'start_cell': 'A2',
            'output_folder': self.output_dir,
            'file_template': 'memory_{valor}.xlsx',
            'enable_chunking': True,
            'max_memory_mb': 100
        }
        
        result = exportar_datos_separados(df, config)
        
        # Medir memoria después
        end_memory = MemoryMonitor.get_memory_mb()
        memory_increase = end_memory - start_memory
        
        # Benchmark: aumento de memoria < 100MB para 5K filas
        self.assertTrue(result['success'])
        self.assertLess(memory_increase, 100, f"Uso de memoria alto: {memory_increase:.1f}MB")
        
        print(f"✓ Benchmark memoria: {memory_increase:.1f}MB aumento (objetivo: < 100MB)")
    
    def tearDown(self):
        """Limpiar después de benchmarks"""
        try:
            if os.path.exists(self.temp_dir):
                shutil.rmtree(self.temp_dir)
        except Exception as e:
            print(f"Error limpiando benchmark: {str(e)}")


if __name__ == '__main__':
    print("=== SUITE DE PRUEBAS DE RENDIMIENTO ===")
    print("Ejecutando tests de rendimiento para exportación separada...")
    
    # Ejecutar tests de rendimiento
    unittest.main(verbosity=2)
    
    # Mostrar resumen final
    print("\n=== RESUMEN DE PRUEBAS DE RENDIMIENTO ===")
    print("Todas las pruebas de rendimiento han sido ejecutadas.")
    print("Revise los logs anteriores para métricas detalladas.")