import unittest
from unittest.mock import Mock, patch, mock_open, MagicMock
import pandas as pd
import numpy as np
import tempfile
import os
import shutil
import sys
import time
import json
import hashlib
import re
import openpyxl
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.data_handler import (
    ExcelTemplateSplitter,
    ExportSeparatedConfig,
    exportar_datos_separados,
    ValidationResult,
    ExportResult,
    SeparationError,
    TemplateError,
    ConfigurationError
)


class TestExportSeparatedIntegrationExpanded(unittest.TestCase):
    """
    Suite expandida de tests de integración para la funcionalidad de exportación separada
    Se enfoca en casos especiales, escenarios avanzados y validación robusta
    """
    def setUp(self):
        """Configurar datos de prueba y entorno"""
        # Crear DataFrame de prueba más complejo
        self.df = pd.DataFrame({
            'Región': ['Norte', 'Sur', 'Este', 'Oeste', 'Norte', 'Sur', None],
            'Producto': ['A', 'B', 'C', 'D', 'E', 'F', 'G'],
            'Ventas': [100, 200, 300, 400, 500, 600, 700],
            'Fecha': pd.date_range('2025-01-01', periods=7, freq='D'),
            'Descuento': [0.1, 0.15, 0.2, 0.1, 0.12, 0.18, 0.2]
        })
        
        # Crear archivos temporales
        self.temp_dir = tempfile.mkdtemp()
        self.template_path = os.path.join(self.temp_dir, 'template.xlsx')
        self.create_test_template()
        
        # Crear directorio de salida
        self.output_dir = os.path.join(self.temp_dir, 'output')
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Configuración de prueba por defecto
        self.config = {
            'separator_column': 'Región',
            'template_path': self.template_path,
            'start_cell': 'A5',
            'output_folder': self.output_dir,
            'file_template': '{valor}_{fecha}.xlsx',
            'column_mapping': {},
            'enable_chunking': False
        }
    
    def create_test_template(self, template_path=None):
        """Crear plantilla Excel de prueba"""
        if template_path is None:
            template_path = self.template_path
            
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Datos'
        
        # Encabezados en fila 1
        sheet['A1'] = 'Fecha'
        sheet['B1'] = 'Región'
        sheet['C1'] = 'Producto'
        sheet['D1'] = 'Ventas'
        sheet['E1'] = 'Descuento'
        
        # Datos de ejemplo en filas 2-4
        sheet['A2'] = '2025-01-01'
        sheet['B2'] = 'Norte'
        sheet['C2'] = 'A'
        sheet['D2'] = 100
        sheet['E2'] = 0.1
        
        sheet['A3'] = '2025-01-02'
        sheet['B3'] = 'Sur'
        sheet['C3'] = 'B'
        sheet['D3'] = 200
        sheet['E3'] = 0.15
        
        sheet['A4'] = '2025-01-03'
        sheet['B4'] = 'Este'
        sheet['C4'] = 'C'
        sheet['D4'] = 300
        sheet['E4'] = 0.2
        
        # Aplicar formato
        for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
            sheet[col].font = openpyxl.styles.Font(bold=True)
            sheet[col].fill = openpyxl.styles.PatternFill(
                start_color='CCCCCC', end_color='CCCCCC', fill_type='solid'
            )
        
        workbook.save(template_path)
    
    def tearDown(self):
        """Limpiar archivos temporales"""
        try:
            shutil.rmtree(self.temp_dir)
        except Exception as e:
            print(f"Error limpiando directorio temporal: {str(e)}")
    
    def test_export_with_unicode_and_special_chars(self):
        """Probar exportación con caracteres especiales y Unicode"""
        # DataFrame con caracteres especiales
        df_special = pd.DataFrame({
            'Región': ['Ñorte', 'Súr', 'Esté', 'Öeste', 'Nor*te', 'Súr'],
            'Producto': ['Á', 'B', 'C', 'D', 'E', 'F'],
            'Ventas': [100, 200, 300, 400, 500, 600]
        })
        
        # Configuración con caracteres especiales
        config = self.config.copy()
        config['file_template'] = 'Reporte_{valor}_{fecha}.xlsx'
        
        # Ejecutar exportación
        result = exportar_datos_separados(df_special, config)
        
        # Verificar resultado
        self.assertTrue(result['success'])
        
        # Verificar que los archivos se crearon con nombres correctos
        for file_path in result['files_created']:
            self.assertTrue(os.path.exists(file_path))
            
            # Verificar que el nombre del archivo está sanitizado
            filename = os.path.basename(file_path)
            self.assertNotIn('<', filename)
            self.assertNotIn('>', filename)
            self.assertNotIn(':', filename)
            self.assertNotIn('"', filename)
            self.assertNotIn('/', filename)
            self.assertNotIn('\\', filename)
            self.assertNotIn('|', filename)
            self.assertNotIn('?', filename)
            self.assertNotIn('*', filename)
            
            # Verificar contenido del archivo
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # Verificar que los datos Unicode se mantienen
            self.assertEqual(sheet['B5'].value, 'Ñorte')  # Primera fila de datos
            
            workbook.close()
    
    def test_export_with_dataframe_many_columns(self):
        """Probar exportación con DataFrame con muchas columnas"""
        # Crear DataFrame con muchas columnas
        num_cols = 50
        col_names = [f'Columna_{i}' for i in range(num_cols)]
        
        df_many_cols = pd.DataFrame({
            col: [f'valor_{i}' for i in range(10)] 
            for col in col_names
        })
        
        # Crear nueva columna de separación con pocos valores únicos
        df_many_cols['Grupo'] = [f'Grupo_{i%3}' for i in range(10)]
        
        # Crear plantilla adecuada
        many_cols_template_path = os.path.join(self.temp_dir, 'many_cols_template.xlsx')
        self.create_template_with_columns(many_cols_template_path, num_cols)
        
        # Configuración
        config = self.config.copy()
        config['separator_column'] = 'Grupo'
        config['template_path'] = many_cols_template_path
        config['start_cell'] = 'A1'
        config['file_template'] = '{valor}.xlsx'
        
        # Ejecutar exportación
        result = exportar_datos_separados(df_many_cols, config)
        
        # Verificar resultado
        self.assertTrue(result['success'])
        self.assertEqual(result['groups_processed'], 3)  # 3 grupos únicos
        
        # Verificar archivos generados
        self.assertEqual(len(result['files_created']), 3)
        
        for file_path in result['files_created']:
            self.assertTrue(os.path.exists(file_path))
            
            # Verificar contenido del archivo
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # Verificar que todas las columnas están presentes
            for i in range(num_cols):
                expected_value = f'valor_{i}'
                # Los datos comienzan en fila 1, así que valor_0 está en columna A, fila 1
                # o en la columna correspondiente mapeada
                self.assertIn(expected_value, [str(cell.value) for cell in sheet[1]])
            
            workbook.close()
    
    def create_template_with_columns(self, template_path, num_cols):
        """Crear plantilla con muchas columnas"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Datos'
        
        # Agregar encabezados para cada columna
        for i in range(num_cols):
            col_letter = openpyxl.utils.get_column_letter(i+1)
            sheet[f'{col_letter}1'] = f'Columna_{i}'
        
        workbook.save(template_path)
    
    def test_export_with_existing_files_and_conflicts(self):
        """Probar exportación con conflictos de nombres de archivos existentes"""
        # Crear algunos archivos existentes en el directorio de salida
        existing_files = ['Norte_2025-11-05.xlsx', 'Sur_2025-11-05.xlsx']
        for filename in existing_files:
            file_path = os.path.join(self.output_dir, filename)
            with open(file_path, 'w') as f:
                f.write(f'Archivo existente: {filename}')
        
        # Ejecutar exportación
        result = exportar_datos_separados(self.df, self.config)
        
        # Verificar resultado
        self.assertTrue(result['success'])
        
        # Verificar que se crearon archivos con numeración
        created_filenames = [os.path.basename(f) for f in result['files_created']]
        
        # Debería haber archivos con numeración para resolver conflictos
        self.assertTrue(any('Norte_' in f for f in created_filenames))
        self.assertTrue(any('Sur_' in f for f in created_filenames))
        
        # Los archivos originales deberían mantenerse
        for filename in existing_files:
            original_path = os.path.join(self.output_dir, filename)
            self.assertTrue(os.path.exists(original_path))
    
    def test_export_with_invalid_template_handling(self):
        """Probar manejo de plantillas inválidas o corruptas"""
        # Crear plantilla inválida
        invalid_template_path = os.path.join(self.temp_dir, 'invalid_template.xlsx')
        with open(invalid_template_path, 'w') as f:
            f.write('Este no es un archivo Excel válido')
        
        # Configuración con plantilla inválida
        config = self.config.copy()
        config['template_path'] = invalid_template_path
        
        # Ejecutar exportación
        result = exportar_datos_separados(self.df, config)
        
        # Verificar que el error se maneja correctamente
        self.assertFalse(result['success'])
        self.assertGreater(len(result['errors']), 0)
        self.assertTrue(any('Excel' in error or 'plantilla' in error.lower() for error in result['errors']))
    
    def test_export_with_compression_and_large_files(self):
        """Probar exportación con compresión y archivos grandes"""
        # Crear DataFrame grande
        num_rows = 10000
        df_large = pd.DataFrame({
            'Región': ['Norte', 'Sur', 'Este'] * (num_rows // 3),
            'Producto': [f'Producto_{i%100}' for i in range(num_rows)],
            'Ventas': [i * 10 for i in range(num_rows)]
        })
        
        # Configuración con chunking habilitado
        config = self.config.copy()
        config['enable_chunking'] = True
        config['max_memory_mb'] = 10  # Límite bajo para forzar chunking
        
        start_time = time.time()
        
        # Ejecutar exportación
        result = exportar_datos_separados(df_large, config)
        
        elapsed_time = time.time() - start_time
        
        # Verificar resultado
        self.assertTrue(result['success'])
        self.assertEqual(result['groups_processed'], 3)  # 3 regiones
        self.assertEqual(result['total_rows'], num_rows)
        
        # Verificar rendimiento
        print(f"Tiempo de procesamiento: {elapsed_time:.2f} segundos para {num_rows} filas")
        self.assertLess(elapsed_time, 120)  # Menos de 2 minutos para 10K filas
        
        # Verificar archivos generados
        for file_path in result['files_created']:
            self.assertTrue(os.path.exists(file_path))
            
            # Verificar que el archivo tiene un tamaño considerable
            file_size_kb = os.path.getsize(file_path) / 1024
            self.assertGreater(file_size_kb, 100)  # Al menos 100 KB
    
    def test_export_with_cancellation_and_progress_tracking(self):
        """Probar cancelación de operación y seguimiento de progreso"""
        # Crear DataFrame con muchos grupos
        num_groups = 50
        df_many_groups = pd.DataFrame({
            'Región': [f'Región_{i}' for i in range(num_groups)],
            'Producto': [f'Producto_{i%10}' for i in range(num_groups)],
            'Ventas': [i * 100 for i in range(num_groups)]
        })
        
        # Crear splitter
        config = ExportSeparatedConfig(**self.config)
        splitter = ExcelTemplateSplitter(df_many_groups, config)
        
        # Configurar callback que cancelará la operación después de unos pocos grupos
        cancellation_requested = False
        progress_updates = []
        
        def progress_callback(current, total):
            progress_updates.append((current, total))
            # Solicitar cancelación después de 5 grupos
            if current >= 5 and not cancellation_requested:
                splitter.cancel_operation()
                nonlocal cancellation_requested
                cancellation_requested = True
        
        config.progress_callback = progress_callback
        
        # Ejecutar separación
        start_time = time.time()
        result = splitter.separate_and_export()
        elapsed_time = time.time() - start_time
        
        # Verificar que la operación se canceló parcialmente
        self.assertTrue(result['success'])  # Algunos grupos se procesaron
        self.assertGreater(result['groups_processed'], 0)
        self.assertLess(result['groups_processed'], num_groups)  # No todos los grupos
        self.assertGreater(len(progress_updates), 0)
        
        # Verificar que hay seguimiento de progreso
        self.assertIn('analysis', result)
        
        print(f"Procesamiento cancelado: {result['groups_processed']}/{num_groups} grupos")
        print(f"Tiempo transcurrido: {elapsed_time:.2f} segundos")
        
        # Limpiar archivos creados durante el test
        for file_path in result['files_created']:
            if os.path.exists(file_path):
                os.remove(file_path)
    
    def test_export_with_different_worksheet_templates(self):
        """Probar exportación con plantillas que tienen múltiples hojas"""
        # Crear plantilla con múltiples hojas
        multi_sheet_template_path = os.path.join(self.temp_dir, 'multi_sheet_template.xlsx')
        
        workbook = openpyxl.Workbook()
        
        # Hoja principal
        main_sheet = workbook.active
        main_sheet.title = 'Datos'
        
        # Encabezados en fila 1
        main_sheet['A1'] = 'Fecha'
        main_sheet['B1'] = 'Región'
        main_sheet['C1'] = 'Producto'
        main_sheet['D1'] = 'Ventas'
        
        # Aplicar formato
        for col in ['A1', 'B1', 'C1', 'D1']:
            main_sheet[col].font = openpyxl.styles.Font(bold=True)
            main_sheet[col].fill = openpyxl.styles.PatternFill(
                start_color='DDDDDD', end_color='DDDDDD', fill_type='solid'
            )
        
        # Hoja secundaria
        summary_sheet = workbook.create_sheet('Resumen')
        summary_sheet['A1'] = 'Región'
        summary_sheet['B1'] = 'Total Ventas'
        
        # Agregar datos de resumen
        summary_sheet['A2'] = 'Norte'
        summary_sheet['B2'] = '=SUMIF(Datos.B:B,A2,Datos.D:D)'
        
        workbook.save(multi_sheet_template_path)
        
        # Configuración con hoja específica
        config = self.config.copy()
        config['template_path'] = multi_sheet_template_path
        config['start_cell'] = 'A2'  # Empezar en fila 2 para preservar encabezados
        
        # Ejecutar exportación
        result = exportar_datos_separados(self.df, config)
        
        # Verificar resultado
        self.assertTrue(result['success'])
        self.assertGreater(len(result['files_created']), 0)
        
        # Verificar que se preservó la hoja secundaria
        for file_path in result['files_created']:
            self.assertTrue(os.path.exists(file_path))
            
            workbook = openpyxl.load_workbook(file_path)
            sheet_names = workbook.sheetnames
            
            # Debería tener tanto la hoja de datos como la hoja de resumen
            self.assertIn('Datos', sheet_names)
            self.assertIn('Resumen', sheet_names)
            
            workbook.close()
    
    def test_export_with_column_type_conversions(self):
        """Probar exportación con conversiones de tipos de datos"""
        # DataFrame con diferentes tipos de datos
        df_types = pd.DataFrame({
            'Región': ['Norte', 'Sur', 'Este', 'Oeste'],
            'Entero': [1, 2, 3, 4],
            'Decimal': [1.5, 2.5, 3.5, 4.5],
            'Fecha': pd.date_range('2025-01-01', periods=4),
            'Bool': [True, False, True, False],
            'Texto': ['A', 'B', 'C', 'D']
        })
        
        # Configuración con mapeo específico
        config = self.config.copy()
        config['column_mapping'] = {
            'Región': 'A',
            'Entero': 'B',
            'Decimal': 'C',
            'Fecha': 'D',
            'Bool': 'E',
            'Texto': 'F'
        }
        
        # Ejecutar exportación
        result = exportar_datos_separados(df_types, config)
        
        # Verificar resultado
        self.assertTrue(result['success'])
        
        # Verificar contenido de un archivo
        file_path = result['files_created'][0]
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Verificar tipos de datos
        self.assertIsInstance(sheet['B5'].value, int)
        self.assertIsInstance(sheet['C5'].value, float)
        self.assertIsInstance(sheet['E5'].value, bool)
        
        workbook.close()
    
    def test_export_with_special_cell_formats(self):
        """Probar preservación de formatos de celdas especiales"""
        # Crear plantilla con formatos especiales
        format_template_path = os.path.join(self.temp_dir, 'format_template.xlsx')
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Datos'
        
        # Agregar diferentes formatos de celda
        sheet['A1'] = 'Fecha'
        sheet['B1'] = 'Cantidad'
        sheet['C1'] = 'Precio'
        
        # Aplicar diferentes formatos
        sheet['A1'].font = openpyxl.styles.Font(bold=True, color='FF0000')
        sheet['A1'].fill = openpyxl.styles.PatternFill(
            start_color='FFFF00', end_color='FFFF00', fill_type='solid'
        )
        
        sheet['B1'].number_format = '#,##0'
        sheet['C1'].number_format = '$#,##0.00'
        
        # Agregar bordes
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        
        for col in ['A1', 'B1', 'C1']:
            sheet[col].border = thin_border
        
        workbook.save(format_template_path)
        
        # Configuración con mapeo
        config = self.config.copy()
        config['template_path'] = format_template_path
        config['column_mapping'] = {
            'Región': 'A',
            'Ventas': 'B',
            'Descuento': 'C'
        }
        
        # Ejecutar exportación
        result = exportar_datos_separados(self.df, config)
        
        # Verificar resultado
        self.assertTrue(result['success'])
        
        # Verificar que los formatos se preservaron
        for file_path in result['files_created']:
            self.assertTrue(os.path.exists(file_path))
            
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # Verificar que los formatos se mantuvieron
            self.assertEqual(sheet['A5'].font.bold, True)
            self.assertEqual(sheet['A5'].font.color.rgb, 'FF0000')
            self.assertEqual(sheet['A5'].fill.start_color.rgb, 'FFFF00')
            
            workbook.close()
    
    def test_export_with_error_handling_and_logging(self):
        """Probar manejo de errores y logging durante exportación"""
        # Simular error en un grupo específico
        with patch.object(ExcelTemplateSplitter, '_create_excel_file_with_template') as mock_create:
            # Configurar mock para fallar solo en el primer grupo
            def side_effect(output_path, data):
                # Extraer nombre del grupo desde la ruta del archivo
                group_name = os.path.splitext(os.path.basename(output_path))[0]
                
                # Fallar solo para el primer grupo
                if 'Norte' in group_name:
                    raise Exception("Error simulado en la creación de archivo Excel")
                
                # Para otros grupos, crear archivo real
                mock_create_calls = mock_create.call_args_list
                if not mock_create_calls:
                    # Crear archivo real
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    
                    # Insertar algunos datos de prueba
                    for i, col_name in enumerate(data.columns):
                        col_letter = openpyxl.utils.get_column_letter(i+1)
                        for j, value in enumerate(data[col_name], start=5):
                            sheet[f'{col_letter}{j}'] = value
                    
                    workbook.save(output_path)
                    workbook.close()
                    return True
                
                return True
            
            mock_create.side_effect = side_effect
            
            # Ejecutar exportación
            result = exportar_datos_separados(self.df, self.config)
            
            # Verificar que se manejó el error correctamente
            self.assertTrue(result['success'])  # Algunos grupos deberían haberse exportado
            self.assertEqual(result['failed_exports'], 1)
            self.assertEqual(result['successful_exports'], 3)  # 3 de 4 grupos exitosos
            self.assertIn('failed_groups', result)
            self.assertIn('Norte', result['failed_groups'])
    
    def test_export_with_very_large_datasets_and_memory_optimization(self):
        """Probar exportación con datasets muy grandes y optimización de memoria"""
        # Crear dataset muy grande
        num_rows = 50000
        num_groups = 10
        df_very_large = pd.DataFrame({
            'Región': [f'Región_{i%num_groups}' for i in range(num_rows)],
            'Producto': [f'Producto_{i%100}' for i in range(num_rows)],
            'Ventas': [i for i in range(num_rows)],
            'Descripcion': [f'Descripción del producto {i%100} con texto largo' * 5 for i in range(num_rows)]
        })
        
        # Configuración con optimización de memoria
        config = self.config.copy()
        config['enable_chunking'] = True
        config['max_memory_mb'] = 10  # Límite muy bajo para forzar optimización
        
        start_time = time.time()
        
        # Ejecutar exportación
        result = exportar_datos_separados(df_very_large, config)
        
        elapsed_time = time.time() - start_time
        
        # Verificar resultado
        self.assertTrue(result['success'])
        self.assertEqual(result['groups_processed'], num_groups)
        self.assertEqual(result['total_rows'], num_rows)
        
        # Verificar rendimiento
        print(f"Tiempo de procesamiento: {elapsed_time:.2f} segundos para {num_rows} filas")
        self.assertLess(elapsed_time, 180)  # Menos de 3 minutos para 50K filas
        
        # Verificar archivos generados
        for file_path in result['files_created']:
            self.assertTrue(os.path.exists(file_path))
            
            # Verificar contenido del archivo
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # Verificar que hay datos
            max_row = sheet.max_row
            self.assertGreater(max_row, 1)  # Al menos encabezados y una fila de datos
            
            workbook.close()
    
    def test_export_end_to_end_with_real_data(self):
        """Prueba end-to-end con datos reales más complejos"""
        # Crear DataFrame más realista
        df_realistic = pd.DataFrame({
            'Empresa': ['EmpresaA', 'EmpresaB', 'EmpresaC', 'EmpresaA', 'EmpresaB'] * 20,
            'Departamento': ['Ventas', 'Marketing', 'IT', 'IT', 'Ventas'] * 20,
            'Empleado': [f'Empleado_{i}' for i in range(100)],
            'Salario': [50000 + (i*1000) for i in range(100)],
            'Fecha_Ingreso': pd.date_range('2020-01-01', periods=100, freq='5D'),
            'Evaluacion': np.random.uniform(1, 5, 100).round(2),
            'Activo': [True if i%3==0 else False for i in range(100)]
        })
        
        # Crear plantilla realista
        realistic_template_path = os.path.join(self.temp_dir, 'realistic_template.xlsx')
        self.create_realistic_template(realistic_template_path)
        
        # Configuración más compleja
        config = self.config.copy()
        config['separator_column'] = 'Empresa'
        config['template_path'] = realistic_template_path
        config['file_template'] = 'Reporte_{valor}_{fecha}.xlsx'
        config['column_mapping'] = {
            'Departamento': 'A',
            'Empleado': 'B',
            'Salario': 'C',
            'Fecha_Ingreso': 'D',
            'Evaluacion': 'E',
            'Activo': 'F'
        }
        
        # Ejecutar exportación
        result = exportar_datos_separados(df_realistic, config)
        
        # Verificar resultado
        self.assertTrue(result['success'])
        self.assertEqual(result['groups_processed'], 3)  # 3 empresas
        
        # Verificar archivos generados
        for file_path in result['files_created']:
            self.assertTrue(os.path.exists(file_path))
            
            # Verificar contenido del archivo
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # Verificar que los datos están en las columnas correctas
            self.assertIsNotNone(sheet['A2'].value)  # Primera fila de datos
            self.assertIsNotNone(sheet['B2'].value)
            
            workbook.close()
    
    def create_realistic_template(self, template_path):
        """Crear plantilla más realista"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Reporte_Empleados'
        
        # Encabezados en fila 1
        sheet['A1'] = 'Departamento'
        sheet['B1'] = 'Empleado'
        sheet['C1'] = 'Salario'
        sheet['D1'] = 'Fecha_Ingreso'
        sheet['E1'] = 'Evaluacion'
        sheet['F1'] = 'Activo'
        
        # Título del reporte
        sheet['A3'] = 'REPORTE DE EMPLEADOS'
        sheet['A3'].font = openpyxl.styles.Font(bold=True, size=14)
        
        # Aplicar formatos
        for col in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
            sheet[col].font = openpyxl.styles.Font(bold=True)
            sheet[col].fill = openpyxl.styles.PatternFill(
                start_color='DDDDDD', end_color='DDDDDD', fill_type='solid'
            )
        
        # Agregar nota al pie
        sheet['A100'] = 'Generado el: ' + datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sheet['A100'].font = openpyxl.styles.Font(italic=True, size=10)
        
        workbook.save(template_path)


if __name__ == '__main__':
    unittest.main()