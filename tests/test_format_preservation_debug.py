#!/usr/bin/env python3
"""
Script de prueba para validar preservación de formato Excel
Ejecuta test completo para verificar que la solución funciona
"""

import sys
import tempfile
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment

# Añadir directorios al path
sys.path.insert(0, str(Path(__file__).resolve().parent))

from core.data_handler import ExcelTemplateSplitter, ExportSeparatedConfig
from core.excel_format_preserver import ExcelFormatPreserver


def create_test_template_with_format() -> pd.DataFrame:
    """Crear plantilla de prueba con formato complejo"""
    template_path = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    
    # 1. Título con formato
    title = ws['A1']
    title.value = "REPORTE DE VENTAS"
    title.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')
    
    # 2. Headers con formato
    headers = ['Región', 'Ventas', 'Meta']
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i)
        cell.value = header
        cell.font = Font(name='Calibri', size=12, bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.border = Border(
            left=openpyxl.styles.Side(border_style='thin', color='000000'),
            right=openpyxl.styles.Side(border_style='thin', color='000000'),
            top=openpyxl.styles.Side(border_style='thin', color='000000'),
            bottom=openpyxl.styles.Side(border_style='thin', color='000000')
        )
    
    # 3. Formato de número
    num_cell = ws['B4']
    num_cell.value = 999999
    num_cell.number_format = '#,##0.00'
    
    # 4. Anchos personalizados
    ws.column_dimensions['A'].width = 20.0
    ws.column_dimensions['B'].width = 15.0
    ws.column_dimensions['C'].width = 12.0
    
    # 5. Datos de ejemplo en fila 5
    ws['A5'] = "Norte"
    ws['B5'] = 5000
    ws['C5'] = 6000
    
    # 6. Celda con formato especial
    special = ws['D2']
    special.value = "FORMATO ESPECIAL"
    special.font = Font(italic=True, color='FF0000', size=10)
    
    wb.save(template_path.name)
    print(f"✓ Plantilla de prueba creada: {template_path.name}")
    
    return template_path.name


def test_format_preservation() -> None:
    """Test principal de preservación de formato"""
    print("\n=== INICIANDO TEST DE PRESERVACIÓN DE FORMATO ===")
    
    # 1. Crear plantilla con formato
    template_path = create_test_template_with_format()
    
    # 2. Crear DataFrame de prueba
    df = pd.DataFrame({
        'Region': ['Norte', 'Sur', 'Este', 'Oeste'],
        'Ventas': [1000, 800, 900, 750],
        'Meta': [1200, 750, 950, 800]
    })
    
    # 3. Crear configuración
    output_folder = tempfile.mkdtemp()
    config = ExportSeparatedConfig(
        separator_column='Region',
        template_path=template_path,
        output_folder=output_folder,
        file_template='Reporte_{valor}.xlsx',
        column_mapping={
            'Region': 'A',
            'Ventas': 'B',
            'Meta': 'C'
        },
        start_cell='A5'
    )
    
    # 4. Crear splitter y procesar
    try:
        print("✓ Creando ExcelTemplateSplitter...")
        splitter = ExcelTemplateSplitter(df, config)
        
        print("✓ Ejecutando separación y exportación...")
        result = splitter.separate_and_export()
        
        if result['success']:
            print("✓ Separación exitosa!")
            print(f"✓ Archivos creados: {len(result['files_created'])}")
            
            # 5. Verificar que se preservó el formato
            for file_path in result['files_created']:
                print(f"\n--- Verificando formato en: {Path(file_path).name}")
                verify_format_preservation(file_path, template_path)
            
        else:
            print("❌ Error en la separación:")
            for error in result.get('errors', []):
                print(f"  - {error}")
            
    except Exception as e:
        print(f"❌ Error durante el test: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        # Limpiar archivos
        try:
            Path(template_path).unlink()
            import shutil
            shutil.rmtree(output_folder)
        except:
            pass


def verify_format_preservation(output_path: str, original_template: str) -> None:
    """Verificar que el formato se preservó correctamente"""
    try:
        # Cargar archivo original para comparar
        original_wb = openpyxl.load_workbook(original_template, data_only=False)
        original_sheet = original_wb.active
        
        # Cargar archivo de salida
        output_wb = openpyxl.load_workbook(output_path, data_only=False)
        output_sheet = output_wb.active
        
        # Verificar título
        print("  📋 Verificando título...")
        original_title = original_sheet['A1']
        output_title = output_sheet['A1']
        
        title_checks = [
            ("valor", original_title.value, output_title.value),
            ("font.name", original_title.font.name, output_title.font.name),
            ("font.size", original_title.font.size, output_title.font.size),
            ("font.bold", original_title.font.bold, output_title.font.bold),
            ("fill.color", original_title.fill.start_color.rgb, output_title.fill.start_color.rgb)
        ]
        
        for check_name, original_val, output_val in title_checks:
            if original_val == output_val:
                print(f"    ✓ {check_name}: {output_val}")
            else:
                print(f"    ❌ {check_name}: esperado {original_val}, obtenido {output_val}")
        
        # Verificar headers
        print("  📋 Verificando headers...")
        for col in ['A', 'B', 'C']:
            original_cell = original_sheet[f'{col}3']
            output_cell = output_sheet[f'{col}3']
            
            format_checks = [
                ("valor", original_cell.value, output_cell.value),
                ("font.name", original_cell.font.name, output_cell.font.name),
                ("font.size", original_cell.font.size, output_cell.font.size),
                ("font.bold", original_cell.font.bold, output_cell.font.bold),
                ("fill.color", original_cell.fill.start_color.rgb, output_cell.fill.start_color.rgb),
                ("border", original_cell.border.left.style is not None, output_cell.border.left.style is not None)
            ]
            
            for check_name, original_val, output_val in format_checks:
                if original_val == output_val:
                    print(f"    ✓ {col}3 {check_name}: {output_val}")
                else:
                    print(f"    ❌ {col}3 {check_name}: esperado {original_val}, obtenido {output_val}")
        
        # Verificar anchos de columna
        print("  📋 Verificando anchos de columna...")
        for col in ['A', 'B', 'C']:
            original_width = original_sheet.column_dimensions[col].width
            output_width = output_sheet.column_dimensions[col].width
            
            if original_width == output_width:
                print(f"    ✓ {col} width: {output_width}")
            else:
                print(f"    ❌ {col} width: esperado {original_width}, obtenido {output_width}")
        
        # Verificar celda con formato especial
        print("  📋 Verificando celda especial...")
        original_special = original_sheet['D2']
        output_special = output_sheet['D2']
        
        special_checks = [
            ("valor", original_special.value, output_special.value),
            ("font.italic", original_special.font.italic, output_special.font.italic),
            ("font.color", original_special.font.color.rgb, output_special.font.color.rgb)
        ]
        
        for check_name, original_val, output_val in special_checks:
            if original_val == output_val:
                print(f"    ✓ {check_name}: {output_val}")
            else:
                print(f"    ❌ {check_name}: esperado {original_val}, obtenido {output_val}")
        
        original_wb.close()
        output_wb.close()
        
        print("  ✅ Verificación de formato completada")
        
    except Exception as e:
        print(f"  ❌ Error verificando formato: {e}")
        import traceback
        traceback.print_exc()


def test_excel_format_preserver_directly() -> None:
    """Test directo del ExcelFormatPreserver"""
    print("\n=== TEST DIRECTO DE EXCELFORMATPRESERVER ===")
    
    try:
        # Crear preserver
        preserver = ExcelFormatPreserver()
        print("✓ ExcelFormatPreserver creado")
        
        # Crear plantilla de prueba
        template_path = create_test_template_with_format()
        
        # Cargar y cachear formatos
        wb = openpyxl.load_workbook(template_path, data_only=False)
        sheet = wb.active
        
        print("✓ Caching formatos...")
        formats_cache = preserver.cache_workbook_formats(wb)
        print(f"  - Cached {len(formats_cache.get('worksheet_formats', {}))} worksheets")
        print(f"  - Cached {len(formats_cache.get('cell_formats', {}))} cell formats")
        
        # Simular inserción de datos
        test_data = {0: {'Region': 'Test', 'Ventas': 1234, 'Meta': 5678}}
        
        print("✓ Insertando datos...")
        preserver.insert_data_preserving_format(
            sheet, 
            test_data, 
            {'Region': 'A', 'Ventas': 'B', 'Meta': 'C'},
            'A5'
        )
        
        # Restaurar formatos
        print("✓ Restaurando formatos...")
        preserver.restore_workbook_formats(wb, formats_cache)
        
        # Guardar archivo de prueba
        test_output = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(test_output.name)
        wb.close()
        
        print("✓ Verificando resultado...")
        verify_format_preservation(test_output.name, template_path)
        
        # Limpiar
        Path(template_path).unlink()
        Path(test_output.name).unlink()
        
    except Exception as e:
        print(f"❌ Error en test directo: {e}")
        import traceback
        traceback.print_exc()


def main() -> None:
    """Función principal"""
    print("🔍 DIAGNÓSTICO Y PRUEBA DE PRESERVACIÓN DE FORMATO EXCEL")
    print("=" * 60)
    
    # Test 1: Test directo del preserver
    test_excel_format_preserver_directly()
    
    # Test 2: Test completo con separación
    test_format_preservation()
    
    print("\n" + "=" * 60)
    print("✅ TESTS COMPLETADOS")
    print("\nRESUMEN:")
    print("- Se ha implementado ExcelFormatPreserver para preservar formato")
    print("- Se ha modificado _create_excel_file_with_template para usar preservación")
    print("- Los tests verifican que el formato se mantiene durante la inserción de datos")


if __name__ == '__main__':
    main()