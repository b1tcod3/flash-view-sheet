"""
Tests para la clase ExcelTemplateSplitter y funcionalidad de separación de datos.
"""

import unittest
import pandas as pd
import tempfile
import os
import json
from unittest.mock import patch, MagicMock, mock_open
import openpyxl

from core.data_handler import ExcelTemplateSplitter, ExportSeparatedConfig


class TestExcelTemplateSplitter(unittest.TestCase):
    """Tests para ExcelTemplateSplitter"""
    
    def setUp(self):
        """Configurar datos de prueba"""
        # Crear DataFrame de prueba
        self.df_test = pd.DataFrame({
            'Region': ['Norte', 'Norte', 'Sur', 'Sur', 'Este', 'Este', 'Oeste'],
            'Producto': ['A', 'B', 'A', 'B', 'A', 'B', 'A'],
            'Ventas': [1000, 1500, 800, 1200, 900, 1300, 750],
            'Fecha': pd.to_datetime(['2024-01-01', '2024-01-02', '2024-01-01', 
                                   '2024-01-02', '2024-01-01', '2024-01-02', '2024-01-01'])
        })
        
        # Crear configuración de prueba
        self.config = ExportSeparatedConfig(
            separator_column='Region',
            template_path='test_template.xlsx',
            output_folder='/tmp/test_output',
            file_template='Reporte_{valor}_{fecha}.xlsx',
            column_mapping={
                'Region': 'A',
                'Producto': 'B', 
                'Ventas': 'C',
                'Fecha': 'D'
            },
            start_cell='A5'
        )
        
        # Crear splitter de prueba
        self.splitter = ExcelTemplateSplitter(self.df_test, self.config)
    
    def test_init(self):
        """Test inicialización del splitter"""
        self.assertEqual(self.splitter.df.shape, (7, 4))
        self.assertEqual(self.splitter.config.separator_column, 'Region')
        self.assertEqual(len(self.splitter.groups_data), 0)
        self.assertFalse(self.splitter.is_cancelled)
        self.assertIsNotNone(self.splitter.progress_callback)
    
    def test_analyze_data(self):
        """Test análisis de datos"""
        result = self.splitter._analyze_data()
        
        self.assertIsInstance(result, dict)
        self.assertIn('total_rows', result)
        self.assertIn('groups_info', result)
        self.assertEqual(result['total_rows'], 7)
        
        # Verificar que se detectan 4 grupos únicos
        self.assertEqual(len(result['groups_info']), 4)
        
        # Verificar información de cada grupo
        for group_info in result['groups_info']:
            self.assertIn('name', group_info)
            self.assertIn('count', group_info)
            self.assertIn('preview', group_info)
            self.assertGreater(group_info['count'], 0)
    
    def test_generate_file_preview(self):
        """Test generación de preview de archivos"""
        # Crear template temporal para testing
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = "Plantilla de Prueba"
            wb.save(tmp.name)
            self.config.template_path = tmp.name
            
            preview = self.splitter._generate_file_preview()
            
            self.assertIsInstance(preview, list)
            self.assertEqual(len(preview), 4)  # 4 grupos
            
            for file_info in preview:
                self.assertIn('filename', file_info)
                self.assertIn('group_name', file_info)
                self.assertIn('row_count', file_info)
                self.assertIn('estimated_size_kb', file_info)
                self.assertTrue(file_info['filename'].endswith('.xlsx'))
    
    def test_validate_configuration(self):
        """Test validación de configuración"""
        # Test con configuración válida
        result = self.splitter._validate_configuration()
        self.assertTrue(result['valid'])
        self.assertEqual(len(result['errors']), 0)
    
    def test_validate_configuration_invalid_column(self):
        """Test validación con columna inexistente"""
        self.config.separator_column = 'ColumnaInexistente'
        self.splitter.config = self.config
        
        result = self.splitter._validate_configuration()
        self.assertFalse(result['valid'])
        self.assertGreater(len(result['errors']), 0)
    
    def test_create_excel_files_success(self):
        """Test creación exitosa de archivos Excel"""
        # Mock del método _create_excel_file
        with patch.object(self.splitter, '_create_excel_file') as mock_create:
            mock_create.return_value = True
            
            # Configurar datos de prueba
            self.splitter.groups_data = [
                {
                    'name': 'Norte',
                    'data': self.df_test[self.df_test['Region'] == 'Norte'],
                    'count': 2
                }
            ]
            
            result = self.splitter._create_excel_files()
            
            self.assertTrue(result['success'])
            self.assertEqual(len(result['files_created']), 1)
            mock_create.assert_called_once()
    
    def test_create_excel_files_failure(self):
        """Test fallo en creación de archivos Excel"""
        # Mock del método _create_excel_file para simular fallo
        with patch.object(self.splitter, '_create_excel_file') as mock_create:
            mock_create.return_value = False
            
            # Configurar datos de prueba
            self.splitter.groups_data = [
                {
                    'name': 'Norte',
                    'data': self.df_test[self.df_test['Region'] == 'Norte'],
                    'count': 2
                }
            ]
            
            result = self.splitter._create_excel_files()
            
            self.assertFalse(result['success'])
            self.assertEqual(len(result['files_created']), 0)
            self.assertGreater(len(result['errors']), 0)
    
    def test_create_excel_file_creates_workbook(self):
        """Test que se crea un workbook Excel correctamente"""
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = os.path.join(tmp_dir, 'test_output.xlsx')
            group_data = self.df_test[self.df_test['Region'] == 'Norte']
            
            # Mock template creation
            template_path = os.path.join(tmp_dir, 'template.xlsx')
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = "Template"
            wb.save(template_path)
            
            # Test creation
            success = self.splitter._create_excel_file(
                'Norte', 
                group_data, 
                output_path, 
                template_path
            )
            
            self.assertTrue(success)
            self.assertTrue(os.path.exists(output_path))
            
            # Verify content
            result_wb = openpyxl.load_workbook(output_path)
            result_ws = result_wb.active
            self.assertEqual(result_ws['A5'].value, 'Norte')
            self.assertEqual(result_ws['B5'].value, 'A')
            result_wb.close()
    
    def test_process_sequences(self):
        """Test procesamiento de secuencias"""
        # Test secuencia completa
        with patch.object(self.splitter, '_validate_configuration') as mock_validate, \
             patch.object(self.splitter, '_analyze_data') as mock_analyze, \
             patch.object(self.splitter, '_generate_file_preview') as mock_preview, \
             patch.object(self.splitter, '_create_excel_files') as mock_create:
            
            mock_validate.return_value = {'valid': True, 'errors': []}
            mock_analyze.return_value = {'total_rows': 7, 'groups_info': []}
            mock_preview.return_value = []
            mock_create.return_value = {'success': True, 'files_created': [], 'errors': []}
            
            result = self.splitter.process_sequences()
            
            self.assertTrue(result['success'])
            self.assertIn('files_created', result)
            self.assertIn('total_groups', result)
            self.assertGreater(result['total_groups'], 0)
            
            # Verificar orden de llamadas
            mock_validate.assert_called_once()
            mock_analyze.assert_called_once()
            mock_preview.assert_called_once()
            mock_create.assert_called_once()
    
    def test_cancel_operation(self):
        """Test cancelación de operación"""
        self.splitter.is_cancelled = False
        self.splitter.cancel_operation()
        self.assertTrue(self.splitter.is_cancelled)
    
    def test_to_excel_coordinates(self):
        """Test conversión a coordenadas Excel"""
        test_cases = [
            (0, 'A'),
            (1, 'B'),
            (25, 'Z'),
            (26, 'AA'),
            (27, 'AB'),
            (701, 'ZZ'),
            (702, 'AAA')
        ]
        
        for number, expected in test_cases:
            result = self.splitter._to_excel_coordinates(number)
            self.assertEqual(result, expected)


class TestExportSeparatedConfig(unittest.TestCase):
    """Tests para ExportSeparatedConfig"""
    
    def test_default_constructor(self):
        """Test constructor por defecto"""
        config = ExportSeparatedConfig()
        
        self.assertEqual(config.separator_column, '')
        self.assertEqual(config.template_path, '')
        self.assertEqual(config.output_folder, '')
        self.assertEqual(config.file_template, '{valor}.xlsx')
        self.assertEqual(config.column_mapping, {})
        self.assertEqual(config.start_cell, 'A1')
        self.assertEqual(config.handle_duplicates, 'overwrite')
        self.assertEqual(config.create_summary, True)
        self.assertEqual(config.preserve_format, True)
    
    def test_custom_constructor(self):
        """Test constructor con valores personalizados"""
        config = ExportSeparatedConfig(
            separator_column='Region',
            template_path='template.xlsx',
            output_folder='/output',
            file_template='Report_{valor}.xlsx',
            start_cell='B3',
            handle_duplicates='skip',
            create_summary=False,
            preserve_format=False
        )
        
        self.assertEqual(config.separator_column, 'Region')
        self.assertEqual(config.template_path, 'template.xlsx')
        self.assertEqual(config.output_folder, '/output')
        self.assertEqual(config.file_template, 'Report_{valor}.xlsx')
        self.assertEqual(config.start_cell, 'B3')
        self.assertEqual(config.handle_duplicates, 'skip')
        self.assertEqual(config.create_summary, False)
        self.assertEqual(config.preserve_format, False)
    
    def test_validate_required_fields(self):
        """Test validación de campos requeridos"""
        # Configuración completa válida
        config = ExportSeparatedConfig(
            separator_column='Region',
            template_path='template.xlsx',
            output_folder='/output',
            file_template='Report.xlsx'
        )
        
        errors = config.validate_required_fields()
        self.assertEqual(len(errors), 0)
    
    def test_validate_required_fields_missing(self):
        """Test validación con campos faltantes"""
        config = ExportSeparatedConfig(
            separator_column='',
            template_path='template.xlsx'
        )
        
        errors = config.validate_required_fields()
        self.assertGreater(len(errors), 0)
        self.assertTrue(any('separator_column' in error for error in errors))
        self.assertTrue(any('output_folder' in error for error in errors))
    
    def test_get_default_mapping(self):
        """Test obtención de mapeo por defecto"""
        config = ExportSeparatedConfig()
        df = pd.DataFrame({'A': [1], 'B': [2], 'C': [3]})
        
        mapping = config.get_default_mapping(df)
        
        self.assertIsInstance(mapping, dict)
        self.assertEqual(mapping['A'], 'A')
        self.assertEqual(mapping['B'], 'B')
        self.assertEqual(mapping['C'], 'C')
    
    def test_to_dict(self):
        """Test conversión a diccionario"""
        config = ExportSeparatedConfig(
            separator_column='Region',
            template_path='template.xlsx',
            output_folder='/output'
        )
        
        config_dict = config.to_dict()
        
        self.assertIsInstance(config_dict, dict)
        self.assertEqual(config_dict['separator_column'], 'Region')
        self.assertEqual(config_dict['template_path'], 'template.xlsx')
        self.assertEqual(config_dict['output_folder'], '/output')
    
    def test_from_dict(self):
        """Test creación desde diccionario"""
        data = {
            'separator_column': 'Region',
            'template_path': 'template.xlsx',
            'output_folder': '/output',
            'file_template': 'Custom_{valor}.xlsx'
        }
        
        config = ExportSeparatedConfig.from_dict(data)
        
        self.assertEqual(config.separator_column, 'Region')
        self.assertEqual(config.template_path, 'template.xlsx')
        self.assertEqual(config.output_folder, '/output')
        self.assertEqual(config.file_template, 'Custom_{valor}.xlsx')


if __name__ == '__main__':
    unittest.main()