"""
Tests específicos para verificar preservación de formato Excel
"""

import unittest
import pandas as pd
import tempfile
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter

from core.data_handler import ExcelTemplateSplitter, ExportSeparatedConfig
from core.excel_format_preserver import ExcelFormatPreserver


class TestExcelFormatPreservation(unittest.TestCase):
    """Tests para verificar que el formato Excel se preserve durante la inserción de datos"""
    
    def setUp(self):
        """Configurar datos y plantillas de prueba con formato"""
        # Crear DataFrame de prueba
        self.df_test = pd.DataFrame({
            'Region': ['Norte', 'Sur', 'Este'],
            'Ventas': [1000, 800, 900],
            'Meta': [1200, 750, 950]
        })
        
        # Crear plantilla con formato específico
        self.temp_template = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self._create_formatted_template(self.temp_template.name)
        
        # Crear configuración
        self.config = ExportSeparatedConfig(
            separator_column='Region',
            template_path=self.temp_template.name,
            output_folder=tempfile.mkdtemp(),
            file_template='{valor}_test.xlsx',
            column_mapping={
                'Region': 'A',
                'Ventas': 'B', 
                'Meta': 'C'
            },
            start_cell='A5'
        )
    
    def tearDown(self):
        """Limpiar archivos temporales"""
        try:
            os.unlink(self.temp_template.name)
            import shutil
            shutil.rmtree(self.config.output_folder)
        except:
            pass
    
    def _create_formatted_template(self, template_path: str):
        """Crear plantilla Excel con formato específico para testing"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Crear diferentes elementos de formato
        
        # 1. Título con formato específico
        title_cell = ws['A1']
        title_cell.value = "REPORTE DE VENTAS"
        title_cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 2. Headers con formato
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}3']
            cell.value = f"Columna {col}"
            cell.font = Font(name='Calibri', size=12, bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.border = Border(
                left=openpyxl.styles.Side(border_style='thin', color='000000'),
                right=openpyxl.styles.Side(border_style='thin', color='000000'),
                top=openpyxl.styles.Side(border_style='thin', color='000000'),
                bottom=openpyxl.styles.Side(border_style='thin', color='000000')
            )
        
        # 3. Celda con formato de número
        number_cell = ws['B4']
        number_cell.value = 999999
        number_cell.number_format = '#,##0.00'
        
        # 4. Fila con datos de ejemplo (que se van a sobrescribir)
        ws['A5'] = "Ejemplo Norte"
        ws['B5'] = 5000
        ws['C5'] = 6000
        
        # 5. Comentarios/formato adicional
        comment_cell = ws['D2']
        comment_cell.value = "TEST"
        comment_cell.font = Font(italic=True, color='FF0000')
        
        # 6. Anchos de columna personalizados
        ws.column_dimensions['A'].width = 20.0
        ws.column_dimensions['B'].width = 15.0
        ws.column_dimensions['C'].width = 12.0
        
        # 7. Alturas de fila personalizadas
        ws.row_dimensions[1].height = 30.0
        
        wb.save(template_path)
    
    def test_format_preservation_basic(self):
        """Test básico de preservación de formato"""
        # Crear splitter
        splitter = ExcelTemplateSplitter(self.df_test, self.config)
        
        # Obtener grupo de datos de prueba
        group_data = self.df_test[self.df_test['Region'] == 'Norte'].copy()
        
        # Crear archivo de salida
        output_path = os.path.join(self.config.output_folder, 'Norte_test_format.xlsx')
        
        # Procesar grupo
        result = splitter._export_group('Norte', group_data)
        
        # Verificar que se creó el archivo
        self.assertTrue(result.success)
        self.assertTrue(os.path.exists(output_path))
        
        # Verificar que el formato se preservó
        self._verify_format_preservation(output_path)
    
    def test_format_preservation_with_preserver(self):
        """Test de preservación usando ExcelFormatPreserver directamente"""
        preserver = ExcelFormatPreserver()
        
        # Cargar plantilla original
        original_wb = openpyxl.load_workbook(self.temp_template.name, data_only=False)
        original_sheet = original_wb.active
        
        # Cachear formatos originales
        original_formats = preserver.cache_workbook_formats(original_wb)
        
        # Simular inserción de datos
        test_data = {0: {'Region': 'Test', 'Ventas': 1234, 'Meta': 5678}}
        preserver.insert_data_preserving_format(
            original_sheet, 
            test_data, 
            {'Region': 'A', 'Ventas': 'B', 'Meta': 'C'},
            'A5'
        )
        
        # Guardar archivo de prueba
        test_output = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        original_wb.save(test_output.name)
        original_wb.close()
        
        try:
            # Verificar que el formato se preservó
            self._verify_format_preservation(test_output.name)
        finally:
            try:
                os.unlink(test_output.name)
            except:
                pass
    
    def _verify_format_preservation(self, output_path: str):
        """Verificar que el formato se preservó en el archivo de salida"""
        # Cargar archivo de salida
        output_wb = openpyxl.load_workbook(output_path, data_only=False)
        output_sheet = output_wb.active
        
        # 1. Verificar título
        title_cell = output_sheet['A1']
        self.assertEqual(title_cell.value, "REPORTE DE VENTAS")
        self.assertEqual(title_cell.font.name, 'Arial')
        self.assertEqual(title_cell.font.size, 16)
        self.assertTrue(title_cell.font.bold)
        self.assertEqual(title_cell.fill.start_color.rgb, '366092')
        
        # 2. Verificar headers
        for col in ['A', 'B', 'C']:
            cell = output_sheet[f'{col}3']
            self.assertEqual(cell.value, f"Columna {col}")
            self.assertEqual(cell.font.name, 'Calibri')
            self.assertEqual(cell.font.size, 12)
            self.assertTrue(cell.font.bold)
            self.assertEqual(cell.fill.start_color.rgb, 'D9E1F2')
            # Verificar que tiene borde
            self.assertIsNotNone(cell.border.left.style)
        
        # 3. Verificar formato de número
        number_cell = output_sheet['B4']
        self.assertEqual(number_cell.value, 999999)
        self.assertEqual(number_cell.number_format, '#,##0.00')
        
        # 4. Verificar que se insertaron nuevos datos en A5
        new_data_cell = output_sheet['A5']
        self.assertEqual(new_data_cell.value, 'Norte')  # Primer valor del DataFrame
        
        # 5. Verificar anchos de columna
        self.assertEqual(output_sheet.column_dimensions['A'].width, 20.0)
        self.assertEqual(output_sheet.column_dimensions['B'].width, 15.0)
        self.assertEqual(output_sheet.column_dimensions['C'].width, 12.0)
        
        # 6. Verificar altura de fila
        self.assertEqual(output_sheet.row_dimensions[1].height, 30.0)
        
        # 7. Verificar celda con comentario/formato especial
        comment_cell = output_sheet['D2']
        self.assertEqual(comment_cell.value, "TEST")
        self.assertTrue(comment_cell.font.italic)
        self.assertEqual(comment_cell.font.color.rgb, 'FFFF0000')  # Rojo
        
        output_wb.close()
    
    def test_multiple_format_elements(self):
        """Test con múltiples elementos de formato complejo"""
        # Crear plantilla más compleja
        complex_template = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Headers con diferentes formatos
        ws['A1'] = "ID"
        ws['B1'] = "Nombre" 
        ws['C1'] = "Valor"
        
        # Formatear cada header diferente
        ws['A1'].font = Font(bold=True, color='FF0000')  # Rojo y bold
        ws['B1'].font = Font(bold=True, color='0000FF')  # Azul y bold
        ws['C1'].font = Font(bold=True, color='008000')  # Verde y bold
        
        # Datos de ejemplo
        ws['A2'] = "001"
        ws['B2'] = "Test"
        ws['C2'] = 100.50
        
        # Merge cells
        ws.merge_cells('D1:F1')
        ws['D1'] = "SECCIÓN COMBINADA"
        ws['D1'].font = Font(size=14, bold=True)
        ws['D1'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        wb.save(complex_template.name)
        
        # Test con configuración actualizada
        config = ExportSeparatedConfig(
            separator_column='Region',
            template_path=complex_template.name,
            output_folder=tempfile.mkdtemp(),
            file_template='{valor}_complex.xlsx',
            column_mapping={
                'Region': 'A',
                'Ventas': 'B', 
                'Meta': 'C'
            },
            start_cell='A3'
        )
        
        try:
            splitter = ExcelTemplateSplitter(self.df_test, config)
            group_data = self.df_test[self.df_test['Region'] == 'Norte'].copy()
            result = splitter._export_group('Norte', group_data)
            
            self.assertTrue(result.success)
            
            # Verificar formato complejo
            output_wb = openpyxl.load_workbook(result.file_path, data_only=False)
            output_sheet = output_wb.active
            
            # Verificar que merge cells se preservó
            merged_ranges = list(output_sheet.merged_cells.ranges)
            self.assertGreater(len(merged_ranges), 0)
            
            # Verificar que formato de headers se preservó
            self.assertEqual(output_sheet['A1'].font.color.rgb, 'FFFF0000')  # Rojo
            self.assertEqual(output_sheet['B1'].font.color.rgb, 'FF0000FF')  # Azul
            self.assertEqual(output_sheet['C1'].font.color.rgb, 'FF008000')  # Verde
            
            output_wb.close()
            
        finally:
            try:
                os.unlink(complex_template.name)
                import shutil
                shutil.rmtree(config.output_folder)
            except:
                pass


if __name__ == '__main__':
    unittest.main()