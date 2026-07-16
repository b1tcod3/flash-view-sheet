"""
Test para verificar la corrección del problema de validación de plantilla
"""
import tempfile
from pathlib import Path
import pandas as pd
import openpyxl
from PySide6.QtWidgets import QApplication
from app.widgets.export_separated_dialog import ExportSeparatedDialog

def test_template_path_stored() -> None:
    """Test que verifica que la ruta de plantilla se almacena correctamente"""
    print("🧪 Iniciando test de corrección de ruta de plantilla...")
    
    # Crear una aplicación Qt
    app = QApplication.instance()
    if app is None:
        app = QApplication([])
    
    # Crear datos de prueba
    df = pd.DataFrame({
        'Region': ['Norte', 'Sur', 'Norte', 'Este'],
        'Ventas': [1000, 2000, 1500, 1800]
    })
    
    # Crear plantilla Excel temporal
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        template_path = temp_file.name
    
    try:
        # Crear archivo Excel de prueba
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Región'
        ws['B1'] = 'Ventas'
        wb.save(template_path)
        wb.close()
        
        # Crear diálogo
        dialog = ExportSeparatedDialog(df)
        
        # Simular selección de plantilla
        dialog._template_path = template_path  # Esto es lo que agregamos en la corrección
        dialog.template_path_label.setText(f"📄 {Path(template_path).name}")
        dialog.template_path_label.setToolTip(template_path)
        
        # Verificar que la ruta se almacena correctamente
        assert hasattr(dialog, '_template_path'), "El diálogo debe tener el atributo _template_path"
        assert dialog._template_path == template_path, f"La ruta debe ser {template_path}"
        
        # Verificar que la configuración puede obtener la ruta
        config = dialog.get_configuration(validate=False)
        assert config is not None, "La configuración no debe ser None"
        assert config.template_path == template_path, "La plantilla en la configuración debe coincidir"
        
        print("✅ Test pasado: La ruta de plantilla se almacena y recupera correctamente")
        return True
        
    except Exception as e:
        print(f"❌ Test falló: {str(e)}")
        return False
    finally:
        # Limpiar archivo temporal
        if Path(template_path).exists():
            Path(template_path).unlink()

def test_validation_with_stored_template() -> None:
    """Test que verifica la validación con plantilla almacenada"""
    print("🧪 Iniciando test de validación con plantilla almacenada...")
    
    # Crear una aplicación Qt
    app = QApplication.instance()
    if app is None:
        app = QApplication([])
    
    # Crear datos de prueba
    df = pd.DataFrame({
        'Region': ['Norte', 'Sur'],
        'Ventas': [1000, 2000]
    })
    
    # Crear plantilla Excel temporal
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        template_path = temp_file.name
    
    try:
        # Crear archivo Excel de prueba
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Región'
        ws['B1'] = 'Ventas'
        wb.save(template_path)
        wb.close()
        
        # Crear diálogo y configurar
        dialog = ExportSeparatedDialog(df)
        dialog._template_path = template_path
        dialog.column_combo.setCurrentText('Region')
        dialog.dest_folder_label.setText("/tmp")
        dialog.filename_template_edit.setText("{valor}.xlsx")
        
        # Intentar validación (esto no debe fallar ahora)
        try:
            validation_result = dialog.validate_configuration()
            print("✅ Test pasado: La validación funciona correctamente con plantilla almacenada")
            return True
        except Exception as e:
            print(f"❌ Test falló en validación: {str(e)}")
            return False
        
    except Exception as e:
        print(f"❌ Test falló: {str(e)}")
        return False
    finally:
        # Limpiar archivo temporal
        if Path(template_path).exists():
            Path(template_path).unlink()

def test_filename_extension_validation() -> None:
    """Test que verifica validación de extensión de archivo de exportación"""
    print("🧪 Iniciando test de validación de extensión de archivo...")
    
    # Crear una aplicación Qt
    app = QApplication.instance()
    if app is None:
        app = QApplication([])
    
    # Crear datos de prueba
    df = pd.DataFrame({
        'Region': ['Norte', 'Sur'],
        'Ventas': [1000, 2000]
    })
    
    # Crear plantilla Excel temporal
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        template_path = temp_file.name
    
    try:
        # Crear archivo Excel de prueba
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Región'
        ws['B1'] = 'Ventas'
        wb.save(template_path)
        wb.close()
        
        # Crear diálogo
        dialog = ExportSeparatedDialog(df)
        dialog._template_path = template_path
        dialog.column_combo.setCurrentText('Region')
        dialog.dest_folder_label.setText("/tmp")
        
        # Test 1: Plantilla con extensión válida .xlsx
        dialog.filename_template_edit.setText("{valor}.xlsx")
        config = dialog.get_configuration(validate=False)
        assert config is not None, "Configuración debe ser válida con extensión .xlsx"
        print("✅ Extensión .xlsx válida - Correcta")
        
        # Test 2: Plantilla con extensión válida .xlsm
        dialog.filename_template_edit.setText("{valor}.xlsm")
        config = dialog.get_configuration(validate=False)
        assert config is not None, "Configuración debe ser válida con extensión .xlsm"
        print("✅ Extensión .xlsm válida - Correcta")
        
        # Test 3: Plantilla con extensión inválida
        dialog.filename_template_edit.setText("{valor}.doc")
        config = dialog.get_configuration(validate=False)
        assert config is None, "Configuración debe ser None con extensión .doc inválida"
        print("✅ Extensión .doc inválida - Correctamente rechazada")
        
        # Test 4: Plantilla sin extensión
        dialog.filename_template_edit.setText("{valor}")
        config = dialog.get_configuration(validate=False)
        assert config is None, "Configuración debe ser None sin extensión"
        print("✅ Sin extensión - Correctamente rechazada")
        
        # Test 5: Plantilla con extensión .XLSX (mayúsculas)
        dialog.filename_template_edit.setText("{valor}.XLSX")
        config = dialog.get_configuration(validate=False)
        assert config is not None, "Configuración debe ser válida con extensión .XLSX (mayúsculas)"
        print("✅ Extensión .XLSX (mayúsculas) válida - Correcta")
        
        print("✅ Todos los tests de validación de extensión pasaron")
        return True
        
    except Exception as e:
        print(f"❌ Test de validación de extensión falló: {str(e)}")
        return False
    finally:
        # Limpiar archivo temporal
        if Path(template_path).exists():
            Path(template_path).unlink()

if __name__ == "__main__":
    print("🔍 Test de Corrección: Validación de Plantilla y Nombre de Archivo")
    print("=" * 60)
    
    test1_passed = test_template_path_stored()
    test2_passed = test_validation_with_stored_template()
    test3_passed = test_filename_extension_validation()
    
    print("\n" + "=" * 60)
    if test1_passed and test2_passed and test3_passed:
        print("🎉 Todos los tests pasaron. Las correcciones funcionan correctamente.")
    else:
        print("⚠️  Algunos tests fallaron. Revisar la implementación.")