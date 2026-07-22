"""
Diálogo de Configuración para Exportación Separada con Plantillas Excel
Diseño compacto con QSplitter (panel dividido lado a lado).
"""

from PySide6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel,
                               QPushButton, QGroupBox, QFormLayout, QGridLayout,
                               QComboBox, QLineEdit, QSpinBox,
                               QMessageBox, QListWidget,
                               QCheckBox, QWidget,
                               QTableWidget, QTableWidgetItem, QHeaderView,
                               QFileDialog, QFrame, QSizePolicy, QSplitter,
                               QProgressBar)
from PySide6.QtCore import Qt, Signal, QTimer, QSettings
from PySide6.QtGui import QFont
import pandas as pd
import os
import openpyxl
from pathlib import Path
from datetime import datetime
from typing import Any

import sys
sys.path.append('..')
from core.data_handler import (
    ExcelTemplateSplitter, ExportSeparatedConfig, ValidationResult,
    SeparationError, TemplateError, ConfigurationError, MemoryError
)


# ──────────────────────────────────────────────────────────────────────
# ColumnMappingWidget
# ──────────────────────────────────────────────────────────────────────

class ColumnMappingWidget(QWidget):
    """Widget para gestionar mapeo de columnas DataFrame ↔ Excel"""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.df_columns: list[str] = []
        self.excel_columns: list[str] = []
        self.mapping: dict[str, str] = {}
        self.setup_ui()

    def setup_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        self.mapping_table = QTableWidget(0, 4)
        self.mapping_table.setHorizontalHeaderLabels([
            "Columna Data", "→", "Columna Excel", "Vista Previa"
        ])

        header = self.mapping_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # type: ignore[attr-defined]
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # type: ignore[attr-defined]
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # type: ignore[attr-defined]
        header.setStretchLastSection(True)
        self.mapping_table.setAlternatingRowColors(True)

        layout.addWidget(self.mapping_table)

        button_layout = QHBoxLayout()
        self.auto_map_btn = QPushButton("Auto-mapear")
        self.add_column_btn = QPushButton("+ Añadir")
        self.remove_column_btn = QPushButton("- Eliminar")

        self.auto_map_btn.clicked.connect(self.auto_map_positional)
        self.add_column_btn.clicked.connect(self.add_mapping_row)
        self.remove_column_btn.clicked.connect(self.remove_selected_rows)

        button_layout.addWidget(self.auto_map_btn)
        button_layout.addWidget(self.add_column_btn)
        button_layout.addWidget(self.remove_column_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)

    def set_dataframe_columns(self, columns: list[str]) -> None:
        self.df_columns = columns
        self.refresh_mapping_table()

    def set_excel_columns(self, columns: list[str]) -> None:
        self.excel_columns = columns
        for row in range(self.mapping_table.rowCount()):
            combo = self.mapping_table.cellWidget(row, 2)
            if combo:
                combo.clear()
                combo.addItems(self.excel_columns)

    def refresh_mapping_table(self) -> None:
        self.mapping_table.setRowCount(len(self.df_columns))
        for row, col_name in enumerate(self.df_columns):
            item = QTableWidgetItem(col_name)
            item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)  # type: ignore[attr-defined]
            self.mapping_table.setItem(row, 0, item)

            arrow_item = QTableWidgetItem("→")
            arrow_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)  # type: ignore[attr-defined]
            arrow_item.setTextAlignment(Qt.AlignCenter)  # type: ignore[attr-defined]
            self.mapping_table.setItem(row, 1, arrow_item)

            combo = QComboBox()
            combo.addItems(self.excel_columns)
            self.mapping_table.setCellWidget(row, 2, combo)

            preview_item = QTableWidgetItem("Sin datos")
            preview_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)  # type: ignore[attr-defined]
            self.mapping_table.setItem(row, 3, preview_item)

    def auto_map_positional(self) -> None:
        for row in range(min(len(self.df_columns), len(self.excel_columns))):
            combo = self.mapping_table.cellWidget(row, 2)
            if combo:
                combo.setCurrentIndex(row)

    def add_mapping_row(self) -> None:
        row = self.mapping_table.rowCount()
        self.mapping_table.insertRow(row)
        self.mapping_table.setItem(row, 0, QTableWidgetItem(""))

        arrow_item = QTableWidgetItem("→")
        arrow_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)  # type: ignore[attr-defined]
        arrow_item.setTextAlignment(Qt.AlignCenter)  # type: ignore[attr-defined]
        self.mapping_table.setItem(row, 1, arrow_item)

        combo = QComboBox()
        combo.addItems(self.excel_columns)
        self.mapping_table.setCellWidget(row, 2, combo)
        self.mapping_table.setItem(row, 3, QTableWidgetItem(""))

    def remove_selected_rows(self) -> None:
        selected_rows = sorted(
            [item.row() for item in self.mapping_table.selectedIndexes()],
            reverse=True
        )
        for row in selected_rows:
            self.mapping_table.removeRow(row)

    def get_mapping(self) -> dict[str, str]:
        mapping: dict[str, str] = {}
        for row in range(self.mapping_table.rowCount()):
            df_col_item = self.mapping_table.item(row, 0)
            combo = self.mapping_table.cellWidget(row, 2)
            if df_col_item and combo:
                df_col = df_col_item.text().strip()
                excel_col = combo.currentText().strip()
                if df_col and excel_col:
                    mapping[df_col] = excel_col
        return mapping


# ──────────────────────────────────────────────────────────────────────
# ExcelTemplateDialog
# ──────────────────────────────────────────────────────────────────────

class ExcelTemplateDialog(QDialog):
    """Diálogo para seleccionar y validar plantilla Excel"""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.template_path: str = ""
        self.selected_sheet: str = ""
        self.available_sheets: list[str] = []
        self.setWindowTitle("Seleccionar Plantilla Excel")
        self.resize(700, 500)
        self.setup_ui()

    def setup_ui(self) -> None:
        layout = QVBoxLayout(self)

        file_group = QGroupBox("Archivo de Plantilla")
        file_layout = QFormLayout(file_group)

        self.file_path_edit = QLineEdit()
        self.file_path_edit.setReadOnly(True)
        file_layout.addRow("Ruta:", self.file_path_edit)

        button_layout = QHBoxLayout()
        self.browse_btn = QPushButton("Explorar...")
        self.browse_btn.clicked.connect(self.browse_for_template)
        button_layout.addWidget(self.browse_btn)
        button_layout.addStretch()
        file_layout.addRow("", button_layout)

        layout.addWidget(file_group)

        info_group = QGroupBox("Información del Archivo")
        info_layout = QFormLayout(info_group)

        self.info_label = QLabel("No hay archivo seleccionado")
        self.info_label.setWordWrap(True)
        info_layout.addRow("Estado:", self.info_label)

        self.sheet_combo = QComboBox()
        self.sheet_combo.setEnabled(False)
        self.sheet_combo.currentTextChanged.connect(self.on_sheet_changed)
        info_layout.addRow("Hoja:", self.sheet_combo)

        layout.addWidget(info_group)

        preview_group = QGroupBox("Vista Previa de la Plantilla")
        preview_layout = QVBoxLayout(preview_group)

        self.preview_table = QTableWidget(0, 0)
        self.preview_table.setMaximumHeight(200)
        self.preview_table.setAlternatingRowColors(True)
        preview_layout.addWidget(self.preview_table)

        layout.addWidget(preview_group)

        buttons = QHBoxLayout()
        self.ok_btn = QPushButton("Aceptar")
        self.ok_btn.clicked.connect(self.accept)
        self.cancel_btn = QPushButton("Cancelar")
        self.cancel_btn.clicked.connect(self.reject)
        buttons.addStretch()
        buttons.addWidget(self.ok_btn)
        buttons.addWidget(self.cancel_btn)
        layout.addLayout(buttons)

    def browse_for_template(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar Plantilla Excel", "",
            "Excel Files (*.xlsx *.xlsm);;All Files (*)"
        )
        if file_path:
            self.load_template(file_path)

    def load_template(self, file_path: str) -> None:
        try:
            if not file_path.lower().endswith(('.xlsx', '.xlsm')):
                self.info_label.setText("Error: El archivo debe ser .xlsx o .xlsm")
                return

            workbook = openpyxl.load_workbook(file_path, data_only=False)
            self.available_sheets = workbook.sheetnames
            workbook.close()

            self.template_path = file_path
            self.file_path_edit.setText(file_path)

            file_size = Path(file_path).stat().st_size
            self.info_label.setText(
                f"Archivo válido — Tamaño: {file_size / 1024:.1f}KB — "
                f"Hojas: {len(self.available_sheets)}"
            )

            self.sheet_combo.clear()
            self.sheet_combo.addItems(self.available_sheets)
            self.sheet_combo.setEnabled(True)
            self.selected_sheet = self.available_sheets[0] if self.available_sheets else ""

            self.show_preview()

        except Exception as e:
            self.info_label.setText(f"Error al cargar plantilla: {e}")
            self.template_path = ""
            self.file_path_edit.clear()
            self.sheet_combo.setEnabled(False)
            self.preview_table.setRowCount(0)
            self.preview_table.setColumnCount(0)

    def show_preview(self) -> None:
        try:
            if not self.template_path:
                return
            workbook = openpyxl.load_workbook(self.template_path, data_only=True)
            sheet = workbook[self.selected_sheet] if self.selected_sheet else workbook.active

            max_rows = min(10, sheet.max_row or 0)
            max_cols = min(10, sheet.max_column or 0)

            self.preview_table.setRowCount(max_rows)
            self.preview_table.setColumnCount(max_cols)

            for row in range(max_rows):
                for col in range(max_cols):
                    cell_value = sheet.cell(row + 1, col + 1).value
                    if cell_value is not None:
                        self.preview_table.setItem(row, col, QTableWidgetItem(str(cell_value)))

            workbook.close()
        except Exception:
            pass

    def on_sheet_changed(self, sheet_name: str) -> None:
        self.selected_sheet = sheet_name
        if self.template_path:
            self.show_preview()

    def get_template_info(self) -> tuple[str, str, list[str]]:
        return self.template_path, self.selected_sheet, self.available_sheets


# ──────────────────────────────────────────────────────────────────────
# FilePreviewDialog
# ──────────────────────────────────────────────────────────────────────

class FilePreviewDialog(QDialog):
    """Diálogo para vista previa de archivos a generar"""

    def __init__(self, files_info: list[dict[str, Any]], parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.files_info = files_info
        self.setWindowTitle("Vista Previa de Archivos")
        self.resize(800, 400)
        self.setup_ui()

    def setup_ui(self) -> None:
        layout = QVBoxLayout(self)

        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Filtros:"))
        self.status_filter = QComboBox()
        self.status_filter.addItems(["Todos", "Listos", "Advertencias", "Errores"])
        self.status_filter.currentTextChanged.connect(self.filter_files)
        filter_layout.addWidget(self.status_filter)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        self.files_table = QTableWidget(0, 5)
        self.files_table.setHorizontalHeaderLabels([
            "Nombre de Archivo", "Grupo", "Filas", "Tamaño", "Estado"
        ])
        header = self.files_table.horizontalHeader()
        header.setStretchLastSection(True)
        self.files_table.setAlternatingRowColors(True)
        layout.addWidget(self.files_table)

        self.summary_label = QLabel()
        layout.addWidget(self.summary_label)

        buttons = QHBoxLayout()
        self.ok_btn = QPushButton("Aceptar")
        self.ok_btn.clicked.connect(self.accept)
        buttons.addStretch()
        buttons.addWidget(self.ok_btn)
        layout.addLayout(buttons)

        self.refresh_table()

    def refresh_table(self) -> None:
        filtered_files = self.get_filtered_files()
        self.files_table.setRowCount(len(filtered_files))

        for row, file_info in enumerate(filtered_files):
            self.files_table.setItem(row, 0, QTableWidgetItem(file_info.get('filename', '')))
            self.files_table.setItem(row, 1, QTableWidgetItem(file_info.get('group', '')))
            self.files_table.setItem(row, 2, QTableWidgetItem(str(file_info.get('rows', 0))))
            self.files_table.setItem(row, 3, QTableWidgetItem(file_info.get('size', '')))

            status_item = QTableWidgetItem(file_info.get('status', ''))
            if file_info.get('status') == 'Listo':
                status_item.setBackground(Qt.green)  # type: ignore[attr-defined]
            elif file_info.get('status') == 'Advertencia':
                status_item.setBackground(Qt.yellow)  # type: ignore[attr-defined]
            elif file_info.get('status') == 'Error':
                status_item.setBackground(Qt.red)  # type: ignore[attr-defined]
            self.files_table.setItem(row, 4, status_item)

        self.update_summary(filtered_files)

    def get_filtered_files(self) -> list[dict[str, Any]]:
        filter_text = self.status_filter.currentText()
        if filter_text == "Todos":
            return self.files_info
        return [f for f in self.files_info if f.get('status') == filter_text]

    def update_summary(self, files: list[dict[str, Any]]) -> None:
        total_files = len(files)
        total_rows = sum(f.get('rows', 0) for f in files)
        self.summary_label.setText(
            f"Resumen: {total_files} archivos — {total_rows} filas totales"
        )

    def filter_files(self) -> None:
        self.refresh_table()


# ──────────────────────────────────────────────────────────────────────
# ExportSeparatedDialog — Vista Compacta
# ──────────────────────────────────────────────────────────────────────

class ExportSeparatedDialog(QDialog):
    """
    Diálogo compacto de configuración de exportación separada.
    Panel dividido (QSplitter): configuración a la izquierda, mapeo a la derecha.
    """

    configuration_changed = Signal()
    validation_updated = Signal(bool, str)

    def __init__(self, dataframe: pd.DataFrame, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.df = dataframe
        self.config: ExportSeparatedConfig | None = None
        self.validation_result: ValidationResult | None = None
        self._template_path: str = ""
        self._template_ext: str = ".xlsx"
        self._dest_path: Path | None = None

        self.setWindowTitle("Configurar Exportación Separada")
        self.setMinimumSize(900, 560)
        self.resize(1000, 600)

        self._setup_ui()
        self._setup_connections()
        self._initialize_data()

    # ── UI ──────────────────────────────────────────────────────────

    def _setup_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setSpacing(6)
        root.setContentsMargins(10, 8, 10, 8)

        splitter = QSplitter(Qt.Horizontal)  # type: ignore[attr-defined]

        # ── Panel izquierdo: Configuración ──
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 8, 0)
        left_layout.setSpacing(6)

        # Configuración principal
        cfg_group = QGroupBox("Configuración Principal")
        cfg_grid = QGridLayout(cfg_group)
        cfg_grid.setSpacing(4)

        cfg_grid.addWidget(QLabel("Columna:"), 0, 0)
        self.column_combo = QComboBox()
        cfg_grid.addWidget(self.column_combo, 0, 1)

        self.include_nulls_check = QCheckBox("Incluir nulos")
        cfg_grid.addWidget(self.include_nulls_check, 0, 2)

        cfg_grid.addWidget(QLabel("Plantilla:"), 1, 0)
        tpl_row = QHBoxLayout()
        self.template_path_label = QLabel("Ninguna")
        self.template_path_label.setStyleSheet("color: #6b7280; font-style: italic;")
        tpl_row.addWidget(self.template_path_label, 1)
        self.select_template_btn = QPushButton("Examinar…")
        self.select_template_btn.setFixedWidth(90)
        tpl_row.addWidget(self.select_template_btn)
        cfg_grid.addLayout(tpl_row, 1, 1, 1, 2)

        cfg_grid.addWidget(QLabel("Hoja / Celda:"), 2, 0)
        sheet_cell_row = QHBoxLayout()
        self.sheet_combo = QComboBox()
        self.sheet_combo.setEnabled(False)
        self.sheet_combo.setMinimumWidth(100)
        sheet_cell_row.addWidget(self.sheet_combo)
        self.start_cell_combo = QComboBox()
        self.start_cell_combo.setEditable(True)
        self.start_cell_combo.addItems(["A1", "A2", "A5", "B1", "B2"])
        self.start_cell_combo.setFixedWidth(80)
        sheet_cell_row.addWidget(self.start_cell_combo)
        sheet_cell_row.addStretch()
        cfg_grid.addLayout(sheet_cell_row, 2, 1, 1, 2)

        left_layout.addWidget(cfg_group)

        # Salida y archivos
        out_group = QGroupBox("Salida y Archivos")
        out_grid = QGridLayout(out_group)
        out_grid.setSpacing(4)

        out_grid.addWidget(QLabel("Destino:"), 0, 0)
        dest_row = QHBoxLayout()
        self.dest_folder_label = QLabel("Sin seleccionar")
        self.dest_folder_label.setStyleSheet("color: #6b7280; font-style: italic;")
        dest_row.addWidget(self.dest_folder_label, 1)
        self.select_folder_btn = QPushButton("Cambiar…")
        self.select_folder_btn.setFixedWidth(90)
        dest_row.addWidget(self.select_folder_btn)
        out_grid.addLayout(dest_row, 0, 1)

        out_grid.addWidget(QLabel("Nombre:"), 1, 0)
        name_row = QHBoxLayout()
        self.filename_template_edit = QLineEdit()
        self.filename_template_edit.setPlaceholderText("{valor}_{fecha}.xlsx")
        self.filename_template_edit.setText("{valor}_{fecha}.xlsx")
        name_row.addWidget(self.filename_template_edit, 1)
        name_row.addWidget(QLabel("({valor}, {fecha}, {contador})"))
        out_grid.addLayout(name_row, 1, 1)

        out_grid.addWidget(QLabel("Duplicados:"), 2, 0)
        self.handle_duplicates_combo = QComboBox()
        self.handle_duplicates_combo.addItems([
            "Sobrescribir", "Numerar automáticamente", "Ignorar"
        ])
        out_grid.addWidget(self.handle_duplicates_combo, 2, 1)

        left_layout.addWidget(out_group)

        # Preview lado a lado
        prev_group = QGroupBox("Previsualización")
        prev_layout = QHBoxLayout(prev_group)
        prev_layout.setSpacing(4)

        val_col = QVBoxLayout()
        val_col.addWidget(QLabel("Valores:"))
        self.values_preview = QListWidget()
        self.values_preview.setMaximumHeight(100)
        val_col.addWidget(self.values_preview)
        prev_layout.addLayout(val_col)

        name_col = QVBoxLayout()
        name_col.addWidget(QLabel("Nombres generados:"))
        self.filenames_preview = QListWidget()
        self.filenames_preview.setMaximumHeight(100)
        name_col.addWidget(self.filenames_preview)
        prev_layout.addLayout(name_col)

        left_layout.addWidget(prev_group)
        left_layout.addStretch()

        # ── Panel derecho: Mapeo ──
        right = QGroupBox("Mapeo de Columnas (DataFrame → Excel)")
        right_layout = QVBoxLayout(right)
        self.mapping_widget = ColumnMappingWidget()
        right_layout.addWidget(self.mapping_widget)

        # ── Ensamblar splitter ──
        splitter.addWidget(left)
        splitter.addWidget(right)
        splitter.setSizes([480, 520])
        root.addWidget(splitter, 1)

        # ── Barra de validación + botones ──
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setMaximumHeight(4)
        root.addWidget(self.progress_bar)

        bottom = QHBoxLayout()
        self.validation_label = QLabel("Configuración incompleta")
        self.validation_label.setStyleSheet("color: #d97706; font-weight: bold;")
        bottom.addWidget(self.validation_label)
        bottom.addStretch()

        self.preview_btn = QPushButton("Vista Previa")
        self.preview_btn.setEnabled(False)
        bottom.addWidget(self.preview_btn)

        self.cancel_btn = QPushButton("Cancelar")
        bottom.addWidget(self.cancel_btn)

        self.export_btn = QPushButton("Exportar")
        self.export_btn.setEnabled(False)
        self.export_btn.setStyleSheet(
            "QPushButton { background-color: #3b82f6; color: white; "
            "border: none; border-radius: 6px; padding: 6px 18px; font-weight: 600; }"
            "QPushButton:hover { background-color: #2563eb; }"
            "QPushButton:disabled { background-color: #94a3b8; }"
        )
        bottom.addWidget(self.export_btn)

        root.addLayout(bottom)

    # ── Conexiones ──────────────────────────────────────────────────

    def _setup_connections(self) -> None:
        self.column_combo.currentTextChanged.connect(self.on_column_changed)
        self.filename_template_edit.textChanged.connect(self.on_template_changed)
        self.start_cell_combo.currentTextChanged.connect(self.on_start_cell_changed)
        self.sheet_combo.currentTextChanged.connect(self.on_sheet_changed)
        self.select_template_btn.clicked.connect(self.select_template)
        self.select_folder_btn.clicked.connect(self.select_destination_folder)
        self.cancel_btn.clicked.connect(self.reject)
        self.export_btn.clicked.connect(self._on_export)
        self.preview_btn.clicked.connect(self.show_file_preview)

        self._validation_timer = QTimer()
        self._validation_timer.setSingleShot(True)
        self._validation_timer.setInterval(500)
        self._validation_timer.timeout.connect(self.validate_configuration)

    # ── Inicialización ──────────────────────────────────────────────

    def _initialize_data(self) -> None:
        if self.df is None or self.df.empty:
            return
        self.column_combo.addItems(self.df.columns.tolist())
        self.mapping_widget.set_dataframe_columns(self.df.columns.tolist())
        self.mapping_widget.set_excel_columns([chr(65 + i) for i in range(26)])
        self._load_settings()
        self.update_values_preview()
        self.validate_configuration()

    # ── Persistencia con QSettings ──────────────────────────────────

    def _save_settings(self) -> None:
        s = QSettings("FlashSheet", "ExportSeparated")
        s.beginGroup("Config")
        s.setValue("separator_column", self.column_combo.currentText())
        s.setValue("template_path", self._template_path)
        s.setValue("template_ext", self._template_ext)
        s.setValue("start_cell", self.start_cell_combo.currentText())
        s.setValue("file_template", self.filename_template_edit.text())
        s.setValue("handle_duplicates", self.handle_duplicates_combo.currentText())
        s.setValue("output_folder", str(self._dest_path) if self._dest_path else "")
        s.setValue("column_mapping", str(self.mapping_widget.get_mapping()))
        s.endGroup()

    def _load_settings(self) -> None:
        s = QSettings("FlashSheet", "ExportSeparated")
        s.beginGroup("Config")

        col = s.value("separator_column", "")
        if col:
            idx = self.column_combo.findText(col)
            if idx >= 0:
                self.column_combo.setCurrentIndex(idx)

        tpl_path = s.value("template_path", "")
        tpl_ext = s.value("template_ext", ".xlsx")
        if tpl_path and Path(tpl_path).exists():
            self._template_path = tpl_path
            self._template_ext = tpl_ext
            self.template_path_label.setText(Path(tpl_path).name)
            self.template_path_label.setStyleSheet("color: #111827;")
            self.template_path_label.setToolTip(tpl_path)
            try:
                wb = openpyxl.load_workbook(tpl_path, read_only=True)
                self.sheet_combo.blockSignals(True)
                self.sheet_combo.clear()
                self.sheet_combo.addItems(wb.sheetnames)
                self.sheet_combo.setEnabled(True)
                self.sheet_combo.blockSignals(False)
                wb.close()
            except Exception:
                pass

        cell = s.value("start_cell", "A1")
        idx = self.start_cell_combo.findText(cell)
        if idx >= 0:
            self.start_cell_combo.setCurrentIndex(idx)
        else:
            self.start_cell_combo.setEditText(cell)

        file_tpl = s.value("file_template", "")
        if file_tpl:
            self.filename_template_edit.setText(file_tpl)

        dup = s.value("handle_duplicates", "")
        if dup:
            idx = self.handle_duplicates_combo.findText(dup)
            if idx >= 0:
                self.handle_duplicates_combo.setCurrentIndex(idx)

        output = s.value("output_folder", "")
        if output and Path(output).is_dir():
            self._dest_path = Path(output)
            self.dest_folder_label.setText(self._dest_path.name)
            self.dest_folder_label.setToolTip(str(self._dest_path))
            self.dest_folder_label.setStyleSheet("color: #111827;")

        mapping_str = s.value("column_mapping", "")
        if mapping_str:
            try:
                import ast
                mapping = ast.literal_eval(mapping_str)
                if isinstance(mapping, dict):
                    for df_col, excel_col in mapping.items():
                        for row in range(self.mapping_widget.mapping_table.rowCount()):
                            item = self.mapping_widget.mapping_table.item(row, 0)
                            if item and item.text() == df_col:
                                combo = self.mapping_widget.mapping_table.cellWidget(row, 2)
                                if combo:
                                    idx = combo.findText(excel_col)
                                    if idx >= 0:
                                        combo.setCurrentIndex(idx)
                                break
            except Exception:
                pass

        s.endGroup()

    # ── Selección de plantilla ──────────────────────────────────────

    def select_template(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar Plantilla Excel", "",
            "Excel (*.xlsx *.xlsm);;Todos (*)"
        )
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo leer la plantilla:\n{e}")
            return

        self._template_path = path
        self._template_ext = Path(path).suffix.lower()
        name = Path(path).name
        self.template_path_label.setText(name)
        self.template_path_label.setStyleSheet("color: #111827;")
        self.template_path_label.setToolTip(path)

        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        self.sheet_combo.addItems(sheets)
        self.sheet_combo.setEnabled(True)
        self.sheet_combo.blockSignals(False)

        # Actualizar extensión en el filename template
        current = self.filename_template_edit.text().strip()
        if current.lower().endswith(('.xlsx', '.xlsm')):
            current = current[:current.rfind('.')] + self._template_ext
        else:
            current += self._template_ext
        self.filename_template_edit.blockSignals(True)
        self.filename_template_edit.setText(current)
        self.filename_template_edit.blockSignals(False)

        self._validation_timer.start()

    # ── Selección de destino ────────────────────────────────────────

    def select_destination_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "Seleccionar Carpeta de Destino")
        if not folder:
            return

        self._dest_path = Path(folder)
        self.dest_folder_label.setText(self._dest_path.name)
        self.dest_folder_label.setToolTip(str(self._dest_path))
        self.dest_folder_label.setStyleSheet("color: #111827;")

        self._validation_timer.start()

    # ── Cambios en campos ───────────────────────────────────────────

    def on_column_changed(self, _column: str) -> None:
        self.update_values_preview()
        self.update_filename_preview()
        self._validation_timer.start()

    def on_template_changed(self) -> None:
        self.update_filename_preview()
        self._validation_timer.start()

    def on_start_cell_changed(self, _text: str) -> None:
        self._validation_timer.start()

    def on_sheet_changed(self, _name: str) -> None:
        self._validation_timer.start()

    # ── Previews ────────────────────────────────────────────────────

    def update_values_preview(self) -> None:
        self.values_preview.clear()
        col = self.column_combo.currentText()
        if not col or self.df is None or col not in self.df.columns:
            return

        sample = self.df[col] if len(self.df) < 5000 else self.df[col].sample(5000)
        counts = sample.value_counts().head(15)
        for val, cnt in counts.items():
            if pd.notna(val):
                self.values_preview.addItem(f"  {val}  ({cnt:,})")

        nulls = self.df[col].isnull().sum()
        if nulls > 0:
            self.values_preview.addItem(f"  [Nulos]  ({nulls:,})")

    def update_filename_preview(self) -> None:
        self.filenames_preview.clear()
        col = self.column_combo.currentText()
        template = self.filename_template_edit.text().strip()
        if not col or not template or self.df is None or col not in self.df.columns:
            return

        if not template.lower().endswith(('.xlsx', '.xlsm')):
            template += self._template_ext

        for i, val in enumerate(self.df[col].dropna().unique()[:8]):
            fname = (template
                     .replace("{valor}", str(val))
                     .replace("{fecha}", datetime.now().strftime("%Y-%m-%d"))
                     .replace("{contador}", f"{i + 1:02d}")
                     .replace("{columna_nombre}", col))
            self.filenames_preview.addItem(f"  {fname}")

    # ── Validación ──────────────────────────────────────────────────

    def validate_configuration(self) -> None:
        config = self._build_config(validate=False)
        if config is None:
            self._set_validation(False, "Configuración incompleta")
            return

        try:
            splitter = ExcelTemplateSplitter(self.df, config)
            result = splitter.validate_configuration()
            self.validation_result = result

            if not result.is_valid:
                self._set_validation(False, f"{len(result.errors)} error(es)")
            elif result.warnings:
                self._set_validation(True, f"Válida ({len(result.warnings)} advertencia(s))")
            else:
                self._set_validation(True, "Configuración válida")
        except Exception as e:
            self._set_validation(False, str(e))

    def _set_validation(self, valid: bool, msg: str) -> None:
        color = "#16a34a" if valid else "#d97706"
        icon = "✅" if valid else "⚠️"
        self.validation_label.setText(f"{icon} {msg}")
        self.validation_label.setStyleSheet(f"color: {color}; font-weight: bold;")
        self.export_btn.setEnabled(valid)
        self.preview_btn.setEnabled(valid)
        self.validation_updated.emit(valid, msg)

    # ── Construcción de config ──────────────────────────────────────

    def _build_config(self, validate: bool = True) -> ExportSeparatedConfig | None:
        col = self.column_combo.currentText()
        if not col:
            return None

        template = self.filename_template_edit.text().strip()
        if not template:
            return None
        if not template.lower().endswith(('.xlsx', '.xlsm')):
            template += self._template_ext

        output = str(self._dest_path) if self._dest_path else ""
        if not output:
            return None

        start_cell = self.start_cell_combo.currentText().strip() or "A1"

        dup_map = {
            "Sobrescribir": "overwrite",
            "Numerar automáticamente": "append",
            "Ignorar": "skip",
        }

        config = ExportSeparatedConfig(
            separator_column=col,
            template_path=self._template_path,
            output_folder=output,
            start_cell=start_cell,
            file_template=template,
            column_mapping=self.mapping_widget.get_mapping(),
            handle_duplicates=dup_map.get(
                self.handle_duplicates_combo.currentText(), "overwrite"
            ),
        )

        if validate:
            try:
                splitter = ExcelTemplateSplitter(self.df, config)
                result = splitter.validate_configuration()
                if not result.is_valid:
                    return None
            except Exception:
                return None

        return config

    # ── Preview de archivos ─────────────────────────────────────────

    def show_file_preview(self) -> None:
        col = self.column_combo.currentText()
        if not col or self.df is None or col not in self.df.columns:
            return

        files_info: list[dict[str, Any]] = []
        for val in self.df[col].dropna().unique():
            group = self.df[self.df[col] == val]
            fname = (self.filename_template_edit.text().strip()
                     .replace("{valor}", str(val))
                     .replace("{fecha}", datetime.now().strftime("%Y-%m-%d")))
            if not fname.lower().endswith(('.xlsx', '.xlsm')):
                fname += self._template_ext
            files_info.append({
                'filename': fname,
                'group': str(val),
                'rows': len(group),
                'size': f"~{len(group) * 0.5:.1f}KB",
                'status': 'Listo'
            })

        FilePreviewDialog(files_info, self).exec()

    # ── Exportar ────────────────────────────────────────────────────

    def _on_export(self) -> None:
        config = self._build_config()
        if config is None:
            QMessageBox.warning(self, "Inválido", "Revise la configuración.")
            return
        self.config = config
        self.accept()

    # ── API pública ─────────────────────────────────────────────────

    def get_configuration(self, validate: bool = True) -> ExportSeparatedConfig | None:
        if self.config is not None and not validate:
            return self.config
        return self._build_config(validate=validate)

    def accept(self) -> None:
        if self.config is None:
            config = self._build_config()
            if config is None:
                QMessageBox.warning(
                    self, "Configuración Inválida",
                    "La configuración actual no es válida. Revise los errores marcados."
                )
                return
            self.config = config
        self._save_settings()
        super().accept()

    # ── Cleanup ─────────────────────────────────────────────────────

    def closeEvent(self, event: Any) -> None:
        self._validation_timer.stop()
        super().closeEvent(event)
